"""Shaxsiy kabinet — TezPOS backend ma'lumotlari."""
from __future__ import annotations

from io import BytesIO
import csv
import json
import logging
import threading
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError, as_completed
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from statistics import mean, pstdev
from types import SimpleNamespace

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST, require_http_methods

import os
import re

from . import tezpos_api
from .auth_views import (
    SESSION_DISPLAY,
    SESSION_SERVER,
    SESSION_TOKEN,
    clear_tezpos_session,
    session_has_tezpos,
)
from .models import DesktopInstaller, LabelTemplate, TenantProfile

logger = logging.getLogger("tezpos.slow")
tg_logger = logging.getLogger("tezpos.telegram")


def _log_slow(name: str, t0: float, extra: str = "") -> float:
    dt = time.time() - t0
    if dt >= 2.0:
        logger.warning("[SLOW] %s took %.2fs %s", name, dt, extra)
    return dt


PAYMENT_LABELS = {
    "cash": "Naqt",
    "card": "Karta",
    "mixed": "Aralash",
    "credit": "Qarz",
    "qarz": "Qarz",
    "debt": "Qarz",
    "nasiya": "Qarz",
    "click": "Click",
    "payme": "Payme",
    "transfer": "O‘tkazma",
}


def get_tenant_for_user(user):
    tenant, _ = TenantProfile.objects.get_or_create(
        user=user, defaults={"business_name": user.get_full_name() or user.username}
    )
    return tenant


_LABEL_EL_KEYS = (
    "name",
    "price",
    "old_price",
    "wholesale",
    "sku",
    "created",
    "custom1",
    "custom2",
    "custom3",
    "old_label",
    "print_date",
    "barcode",
    "logo",
)


def _label_shop_key(request) -> str:
    server = (request.session.get(SESSION_SERVER) or "").strip()
    if server:
        return server.lower()
    return f"user:{request.user.pk}"


def _label_num(value, default=0.0) -> float:
    try:
        n = float(value)
        if n != n:  # NaN
            return float(default)
        return n
    except (TypeError, ValueError):
        return float(default)


def _sanitize_label_template(body: dict) -> dict:
    if not isinstance(body, dict):
        return {}
    out = {}
    name = str(body.get("name") or "").strip()[:80]
    if name:
        out["name"] = name
    out["widthMm"] = max(10, min(200, _label_num(body.get("widthMm"), 38)))
    out["heightMm"] = max(10, min(200, _label_num(body.get("heightMm"), 58)))
    out["formatPrice"] = bool(body.get("formatPrice", True))
    out["priceSuffix"] = str(body.get("priceSuffix") or "")[:24]
    enabled_in = body.get("enabled") if isinstance(body.get("enabled"), dict) else {}
    out["enabled"] = {k: bool(enabled_in.get(k, False)) for k in _LABEL_EL_KEYS}
    styles_in = body.get("styles") if isinstance(body.get("styles"), dict) else {}
    styles = {}
    for key in _LABEL_EL_KEYS:
        st = styles_in.get(key) if isinstance(styles_in.get(key), dict) else {}
        align = str(st.get("align") or "center").strip().lower()[:16]
        if align not in ("left", "center", "right"):
            align = "center"
        styles[key] = {
            "x": _label_num(st.get("x"), 6),
            "y": _label_num(st.get("y"), 6),
            "w": _label_num(st.get("w"), 88),
            "h": _label_num(st.get("h"), 12),
            "size": _label_num(st.get("size"), 14),
            "weight": int(_label_num(st.get("weight"), 700)),
            "rotate": _label_num(st.get("rotate"), 0),
            "align": align,
            "text": str(st.get("text") or "")[:200],
        }
    out["styles"] = styles
    return out


def _label_template_payload(request) -> dict:
    row = LabelTemplate.objects.filter(shop_key=_label_shop_key(request)).first()
    data = row.data if row and isinstance(row.data, dict) else {}
    return data or {}


def _label_template_json(request) -> str:
    return json.dumps(_label_template_payload(request), ensure_ascii=False)


@require_GET
def download_installer(request):
    """Faol .exe ni yuklab beradi (Install tugmasi)."""
    installer = DesktopInstaller.get_active()
    if not installer or not installer.file:
        raise Http404("Installer hali yuklanmagan.")
    try:
        fh = installer.file.open("rb")
    except Exception as exc:
        raise Http404("Fayl topilmadi.") from exc

    filename = os.path.basename(installer.file.name) or "TezPOS-Setup.exe"
    if not filename.lower().endswith(".exe"):
        filename = f"{filename}.exe"
    response = FileResponse(fh, as_attachment=True, filename=filename)
    response["Content-Type"] = "application/octet-stream"
    return response


def _dec(value, default="0") -> Decimal:
    try:
        if value is None or value == "":
            return Decimal(default)
        return Decimal(str(value).replace(",", ".").replace(" ", "").replace("\u00a0", ""))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


def _first_dec(*values) -> Decimal:
    for value in values:
        if value is None or value == "":
            continue
        parsed = _dec(value)
        return parsed
    return Decimal("0")


def _parse_list_prices(raw) -> dict[str, Decimal]:
    """TezPOS list_prices: {id: narx} yoki [{price_list_id, price}]."""
    out: dict[str, Decimal] = {}
    if isinstance(raw, dict):
        for key, val in raw.items():
            if isinstance(val, dict):
                out[str(key)] = _first_dec(
                    val.get("price"), val.get("value"), val.get("amount")
                )
            else:
                out[str(key)] = _dec(val)
        return out
    if isinstance(raw, list):
        for row in raw:
            if not isinstance(row, dict):
                continue
            lid = (
                row.get("price_list_id")
                or row.get("list_id")
                or row.get("price_list")
                or row.get("id")
            )
            if isinstance(lid, dict):
                lid = lid.get("id")
            if not lid:
                continue
            out[str(lid)] = _first_dec(
                row.get("price"),
                row.get("value"),
                row.get("amount"),
                row.get("selling_price"),
            )
    return out
    try:
        if value is None or value == "":
            return Decimal(default)
        return Decimal(str(value).replace(",", ".").replace(" ", "").replace("\u00a0", ""))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


# TezPOS javoblarini qisqa muddat xotirada saqlash (SSR/AJAX tezligi)
_TEZPOS_MEMO: dict[str, tuple[float, object]] = {}


def _memo_get(key: str, ttl: float, loader, *, skip_empty: bool = False):
    now = time.time()
    hit = _TEZPOS_MEMO.get(key)
    if hit and now - hit[0] < ttl:
        if not (skip_empty and hit[1] in (None, [], {})):
            return hit[1]
    val = loader()
    if skip_empty and val in (None, [], {}):
        return hit[1] if hit else val
    _TEZPOS_MEMO[key] = (now, val)
    # Juda katta bo‘lib ketmasin
    if len(_TEZPOS_MEMO) > 96:
        oldest = sorted(_TEZPOS_MEMO.items(), key=lambda x: x[1][0])[:24]
        for k, _ in oldest:
            _TEZPOS_MEMO.pop(k, None)
    return val


def _memo_peek(key: str):
    hit = _TEZPOS_MEMO.get(key)
    if not hit:
        return None
    val = hit[1]
    if val in (None, [], {}):
        return None
    return val


def _products_payload_list(products: list, *, lite: bool = False) -> list[dict]:
    out = []
    for p in products:
        row = {
            "id": p.id,
            "name": p.name,
            "barcode": p.barcode or "",
            "barcodes": getattr(p, "barcode_list", []) or [],
            "unit": p.unit or "dona",
            "category": p.category or "",
            "brand": p.brand or "",
            "selling_price": float(p.selling_price),
            "wholesale_price": float(p.wholesale_price or 0),
            "cost_price": float(p.cost_price or 0),
            "stock_qty": float(p.stock_qty or 0),
            "min_stock": float(p.min_stock or 0),
            "is_favorite": bool(p.is_favorite),
        }
        if not lite:
            images = []
            try:
                if hasattr(p, "images") and hasattr(p.images, "all"):
                    images = [
                        {"id": img.id, "url": img.image.url, "is_primary": img.is_primary}
                        for img in p.images.all()
                    ]
            except Exception:
                images = []
            row["list_prices"] = {
                str(k): float(v) for k, v in (p.list_prices or {}).items()
            }
            row["image"] = p.display_image
            row["image_url"] = getattr(p, "image_url", "") or ""
            row["images"] = images
        out.append(row)
    return out


@login_required
@require_GET
def cabinet_api_status(request):
    """Kabinet: TezPOS API ulanishini tekshirish (diagnostika)."""
    if not session_has_tezpos(request):
        return JsonResponse({"ok": False, "error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    base = tezpos_api.normalize_api_base()
    t0 = time.time()
    try:
        # Engil so‘rov — price-lists yoki products 1 page
        tezpos_api.api_request(
            "GET",
            "/api/catalog/price-lists/",
            token=token,
            server_name=server,
            timeout=6,
        )
        ms = int((time.time() - t0) * 1000)
        return JsonResponse(
            {
                "ok": True,
                "api": base,
                "server": server,
                "ms": ms,
            }
        )
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"ok": False, "error": "auth", "api": base}, status=401)
        return JsonResponse(
            {
                "ok": False,
                "api": base,
                "error": str(exc),
                "ms": int((time.time() - t0) * 1000),
            },
            status=502,
        )
    except (TimeoutError, OSError) as exc:
        return JsonResponse(
            {
                "ok": False,
                "api": base,
                "error": f"Timeout / ulanish: {exc}",
                "ms": int((time.time() - t0) * 1000),
            },
            status=504,
        )


@login_required
@require_GET
def cabinet_catalog(request):
    """Mahsulotlar AJAX. ?page=N — tez sahifa; brauzer hammasi yig‘adi."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    skip_pl = (request.GET.get("skip_pl") or "").strip() in ("1", "true", "yes")
    want_full = (request.GET.get("full") or "").strip() in ("1", "true", "yes")

    def _price_lists_payload():
        if skip_pl:
            return []
        try:
            raw_pl = _memo_get(
                f"{memo_prefix}|price_lists",
                600.0,
                lambda: tezpos_api.get_price_lists(token, server) or [],
            ) or []
            return [
                {
                    "id": str(pl.get("id") or ""),
                    "name": (pl.get("name") or "").strip() or "Narxlar",
                    "is_selling": bool(pl.get("is_selling")),
                }
                for pl in raw_pl
                if isinstance(pl, dict) and pl.get("is_active", True) and str(pl.get("id") or "")
            ]
        except Exception:
            return []

    page_raw = (request.GET.get("page") or "").strip()
    # full=1 butun katalogni bir requestga yig‘ib worker ni 504 qiladi — sahifalab yuboriladi
    if want_full and not page_raw.isdigit():
        page_raw = "1"
    if page_raw.isdigit():
        page_n = max(1, int(page_raw))
        try:
            page_size = max(20, min(int(request.GET.get("page_size") or 100), 100))
        except (TypeError, ValueError):
            page_size = 100
        t0 = time.time()
        try:
            pack = tezpos_api.get_products_page(
                token,
                server,
                page=page_n,
                page_size=page_size,
                timeout=25 if page_n == 1 else 12,
                retries=2,
            )
        except tezpos_api.TezPosApiError as exc:
            if getattr(exc, "status", None) in (401, 403):
                clear_tezpos_session(request)
                return JsonResponse({"error": "auth"}, status=401)
            return JsonResponse({"ok": False, "error": str(exc), "products": []}, status=200)
        except (TimeoutError, OSError) as exc:
            return JsonResponse({"ok": False, "error": str(exc), "products": []}, status=200)

        products = []
        for row in pack.get("rows") or []:
            if not isinstance(row, dict):
                continue
            try:
                products.append(_map_product(row))
            except Exception:
                continue
        price_lists_payload = []
        if not skip_pl:
            try:
                raw_pl = _memo_get(
                    f"{memo_prefix}|price_lists",
                    600.0,
                    lambda: tezpos_api.get_price_lists(token, server) or [],
                ) or []
                price_lists_payload = [
                    {
                        "id": str(pl.get("id") or ""),
                        "name": (pl.get("name") or "").strip() or "Narxlar",
                        "is_selling": bool(pl.get("is_selling")),
                    }
                    for pl in raw_pl
                    if isinstance(pl, dict) and pl.get("is_active", True) and str(pl.get("id") or "")
                ]
            except Exception:
                price_lists_payload = []
        actual = len(products)
        requested = page_size
        total = int(pack.get("total") or 0)
        api_page_size = int(pack.get("page_size") or actual or page_size)
        loaded = (page_n - 1) * max(api_page_size, 1) + actual
        catalog_count = int(pack.get("catalog_count") or 0)
        if page_n == 1 and not catalog_count:
            try:
                catalog_count = int(
                    _memo_get(
                        f"{memo_prefix}|product_count",
                        120.0,
                        lambda: tezpos_api.get_product_count(token, server) or 0,
                    )
                    or 0
                )
            except Exception:
                catalog_count = 0
        if actual >= api_page_size and total and total <= loaded and total % 50 == 0:
            if not catalog_count or catalog_count <= loaded:
                total = 0
        if catalog_count > total:
            total = catalog_count
        has_more = tezpos_api.catalog_has_more(
            actual=actual,
            requested=requested,
            has_next=bool(pack.get("has_more")),
            total=total or catalog_count,
            loaded=loaded,
        )
        if not products:
            has_more = False
        return JsonResponse(
            {
                "ok": True,
                "products": _products_payload_list(products, lite=skip_pl),
                "priceLists": price_lists_payload,
                "page": page_n,
                "page_size": actual if actual else api_page_size,
                "total": total,
                "catalog_count": catalog_count or total,
                "has_more": has_more,
                "count": actual,
                "complete": not has_more,
                "partial": has_more,
                "api": tezpos_api.normalize_api_base(),
            }
        )

    # page yo‘q: birinchi sahifa (to‘liq dump gunicorn ni 504 qiladi)
    t0 = time.time()
    try:
        pack = tezpos_api.get_products_page(token, server, page=1, page_size=100, timeout=12, retries=2)
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc), "products": []}, status=200)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc), "products": []}, status=200)
    products = []
    for row in pack.get("rows") or []:
        if not isinstance(row, dict):
            continue
        try:
            products.append(_map_product(row))
        except Exception:
            continue
    _log_slow("catalog", t0, f"n={len(products)}")
    return JsonResponse(
        {
            "ok": True,
            "products": _products_payload_list(products, lite=True),
            "priceLists": [],
            "page": 1,
            "page_size": len(products) or 100,
            "total": int(pack.get("total") or 0),
            "has_more": True,
            "count": len(products),
            "complete": False,
            "partial": True,
            "api": tezpos_api.normalize_api_base(),
        }
    )


@login_required
@require_GET
def cabinet_warm(request):
    """API memo ni fonida isitish — navigatsiya kutmasin."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    today = timezone.localdate().isoformat()

    def _warm() -> None:
        # Fon isitish — web requestni bloklamaydi. Optom uchun list_prices kerak.
        try:
            _memo_get(
                f"{memo_prefix}|price_lists",
                600.0,
                lambda: tezpos_api.get_price_lists(token, server) or [],
            )
        except Exception:
            pass
        try:
            _memo_get(
                f"{memo_prefix}|products|4",
                600.0,
                lambda: tezpos_api.get_products(token, server, max_pages=4, timeout=12) or [],
            )
        except Exception:
            pass
        try:
            _memo_get(
                f"{memo_prefix}|dayv3|{today}",
                120.0,
                lambda: tezpos_api.get_sales_for_day(token, server, today),
            )
        except Exception:
            pass

    threading.Thread(target=_warm, daemon=True, name="tezpos-warm").start()
    return JsonResponse({"ok": True, "warming": True})


@login_required
@require_GET
def cabinet_day_sales(request):
    """Kunlik sotuvlar ro‘yxati (AJAX) — sahifa darhol ochilsin."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    sale_date = _parse_sale_date(request.GET.get("sale_date"))
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    try:
        day_sales_raw = _memo_get(
            f"{memo_prefix}|dayv4|{sale_date.isoformat()}",
            90.0,
            lambda: tezpos_api.get_sales_for_day(token, server, sale_date.isoformat()),
        )
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"error": str(exc)}, status=502)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"error": str(exc)}, status=504)

    day_sales_raw = [s for s in (day_sales_raw or []) if isinstance(s, dict)]
    day_sales_raw.sort(
        key=lambda s: _parse_dt(s.get("completed_at") or s.get("created_at"))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    products_raw = []
    try:
        products_raw = _memo_get(
            f"{memo_prefix}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=14) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError, Exception):
        products_raw = []
    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    cashier = _cashier_name(request)

    need_fetch = [
        str(s.get("id"))
        for s in day_sales_raw
        if s.get("id")
        and (
            not _sale_items(s)
            or not _display_receipt_number(s)
            or _sale_level_cost_profit(s, _dec(s.get("total"))) is None
        )
    ]
    details = {}
    if need_fetch:
        details = _fetch_sale_details(
            token,
            server,
            need_fetch,
            limit=min(len(need_fetch), 250),
            per_sale_timeout=2.2,
            overall_timeout=40.0,
        )

    payload = []
    day_gross = Decimal("0")
    day_cost = Decimal("0")
    day_profit = Decimal("0")
    for s in day_sales_raw:
        sid = str(s.get("id") or "")
        detail = details.get(sid) if sid else None
        if isinstance(detail, dict):
            merged = {**s, **detail}
            if not _sale_items(merged) and _sale_items(s):
                merged = {**detail, **s}
        else:
            merged = s
        total = _dec(merged.get("total") or s.get("total"))
        day_gross += total
        row = _serialize_sale_payload(
            merged, cashier, products_by_id, products_by_name
        )
        # Ro‘yxatda chek raqami bo‘lmasa — asosiy sale dan
        if not _display_receipt_number({"receipt_number": row.get("receipt_number")}):
            rn = _display_receipt_number(s, merged)
            if rn:
                row["receipt_number"] = rn
                row["receipt_no"] = rn
        day_cost += _dec(row.get("total_cost"))
        day_profit += _dec(row.get("profit"))
        payload.append(row)

    return JsonResponse(
        {
            "ok": True,
            "sale_date": sale_date.isoformat(),
            "count": len(payload),
            "gross": float(day_gross),
            "cost": float(day_cost),
            "profit": float(day_profit),
            "sales": payload,
        }
    )


@login_required
@require_GET
def cabinet_sale_detail(request):
    """Bitta chek — tovarlar va summalar (drawer uchun AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    sale_id = (request.GET.get("id") or "").strip()
    if not sale_id:
        return JsonResponse({"ok": False, "error": "id kerak"}, status=400)

    memo_prefix = f"{server}|{(token or '')[-12:]}"
    try:
        detail = _memo_get(
            f"{memo_prefix}|sale|{sale_id}",
            45.0,
            lambda: tezpos_api.get_sale(token, server, sale_id),
        )
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc)}, status=502)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=504)

    if not isinstance(detail, dict):
        return JsonResponse({"ok": False, "error": "Chek topilmadi"}, status=404)

    products_raw = []
    try:
        products_raw = _memo_get(
            f"{memo_prefix}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=14) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError, Exception):
        products_raw = []
    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    sale = _serialize_sale_payload(
        detail,
        _cashier_name(request),
        products_by_id,
        products_by_name,
    )
    return JsonResponse({"ok": True, "sale": sale})


@login_required
@require_GET
def cabinet_day_sales_export(request):
    """Kunlik sotuvlar — Excel/CSV (FAST_SHELL dan tashqari)."""
    if not session_has_tezpos(request):
        return redirect("login")
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    sale_date = _parse_sale_date(request.GET.get("sale_date"))

    try:
        day_sales_raw = tezpos_api.get_sales_for_day(token, server, sale_date.isoformat())
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return redirect("login")
        return HttpResponse(str(exc), status=502, content_type="text/plain; charset=utf-8")
    except (TimeoutError, OSError) as exc:
        return HttpResponse(str(exc), status=504, content_type="text/plain; charset=utf-8")

    day_sales_raw = [s for s in (day_sales_raw or []) if isinstance(s, dict)]
    day_sales_raw.sort(
        key=lambda s: _parse_dt(s.get("completed_at") or s.get("created_at"))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    memo_prefix = f"{server}|{(token or '')[-12:]}"
    products_raw = []
    try:
        products_raw = _memo_get(
            f"{memo_prefix}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=14) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError, Exception):
        products_raw = []
    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    cashier = _cashier_name(request)

    need_fetch = [
        str(s.get("id"))
        for s in day_sales_raw
        if s.get("id") and not _sale_items(s)
    ]
    details = _fetch_sale_details(
        token,
        server,
        need_fetch,
        limit=len(need_fetch) or 1,
        per_sale_timeout=4.0,
        overall_timeout=90.0,
    )

    day_sales_payload = []
    for s in day_sales_raw:
        sid = str(s.get("id") or "")
        detail = details.get(sid) or s
        if not _sale_items(detail):
            detail = {**s, "items": []}
        day_sales_payload.append(
            _serialize_sale_payload(
                detail,
                cashier,
                products_by_id,
                products_by_name,
            )
        )
    return _export_daily_sales_excel(day_sales_payload, sale_date, cashier)


def _collect_shifts_payload(token: str, server: str, *, days: int = 90) -> tuple[list[dict], str]:
    """Smenalar ro‘yxati — desktop /api/auth/shift bilan mos."""
    today = timezone.localdate()
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    date_from = (today - timedelta(days=days)).isoformat()
    date_to = today.isoformat()

    raw_shifts_api: list = []
    current_raw = None
    try:
        current_raw = tezpos_api.get_current_shift(token, server)
    except (tezpos_api.TezPosApiError, TimeoutError, OSError):
        current_raw = None
    try:
        raw_shifts_api = _memo_get(
            f"{memo_prefix}|shiftsv2|{date_from}|{date_to}",
            60.0,
            lambda: tezpos_api.get_shifts(
                token,
                server,
                date_from=date_from,
                date_to=date_to,
                timeout=12,
                max_pages=20,
            )
            or [],
        )
    except (tezpos_api.TezPosApiError, TimeoutError, OSError):
        raw_shifts_api = []

    margin_ratio = Decimal("0")
    shifts_payload: list[dict] = []
    shifts_source = "none"
    seen_ids: set[str] = set()

    def _push(raw: dict) -> None:
        if not isinstance(raw, dict):
            return
        sh = _normalize_api_shift(raw, today)
        sid = str(sh.get("id") or "")
        if sid and sid in seen_ids:
            return
        if sid:
            seen_ids.add(sid)
        sh.pop("raw", None)
        shifts_payload.append(sh)

    if current_raw:
        _push(current_raw)
    if raw_shifts_api:
        shifts_source = "api"
        for raw in raw_shifts_api:
            _push(raw)
    elif current_raw:
        shifts_source = "api"

    open_id = ""
    if current_raw:
        open_id = str(current_raw.get("id") or current_raw.get("uuid") or "").strip()
    for sh in shifts_payload:
        sid = str(sh.get("id") or "")
        if sh.get("status") == "open" and (not open_id or sid != open_id):
            sh["status"] = "closed"
            sh["status_label"] = "Yopilgan"
            if not sh.get("closed_at"):
                sh["closed_display"] = sh.get("opened_display") or "—"

    if not shifts_payload:
        shifts_source = "sales"

    sales_for_shifts: list = []
    try:
        sales_for_shifts = _memo_get(
            f"{memo_prefix}|salesv3|{date_from}|{date_to}|shifts",
            60.0,
            lambda: tezpos_api.get_sales(
                token,
                server,
                date_from=date_from,
                date_to=date_to,
                timeout=16,
                max_pages=50,
            )
            or [],
        )
    except Exception:
        sales_for_shifts = []
    sales_for_shifts = [s for s in (sales_for_shifts or []) if isinstance(s, dict)]

    if not shifts_payload:
        shifts_payload = _build_shifts_from_sales(
            sales_for_shifts,
            today=today,
            margin_ratio=margin_ratio,
        )
    elif sales_for_shifts:
        for sh in shifts_payload:
            _enrich_shift_with_sales(sh, sales_for_shifts, margin_ratio=margin_ratio)

    shifts_payload.sort(key=lambda s: s.get("opened_at") or "", reverse=True)
    return shifts_payload[:120], shifts_source


@login_required
@require_GET
def cabinet_shifts(request):
    """Smenalar ro‘yxati (AJAX) — sahifa darhol ochilsin."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    t0 = time.time()
    try:
        shifts, source = _collect_shifts_payload(token, server, days=90)
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc), "shifts": []}, status=200)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc), "shifts": []}, status=200)

    _log_slow("shifts", t0, f"n={len(shifts)} src={source}")
    return JsonResponse(
        {
            "ok": True,
            "shifts": shifts,
            "source": source,
            "count": len(shifts),
        }
    )


@login_required
@require_GET
def cabinet_reports(request):
    """Hisobotlar grafigi (AJAX) — sahifa kutmasin."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    today = timezone.localdate()
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    start = today - timedelta(days=6)
    t0 = time.time()
    try:
        sales = _memo_get(
            f"{memo_prefix}|salesv3|{start.isoformat()}|{today.isoformat()}|40",
            90.0,
            lambda: tezpos_api.get_sales(
                token,
                server,
                date_from=start.isoformat(),
                date_to=today.isoformat(),
                timeout=16,
                max_pages=40,
            )
            or [],
        )
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc), "reports": {}}, status=200)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc), "reports": {}}, status=200)

    sales = [s for s in (sales or []) if isinstance(s, dict)]
    charts = _build_charts_from_sales(sales, today)
    d_pack = charts.get("d7") or {"labels": [], "totals": [], "counts": []}
    daily = _chart_pack_for_dates(
        [s for s in sales if _sale_day(s) == today], today, today
    )
    weekly = _chart_pack_for_dates(sales, today - timedelta(days=6), today)
    monthly = charts.get("m6") or charts.get("m3") or d_pack

    margin_ratio = Decimal("0.25")

    gross = float(sum((_dec(s.get("total")) for s in sales), Decimal("0")))
    checks = len(sales)
    profit = gross * float(margin_ratio)
    today_sales = [s for s in sales if _sale_day(s) == today]
    today_gross = float(sum((_dec(s.get("total")) for s in today_sales), Decimal("0")))
    today_profit = today_gross * float(margin_ratio)

    _log_slow("reports", t0, f"n={len(sales)}")
    return JsonResponse(
        {
            "ok": True,
            "reports": {
                "daily": daily,
                "weekly": weekly,
                "monthly": {
                    "labels": monthly.get("labels") or [],
                    "totals": monthly.get("totals") or [],
                    "counts": monthly.get("counts") or [],
                },
            },
            "summary": {
                "checks": checks,
                "gross": gross,
                "cost": gross - profit,
                "profit": profit,
                "margin": (profit / gross * 100.0) if gross else 0.0,
                "today_profit": today_profit,
                "today_count": len(today_sales),
                "today_gross": today_gross,
            },
        }
    )


@login_required
@require_GET
def cabinet_abc(request):
    """ABC-XYZ (AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    today = timezone.localdate()
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    start = today - timedelta(days=14)
    t0 = time.time()
    try:
        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_p = pool.submit(
                lambda: _memo_get(
                    f"{memo_prefix}|products|1",
                    600.0,
                    lambda: tezpos_api.get_products(token, server, max_pages=1, timeout=10) or [],
                )
            )
            fut_s = pool.submit(
                lambda: _memo_get(
                    f"{memo_prefix}|salesv3|{start.isoformat()}|{today.isoformat()}|40",
                    90.0,
                    lambda: tezpos_api.get_sales(
                        token,
                        server,
                        date_from=start.isoformat(),
                        date_to=today.isoformat(),
                        timeout=16,
                        max_pages=40,
                    )
                    or [],
                )
            )
            products_raw = fut_p.result(timeout=18) or []
            sales = fut_s.result(timeout=20) or []
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc), "rows": []}, status=200)
    except (TimeoutError, OSError, FuturesTimeoutError) as exc:
        return JsonResponse({"ok": False, "error": str(exc), "rows": []}, status=200)

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    sales = [s for s in sales if isinstance(s, dict)]
    item_rows = []
    sale_ids = [str(s.get("id")) for s in sales if s.get("id")][:12]
    details = _fetch_sale_details(
        token,
        server,
        sale_ids,
        limit=12,
        per_sale_timeout=2.5,
        overall_timeout=6.0,
    )
    for detail in details.values():
        for item in detail.get("items") or []:
            if not isinstance(item, dict):
                continue
            pid = str(item.get("product_id") or item.get("product") or "")
            if not pid or pid.startswith("{"):
                name = (item.get("product_name") or item.get("name") or "").strip()
                if not name:
                    continue
                pid = f"name:{name.casefold()}"
            qty = _dec(item.get("quantity"))
            unit_price = _dec(item.get("unit_price"))
            item_rows.append(
                {
                    "product_id": pid,
                    "quantity": qty,
                    "unit_price": unit_price,
                    "day": _sale_day(detail) or today,
                }
            )
    abc_rows, abc_matrix, abc_total = _build_abc_xyz(products, item_rows, today)
    rows_out = []
    for row in abc_rows[:200]:
        p = row.get("product")
        rows_out.append(
            {
                "name": getattr(p, "name", None) or row.get("name") or "—",
                "abc": row.get("abc"),
                "xyz": row.get("xyz"),
                "group": row.get("group"),
                "revenue": float(row.get("revenue") or 0),
                "share": float(row.get("share") or 0),
                "stock": float(getattr(p, "stock_qty", 0) or 0),
            }
        )
    _log_slow("abc", t0, f"rows={len(rows_out)}")
    return JsonResponse(
        {
            "ok": True,
            "rows": rows_out,
            "matrix": abc_matrix,
            "total": float(abc_total or 0),
        }
    )


def _parse_sale_date(raw):
    if not raw:
        return timezone.localdate()
    try:
        return date.fromisoformat(str(raw)[:10])
    except ValueError:
        return timezone.localdate()


def _cashier_name(request) -> str:
    display = (request.session.get(SESSION_DISPLAY) or "").strip()
    if display:
        return display
    user = request.user
    return (user.get_full_name() or "").strip() or user.username or "Kassir"


def _payment_label(method):
    return PAYMENT_LABELS.get((method or "").lower(), method or "Naqt")


def _sale_items(detail: dict | None) -> list[dict]:
    if not isinstance(detail, dict):
        return []
    for key in ("items", "lines", "details", "sale_items", "products"):
        rows = detail.get(key)
        if isinstance(rows, list) and rows:
            return [x for x in rows if isinstance(x, dict)]
    return []


def _receipt_number(*sources: dict | None) -> str:
    keys = (
        "receipt_number",
        "receipt_no",
        "check_number",
        "check_no",
        "checkNumber",
    )
    for src in sources:
        if not isinstance(src, dict):
            continue
        for key in keys:
            val = src.get(key)
            if val not in (None, ""):
                return str(val).strip()
    for src in sources:
        if isinstance(src, dict) and src.get("id") not in (None, ""):
            return str(src.get("id"))
    return ""


def _display_receipt_number(*sources: dict | None) -> str:
    """TezPOS dasturidagi chek raqami (UUID emas)."""
    keys = (
        "receipt_number",
        "receipt_no",
        "check_number",
        "check_no",
        "checkNumber",
    )
    for src in sources:
        if not isinstance(src, dict):
            continue
        for key in keys:
            val = src.get(key)
            if val in (None, ""):
                continue
            text = str(val).strip()
            # UUID o‘xshash qiymatlarni rad etamiz
            if len(text) >= 32 and text.count("-") >= 4:
                continue
            return text
    return ""


def _sale_level_cost_profit(sale: dict, total: Decimal) -> tuple[Decimal, Decimal] | None:
    """API chekda tayyor tannarx/foyda bo‘lsa — shundan foydalanamiz."""
    cost = _dec(
        sale.get("total_cost")
        or sale.get("cost_total")
        or sale.get("purchase_total")
        or sale.get("cogs")
        or sale.get("cost_amount")
    )
    profit = _dec(
        sale.get("profit")
        or sale.get("net_profit")
        or sale.get("gross_profit")
    )
    if cost > 0:
        if profit == 0 and total > 0:
            profit = (total - cost).quantize(Decimal("0.01"))
        return cost.quantize(Decimal("0.01")), profit.quantize(Decimal("0.01"))
    if profit != 0 and total > 0:
        cost = (total - profit).quantize(Decimal("0.01"))
        if cost >= 0:
            return cost, profit.quantize(Decimal("0.01"))
    return None


def _item_qty(item: dict) -> Decimal:
    return _dec(item.get("quantity") or item.get("qty") or item.get("count"))


def _item_unit_price(item: dict) -> Decimal:
    return _dec(
        item.get("unit_price")
        or item.get("price")
        or item.get("selling_price")
        or item.get("sale_price")
    )


def _item_unit(item: dict, product=None) -> str:
    unit = (
        item.get("unit")
        or item.get("unit_name")
        or (getattr(product, "unit", None) if product else None)
        or "dona"
    )
    return str(unit).strip() or "dona"


def _sale_products_text(
    items: list[dict],
    products_by_id: dict,
    products_by_name: dict | None = None,
) -> str:
    from . import telegram_bot as tg

    lines: list[str] = []
    for item in items:
        qty = float(_item_qty(item))
        if qty <= 0:
            continue
        pid = str(item.get("product_id") or item.get("product") or "").strip()
        name = (item.get("product_name") or item.get("name") or "").strip()
        p = _find_product(
            products_by_id,
            products_by_name,
            product_id=pid,
            product_name=name,
        )
        if not name and p:
            name = p.name
        unit_price = float(_item_unit_price(item))
        line_total = float(_dec(item.get("total") or item.get("line_total"), str(qty * unit_price)))
        lines.append(
            tg.format_sale_product_line(
                name or "Mahsulot",
                qty,
                _item_unit(item, p),
                unit_price,
                line_total,
            )
        )
    return "\n".join(lines)


def _parse_dt(raw) -> datetime | None:
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    dt = parse_datetime(str(raw).replace("Z", "+00:00"))
    if dt is None:
        try:
            dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
        except ValueError:
            return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _sale_day(sale: dict) -> date | None:
    dt = _parse_dt(sale.get("completed_at") or sale.get("created_at"))
    if not dt:
        return None
    return timezone.localtime(dt).date()


def _rel_image_url(url: str | None) -> str:
    if not url:
        return ""
    url = str(url)
    if url.startswith("http://") or url.startswith("https://"):
        return url
    base = tezpos_api.normalize_api_base()
    if url.startswith("/"):
        return f"{base}{url}"
    return f"{base}/{url}"


def _barcode_token(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        return str(
            value.get("barcode")
            or value.get("code")
            or value.get("value")
            or value.get("ean")
            or value.get("barcode_value")
            or ""
        ).strip()
    text = str(value).strip()
    if text.lower() in ("none", "null"):
        return ""
    return text


def collect_product_barcodes(*sources) -> list[str]:
    """Mahsulotdagi barcha shtrixkodlar (30–40 ta ham). Tartib saqlanadi."""
    out: list[str] = []
    seen: set[str] = set()

    def _add(code: str) -> None:
        if not code:
            return
        if any(sep in code for sep in (",", "\n", "\r", ";", "|")):
            for piece in (
                code.replace("\r\n", "\n")
                .replace("\r", "\n")
                .replace(";", ",")
                .replace("|", ",")
                .replace("\n", ",")
                .split(",")
            ):
                part = piece.strip()
                if part and part not in seen:
                    seen.add(part)
                    out.append(part)
            return
        if code in seen:
            return
        seen.add(code)
        out.append(code)

    for src in sources:
        if src is None or src is False:
            continue
        if isinstance(src, (list, tuple, set)):
            for item in src:
                _add(_barcode_token(item))
        else:
            _add(_barcode_token(src))
    return out


def format_barcodes_excel_cell(codes) -> str:
    """Excel katak: har kod o‘z qatorida, oxirida vergul (Wrap Text)."""
    rows = collect_product_barcodes(codes)
    if not rows:
        return ""
    return "\n".join(f"{code}," for code in rows)


def parse_barcodes_cell(value) -> list[str]:
    """Shablon: kod,\\nkod,  — vergul/qator/nuqtali vergul."""
    if value is None or value is False:
        return []
    if isinstance(value, (list, tuple, set)):
        return collect_product_barcodes(value)
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    chunks: list[str] = []
    for line in text.split("\n"):
        for piece in (
            line.replace(";", ",").replace("|", ",").replace("\u00a0", " ").split(",")
        ):
            chunks.append(piece.strip())
    return collect_product_barcodes(chunks)


def _map_product(raw: dict) -> SimpleNamespace:
    codes = collect_product_barcodes(
        raw.get("barcode"),
        raw.get("barcodes"),
        raw.get("barcode_list"),
        raw.get("barcode_codes"),
        raw.get("extra_barcodes"),
        raw.get("additional_barcodes"),
        raw.get("codes"),
    )
    barcode = codes[0] if codes else ""

    list_prices = _parse_list_prices(
        raw.get("list_prices") or raw.get("price_lists") or raw.get("prices")
    )
    wholesale = Decimal("0")
    if list_prices:
        try:
            wholesale = min(list_prices.values())
        except ValueError:
            wholesale = Decimal("0")

    image_url = _rel_image_url(raw.get("image_url") or raw.get("image"))
    images_raw = raw.get("images") or []
    images = []
    for img in images_raw:
        if not isinstance(img, dict):
            continue
        images.append(
            SimpleNamespace(
                id=img.get("id"),
                image=SimpleNamespace(url=_rel_image_url(img.get("url"))),
                is_primary=bool(img.get("is_primary")),
            )
        )

    stock = _first_dec(
        raw.get("quantity"),
        raw.get("stock_qty"),
        raw.get("stock"),
        raw.get("qty"),
    )
    selling = _first_dec(
        raw.get("price"),
        raw.get("selling_price"),
        raw.get("sale_price"),
        raw.get("sell_price"),
    )
    cost = _first_dec(
        raw.get("cost_price"),
        raw.get("purchase_price"),
        raw.get("buy_price"),
        raw.get("cost"),
    )
    min_stock = _dec(raw.get("min_stock"), "0")
    category = (raw.get("category_name") or "").strip()
    brand = (raw.get("brand_name") or "").strip()
    name = (raw.get("name") or "").strip() or "Mahsulot"
    pid = str(raw.get("id") or "")
    if cost > 0:
        margin_pct = float((selling - cost) / cost * 100)
    else:
        margin_pct = 0.0

    p = SimpleNamespace(
        id=pid,
        name=name,
        sku=(raw.get("sku") or "").strip(),
        barcode=barcode or (codes[0] if codes else ""),
        barcodes=codes,
        barcode_list=codes,
        unit=(raw.get("unit") or "dona").strip() or "dona",
        category=category,
        brand=brand,
        selling_price=selling,
        wholesale_price=wholesale,
        list_prices=list_prices,
        cost_price=cost,
        stock_qty=stock,
        min_stock=min_stock,
        margin_pct=margin_pct,
        is_favorite=bool(raw.get("is_favorite")),
        is_active=bool(raw.get("is_active", True)),
        image_url=image_url,
        display_image=image_url,
        is_low_stock=stock <= min_stock,
        images=SimpleNamespace(
            all=lambda: images,
            exists=lambda: bool(images),
            filter=lambda **kw: SimpleNamespace(
                first=lambda: next(
                    (i for i in images if all(getattr(i, k) == v for k, v in kw.items())),
                    None,
                )
            ),
            first=lambda: images[0] if images else None,
        ),
        _raw=raw,
    )
    return p


def _empty_form(errors=None):
    return SimpleNamespace(errors=errors or {})


def _build_charts_from_sales(sales: list[dict], today: date):
    def bucket_daily(start: date, days: int, fmt="%d.%m"):
        totals_map = defaultdict(float)
        counts_map = defaultdict(int)
        for s in sales:
            d = _sale_day(s)
            if not d or d < start or d > start + timedelta(days=days - 1):
                continue
            totals_map[d] += float(_dec(s.get("total")))
            counts_map[d] += 1
        labels, totals, counts = [], [], []
        for i in range(days):
            day = start + timedelta(days=i)
            labels.append(day.strftime(fmt))
            totals.append(totals_map.get(day, 0.0))
            counts.append(counts_map.get(day, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    def hourly_today():
        totals_map = defaultdict(float)
        counts_map = defaultdict(int)
        for s in sales:
            if _sale_day(s) != today:
                continue
            dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
            if not dt:
                continue
            hour = timezone.localtime(dt).hour
            totals_map[hour] += float(_dec(s.get("total")))
            counts_map[hour] += 1
        labels, totals, counts = [], [], []
        for hour in range(24):
            labels.append(f"{hour:02d}:00")
            totals.append(totals_map.get(hour, 0.0))
            counts.append(counts_map.get(hour, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    def monthly(months: int):
        y, m = today.year, today.month
        starts = []
        for _ in range(months):
            starts.append(date(y, m, 1))
            m -= 1
            if m <= 0:
                m = 12
                y -= 1
        starts.reverse()
        totals_map = defaultdict(float)
        counts_map = defaultdict(int)
        for s in sales:
            d = _sale_day(s)
            if not d:
                continue
            key = date(d.year, d.month, 1)
            totals_map[key] += float(_dec(s.get("total")))
            counts_map[key] += 1
        labels, totals, counts = [], [], []
        for start in starts:
            labels.append(start.strftime("%m.%Y"))
            totals.append(totals_map.get(start, 0.0))
            counts.append(counts_map.get(start, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    def weekly(start: date, weeks: int):
        totals_map = defaultdict(float)
        counts_map = defaultdict(int)
        for s in sales:
            d = _sale_day(s)
            if not d or d < start:
                continue
            week_start = d - timedelta(days=d.weekday())
            totals_map[week_start] += float(_dec(s.get("total")))
            counts_map[week_start] += 1
        labels, totals, counts = [], [], []
        cursor = start - timedelta(days=start.weekday())
        for i in range(weeks):
            key = cursor + timedelta(weeks=i)
            labels.append(key.strftime("%d.%m"))
            totals.append(totals_map.get(key, 0.0))
            counts.append(counts_map.get(key, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    month_start = today.replace(day=1)
    sales_stats = {
        "d1": hourly_today(),
        "d7": bucket_daily(today - timedelta(days=6), 7),
        "d15": bucket_daily(today - timedelta(days=14), 15),
        "d30": bucket_daily(today - timedelta(days=29), 30),
        "m1": bucket_daily(month_start, max((today - month_start).days + 1, 1)),
        "m3": weekly(today - timedelta(weeks=12), 13),
        "m6": monthly(6),
        "y1": monthly(12),
    }
    return sales_stats


SELLING_LIST_ID = "__selling__"
OTHER_LIST_ID = "__other__"


def _range_window(key: str, today: date) -> tuple[date, date]:
    if key == "d1":
        return today, today
    if key == "d7":
        return today - timedelta(days=6), today
    if key == "d15":
        return today - timedelta(days=14), today
    if key == "d30":
        return today - timedelta(days=29), today
    if key == "m1":
        return today.replace(day=1), today
    if key == "m3":
        return today - timedelta(weeks=12), today
    if key in ("m6", "y1"):
        months = 6 if key == "m6" else 12
        y, m = today.year, today.month
        m -= months - 1
        while m <= 0:
            m += 12
            y -= 1
        return date(y, m, 1), today
    return today - timedelta(days=6), today


def _attach_range_summaries(
    sales_stats: dict, sales: list[dict], today: date, margin_ratio: Decimal
) -> dict:
    ratio = float(margin_ratio or 0)
    sale_dates = [
        d
        for s in sales
        if isinstance(s, dict)
        for d in [_sale_day(s)]
        if d
    ]
    earliest = min(sale_dates) if sale_dates else None
    for key, pack in sales_stats.items():
        start, end = _range_window(key, today)
        checks = 0
        gross = 0.0
        for s in sales:
            if not isinstance(s, dict):
                continue
            d = _sale_day(s)
            if not d or d < start or d > end:
                continue
            checks += 1
            gross += float(_dec(s.get("total")))
        profit = gross * ratio
        pack["summary"] = {
            "checks": checks,
            "gross": gross,
            "profit": profit,
            "margin": (profit / gross * 100.0) if gross > 0 else 0.0,
        }
        # Yuklangan sotuvlar oynasi to'liq qoplamasa — AJAX kerak
        pack["partial"] = bool(earliest is None or earliest > start)
    return sales_stats


def _chart_pack_from_sales(sales: list[dict], today: date, range_key: str) -> dict:
    """AJAX uchun bitta davr grafigi."""
    stats = _build_charts_from_sales(sales, today)
    pack = stats.get(range_key) or {"labels": [], "totals": [], "counts": []}
    return {
        "labels": pack.get("labels") or [],
        "totals": pack.get("totals") or [],
        "counts": pack.get("counts") or [],
    }


def _parse_iso_date(value: str | None) -> date | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def _chart_pack_for_dates(sales: list[dict], start: date, end: date) -> dict:
    """Tanlangan kun yoki oralik uchun grafik (soatlik / kunlik / haftalik / oylik)."""
    if end < start:
        start, end = end, start
    days = (end - start).days + 1
    sales = [s for s in sales if isinstance(s, dict)]

    if days <= 1:
        totals_map: dict[int, float] = defaultdict(float)
        counts_map: dict[int, int] = defaultdict(int)
        for s in sales:
            if _sale_day(s) != start:
                continue
            dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
            if not dt:
                continue
            hour = timezone.localtime(dt).hour
            totals_map[hour] += float(_dec(s.get("total")))
            counts_map[hour] += 1
        labels, totals, counts = [], [], []
        for hour in range(24):
            labels.append(f"{hour:02d}:00")
            totals.append(totals_map.get(hour, 0.0))
            counts.append(counts_map.get(hour, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    if days <= 62:
        totals_map_d: dict[date, float] = defaultdict(float)
        counts_map_d: dict[date, int] = defaultdict(int)
        for s in sales:
            d = _sale_day(s)
            if not d or d < start or d > end:
                continue
            totals_map_d[d] += float(_dec(s.get("total")))
            counts_map_d[d] += 1
        labels, totals, counts = [], [], []
        for i in range(days):
            day = start + timedelta(days=i)
            labels.append(day.strftime("%d.%m"))
            totals.append(totals_map_d.get(day, 0.0))
            counts.append(counts_map_d.get(day, 0))
        return {"labels": labels, "totals": totals, "counts": counts}

    if days <= 180:
        totals_map_w: dict[date, float] = defaultdict(float)
        counts_map_w: dict[date, int] = defaultdict(int)
        for s in sales:
            d = _sale_day(s)
            if not d or d < start or d > end:
                continue
            week_start = d - timedelta(days=d.weekday())
            totals_map_w[week_start] += float(_dec(s.get("total")))
            counts_map_w[week_start] += 1
        labels, totals, counts = [], [], []
        cursor = start - timedelta(days=start.weekday())
        while cursor <= end:
            labels.append(cursor.strftime("%d.%m"))
            totals.append(totals_map_w.get(cursor, 0.0))
            counts.append(counts_map_w.get(cursor, 0))
            cursor += timedelta(weeks=1)
        return {"labels": labels, "totals": totals, "counts": counts}

    totals_map_m: dict[date, float] = defaultdict(float)
    counts_map_m: dict[date, int] = defaultdict(int)
    for s in sales:
        d = _sale_day(s)
        if not d or d < start or d > end:
            continue
        key = date(d.year, d.month, 1)
        totals_map_m[key] += float(_dec(s.get("total")))
        counts_map_m[key] += 1
    labels, totals, counts = [], [], []
    y, m = start.year, start.month
    while date(y, m, 1) <= end:
        key = date(y, m, 1)
        labels.append(key.strftime("%m.%Y"))
        totals.append(totals_map_m.get(key, 0.0))
        counts.append(counts_map_m.get(key, 0))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return {"labels": labels, "totals": totals, "counts": counts}


def _products_by_name(products: list) -> dict[str, SimpleNamespace]:
    out: dict[str, SimpleNamespace] = {}
    for p in products:
        name = (getattr(p, "name", "") or "").strip()
        if not name:
            continue
        if name not in out:
            out[name] = p
        key = " ".join(name.casefold().split())
        if key and key not in out:
            out[key] = p
    return out


def _find_product(
    products_by_id: dict[str, SimpleNamespace],
    products_by_name: dict[str, SimpleNamespace] | None,
    *,
    product_id: str = "",
    product_name: str = "",
) -> SimpleNamespace | None:
    if product_id and str(product_id) in products_by_id:
        return products_by_id[str(product_id)]
    by_name = products_by_name or {}
    name = (product_name or "").strip()
    if not name:
        return None
    if name in by_name:
        return by_name[name]
    key = " ".join(name.casefold().split())
    if key in by_name:
        return by_name[key]
    # Qisman moslik (chekdagi nom biroz farq qilsa)
    for k, p in by_name.items():
        if not isinstance(k, str):
            continue
        kk = " ".join(k.casefold().split())
        if kk and (kk in key or key in kk) and abs(len(kk) - len(key)) <= 8:
            return p
    return None


def _price_within(unit: float, catalog: float, ratio: float = 0.01, floor: float = 1.0) -> bool:
    """Birlik narxi katalog narxiga mos keladimi."""
    if catalog <= 0:
        return False
    return abs(unit - catalog) <= max(floor, abs(catalog) * ratio)


def _is_api_selling_list(pl: dict) -> bool:
    name = (pl.get("name") or "").strip().casefold()
    return bool(
        pl.get("is_selling")
        or name in {"sotuv", "sotish", "sotuv narxi", "retail", "selling"}
    )


def _match_price_list_id(
    unit_price: Decimal,
    product: SimpleNamespace | None,
    price_lists: list[dict],
) -> str:
    """
    Birlik narxini Sotuv yoki Optom (narxlar ro'yxati) ga biriktiradi.
    Mahsulot topilsa — eng yaqin katalog narxi (Boshqa bo'lmaydi).
    Topilmasa — Sotuv (oddiy chakana savdo).
    """
    if not product:
        return SELLING_LIST_ID

    up = float(unit_price)
    if up <= 0:
        return SELLING_LIST_ID

    selling = float(getattr(product, "selling_price", 0) or 0)
    list_prices = getattr(product, "list_prices", None) or {}
    candidates: list[tuple[str, float]] = []
    if selling > 0:
        candidates.append((SELLING_LIST_ID, selling))

    for pl in price_lists:
        lid = str(pl.get("id") or "")
        if not lid or lid not in list_prices:
            continue
        if _is_api_selling_list(pl):
            continue
        lp = float(list_prices[lid] or 0)
        if lp > 0:
            candidates.append((lid, lp))

    if not candidates:
        return SELLING_LIST_ID

    # Eng yaqin narx
    best_id, best_price = min(candidates, key=lambda x: abs(up - x[1]))
    best_dist = abs(up - best_price)

    # Agar sotuv ham deyarli yaqin bo'lsa — Sotuv ustun (bir xil masofa)
    if selling > 0:
        sell_dist = abs(up - selling)
        if sell_dist <= best_dist + 0.5:
            # Qattiq moslik: sotuv ±1.5% ichida
            if _price_within(up, selling, ratio=0.015, floor=1.0):
                return SELLING_LIST_ID
            # Optom aniqroq (kamida 1.5x yaqinroq) bo'lmasa — Sotuv
            if best_id != SELLING_LIST_ID and best_dist * 1.5 < sell_dist:
                return best_id
            return SELLING_LIST_ID

    return best_id


def _aggregate_price_list_stats(
    sale_details: list[dict],
    products_by_id: dict[str, SimpleNamespace],
    products_by_name: dict[str, SimpleNamespace],
    price_lists: list[dict],
) -> list[dict]:
    """Har bir narxlar ro'yxati + sotuv narxi bo'yicha qty/chek/foyda/marja."""
    buckets: dict[str, dict] = {
        SELLING_LIST_ID: {
            "id": SELLING_LIST_ID,
            "name": "Sotuv",
            "qty": 0.0,
            "checks": set(),
            "revenue": 0.0,
            "cost": 0.0,
            "costed_revenue": 0.0,
        },
    }
    selling_list_ids = {
        str(pl.get("id")) for pl in price_lists if pl.get("id") and _is_api_selling_list(pl)
    }
    for pl in price_lists:
        lid = str(pl.get("id") or "")
        if not lid or lid in selling_list_ids or _is_api_selling_list(pl):
            continue
        buckets[lid] = {
            "id": lid,
            "name": (pl.get("name") or "Ro‘yxat").strip() or "Ro‘yxat",
            "qty": 0.0,
            "checks": set(),
            "revenue": 0.0,
            "cost": 0.0,
            "costed_revenue": 0.0,
        }

    for detail in sale_details:
        if not isinstance(detail, dict):
            continue
        sid = str(detail.get("id") or "")
        sale_pl = (
            detail.get("price_list_id")
            or detail.get("price_list")
            or detail.get("list_id")
        )
        if isinstance(sale_pl, dict):
            sale_pl = sale_pl.get("id")
        sale_pl_name = (
            (detail.get("price_list_name") or detail.get("list_name") or "")
            if not isinstance(detail.get("price_list"), dict)
            else (detail.get("price_list") or {}).get("name")
        )
        for item in _sale_items(detail) or []:
            qty_dec = _item_qty(item)
            qty = float(qty_dec)
            unit_price = _item_unit_price(item)
            line_total_dec = _dec(item.get("total") or item.get("line_total"), str(qty_dec * unit_price))
            line_total = float(line_total_dec)
            if unit_price <= 0 and qty_dec > 0 and line_total_dec > 0:
                unit_price = (line_total_dec / qty_dec).quantize(Decimal("0.01"))
            pid, name, _nested = _item_product_ref(item)
            p = _find_product(
                products_by_id,
                products_by_name,
                product_id=pid,
                product_name=name,
            )
            raw_pl = (
                item.get("price_list_id")
                or item.get("price_list")
                or item.get("list_id")
                or sale_pl
            )
            if isinstance(raw_pl, dict):
                raw_pl = raw_pl.get("id")
            list_id = str(raw_pl).strip() if raw_pl not in (None, "") else ""
            if list_id in ("selling", "retail", SELLING_LIST_ID) or list_id in selling_list_ids:
                list_id = SELLING_LIST_ID
            elif list_id and list_id in buckets:
                pass
            elif list_id:
                # API bergan noma’lum optom ro‘yxati — Sotuvga tashlamaymiz
                pl_name = (
                    item.get("price_list_name")
                    or item.get("list_name")
                    or sale_pl_name
                    or "Optom"
                )
                if isinstance(item.get("price_list"), dict):
                    pl_name = item["price_list"].get("name") or pl_name
                buckets[list_id] = {
                    "id": list_id,
                    "name": str(pl_name).strip() or "Optom",
                    "qty": 0.0,
                    "checks": set(),
                    "revenue": 0.0,
                    "cost": 0.0,
                    "costed_revenue": 0.0,
                }
            else:
                list_id = _match_price_list_id(unit_price, p, price_lists)
            if list_id not in buckets:
                list_id = SELLING_LIST_ID
            unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
            bucket = buckets[list_id]
            bucket["qty"] += qty
            bucket["revenue"] += line_total
            if unit_cost > 0:
                bucket["cost"] += float(unit_cost * qty_dec)
                bucket["costed_revenue"] += line_total
            if sid:
                bucket["checks"].add(sid)

    out = []
    order = [SELLING_LIST_ID] + [
        str(pl.get("id"))
        for pl in price_lists
        if pl.get("id") and str(pl.get("id")) in buckets
    ]
    seen = set()
    total_rev = 0.0
    total_profit = 0.0
    total_cost = 0.0
    total_costed = 0.0
    total_qty = 0.0
    all_checks: set = set()
    for lid in order:
        if lid in seen or lid not in buckets:
            continue
        seen.add(lid)
        b = buckets[lid]
        rev = float(b["revenue"])
        if rev <= 0:
            continue
        cost = float(b["cost"])
        costed_rev = float(b.get("costed_revenue") or 0)
        # Foyda: faqat sotib olish narxi bor qatorlar (1200 - 1000 = 200)
        profit = costed_rev - cost
        total_rev += rev
        total_profit += profit
        total_cost += cost
        total_costed += costed_rev
        total_qty += float(b["qty"])
        all_checks |= b["checks"]
        markup = (profit / cost * 100.0) if cost > 0 else 0.0
        margin = (profit / costed_rev * 100.0) if costed_rev > 0 else 0.0
        out.append(
            {
                "id": b["id"],
                "name": b["name"],
                "qty": round(float(b["qty"]), 3),
                "checks": len(b["checks"]),
                "revenue": rev,
                "cost": cost,
                "costed_revenue": costed_rev,
                "profit": profit,
                "margin": margin,
                "markup": markup,
                "share": 0.0,
                "is_total": False,
            }
        )
    if total_rev > 0:
        for row in out:
            row["share"] = row["revenue"] / total_rev * 100.0
    if out:
        out.append(
            {
                "id": "__all__",
                "name": "Jami",
                "qty": round(total_qty, 3),
                "checks": len(all_checks),
                "revenue": total_rev,
                "cost": total_cost,
                "costed_revenue": total_costed,
                "profit": total_profit,
                "margin": (total_profit / total_costed * 100.0) if total_costed > 0 else 0.0,
                "markup": (total_profit / total_cost * 100.0) if total_cost > 0 else 0.0,
                "share": 100.0,
                "is_total": True,
            }
        )
    return out


def _scale_price_list_stats(
    lists: list[dict],
    target_gross: float,
    target_checks: int,
) -> list[dict]:
    """
    Namunadagi Sotuv/Optom ulushini to'liq davr tushumiga (gross) masshtablaydi.
    Shunda Jami = kunlik/davr jami (masalan 231 mln) bilan mos keladi.
    """
    parts = [dict(r) for r in lists if isinstance(r, dict) and not r.get("is_total")]
    if not parts:
        if target_gross > 0:
            return [
                {
                    "id": "__all__",
                    "name": "Jami",
                    "qty": 0.0,
                    "checks": int(target_checks),
                    "revenue": float(target_gross),
                    "cost": 0.0,
                    "profit": 0.0,
                    "margin": 0.0,
                    "markup": 0.0,
                    "share": 100.0,
                    "is_total": True,
                    "scaled": True,
                }
            ]
        return lists

    sample_rev = sum(float(r.get("revenue") or 0) for r in parts)
    if sample_rev <= 0 or target_gross <= 0:
        return lists

    # Allaqachon to'liq qamrab olgan bo'lsa — masshtablash shart emas
    if abs(sample_rev - target_gross) / target_gross < 0.02 and sample_rev >= target_gross * 0.95:
        # Jami ni baribir to'liq gross ga tenglashtirish
        out = [r for r in lists if not r.get("is_total")]
        total_cost = sum(float(r.get("cost") or 0) for r in out)
        total_profit = sum(float(r.get("profit") or 0) for r in out)
        total_costed = sum(float(r.get("costed_revenue") or 0) for r in out)
        total_qty = sum(float(r.get("qty") or 0) for r in out)
        for r in out:
            r["share"] = (float(r["revenue"]) / target_gross * 100.0) if target_gross else 0.0
        out.append(
            {
                "id": "__all__",
                "name": "Jami",
                "qty": round(total_qty, 3),
                "checks": int(target_checks),
                "revenue": float(target_gross),
                "cost": total_cost,
                "costed_revenue": total_costed,
                "profit": total_profit,
                "margin": (total_profit / total_costed * 100.0) if total_costed else 0.0,
                "markup": (total_profit / total_cost * 100.0) if total_cost > 0 else 0.0,
                "share": 100.0,
                "is_total": True,
                "scaled": False,
            }
        )
        return out

    scale = target_gross / sample_rev
    scaled: list[dict] = []
    for r in parts:
        share = float(r.get("revenue") or 0) / sample_rev
        rev = float(r["revenue"]) * scale
        cost = float(r.get("cost") or 0) * scale
        costed = float(r.get("costed_revenue") or 0) * scale
        profit = costed - cost
        qty = float(r.get("qty") or 0) * scale
        checks_est = int(round(target_checks * share)) if target_checks and share > 0 else int(r.get("checks") or 0)
        scaled.append(
            {
                "id": r.get("id"),
                "name": r.get("name"),
                "qty": round(qty, 3),
                "checks": max(checks_est, 1) if rev > 0 else 0,
                "revenue": rev,
                "cost": cost,
                "costed_revenue": costed,
                "profit": profit,
                "margin": (profit / costed * 100.0) if costed > 0 else 0.0,
                "markup": (profit / cost * 100.0) if cost > 0 else 0.0,
                "share": share * 100.0,
                "is_total": False,
                "scaled": True,
            }
        )

    total_cost = sum(float(r["cost"]) for r in scaled)
    total_profit = sum(float(r["profit"]) for r in scaled)
    total_costed = sum(float(r.get("costed_revenue") or 0) for r in scaled)
    total_qty = sum(float(r["qty"]) for r in scaled)
    scaled.append(
        {
            "id": "__all__",
            "name": "Jami",
            "qty": round(total_qty, 3),
            "checks": int(target_checks),
            "revenue": float(target_gross),
            "cost": total_cost,
            "costed_revenue": total_costed,
            "profit": total_profit,
            "margin": (total_profit / total_costed * 100.0) if total_costed else 0.0,
            "markup": (total_profit / total_cost * 100.0) if total_cost > 0 else 0.0,
            "share": 100.0,
            "is_total": True,
            "scaled": True,
        }
    )
    return scaled


def _shift_dt(value) -> datetime | None:
    return _parse_dt(value)


def _normalize_api_shift(raw: dict, today: date) -> dict:
    """TezPOS smena obyektini bir xil formatga keltiradi."""
    opened = _shift_dt(
        raw.get("opened_at")
        or raw.get("opened_at_local")
        or raw.get("start_at")
        or raw.get("started_at")
        or raw.get("open_time")
        or raw.get("created_at")
    )
    closed = _shift_dt(
        raw.get("closed_at")
        or raw.get("closed_at_local")
        or raw.get("end_at")
        or raw.get("ended_at")
        or raw.get("close_time")
        or raw.get("finished_at")
    )
    status_raw = (raw.get("status") or "").strip().lower()
    is_open = status_raw in ("open", "opened", "active", "ochiq")
    if not status_raw:
        is_open = bool(opened) and not closed
    if closed and status_raw not in ("open", "opened", "active", "ochiq"):
        is_open = False
    if not status_raw:
        status_raw = "open" if is_open else "closed"
    gross = _dec(
        raw.get("sales_total")
        or raw.get("total_sales")
        or raw.get("gross")
        or raw.get("total")
        or raw.get("revenue")
        or 0
    )
    checks = int(
        raw.get("sales_count")
        or raw.get("receipts_count")
        or raw.get("checks_count")
        or raw.get("orders_count")
        or 0
    )
    cashier = (
        raw.get("cashier_name")
        or raw.get("user_name")
        or raw.get("opened_by_name")
        or raw.get("employee_name")
        or ""
    )
    sid = str(raw.get("id") or raw.get("uuid") or "")
    return {
        "id": sid or (opened.isoformat() if opened else ""),
        "source": "api",
        "status": "open" if is_open else "closed",
        "status_label": "Ochiq" if is_open else "Yopilgan",
        "opened_at": opened.isoformat() if opened else "",
        "closed_at": closed.isoformat() if closed else "",
        "opened_display": timezone.localtime(opened).strftime("%d.%m.%Y %H:%M") if opened else "—",
        "closed_display": (
            timezone.localtime(closed).strftime("%d.%m.%Y %H:%M") if closed else ("Hozir" if is_open else "—")
        ),
        "cashier": (cashier or "").strip() or "Kassir",
        "opening_cash": float(_dec(raw.get("opening_cash") or raw.get("open_cash") or 0)),
        "closing_cash": float(_dec(raw.get("closing_cash") or raw.get("close_cash") or 0)),
        "gross": float(gross),
        "checks": checks,
        "profit": float(_dec(raw.get("profit") or 0)),
        "margin": float(_dec(raw.get("margin") or 0)),
        "sale_ids": [str(x) for x in (raw.get("sale_ids") or []) if x],
        "price_lists": [],
        "raw": raw,
    }


def _build_shifts_from_sales(
    sales: list[dict],
    *,
    today: date,
    gap_hours: float = 4.0,
    margin_ratio: Decimal = Decimal("0"),
) -> list[dict]:
    """
    Agar TezPOS /shifts API yo'q bo'lsa — sotuvlar oralig'idan smena yig'adi.
    Ketma-ket cheklar orasida gap_hours soatdan ko'p bo'lsa yangi smena.
    """
    timed: list[tuple[datetime, dict]] = []
    for s in sales:
        if not isinstance(s, dict):
            continue
        dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
        if not dt:
            continue
        timed.append((timezone.localtime(dt), s))
    timed.sort(key=lambda x: x[0])
    if not timed:
        return []

    gap = timedelta(hours=gap_hours)
    groups: list[list[tuple[datetime, dict]]] = [[timed[0]]]
    for dt, sale in timed[1:]:
        prev_dt = groups[-1][-1][0]
        if dt - prev_dt > gap:
            groups.append([(dt, sale)])
        else:
            groups[-1].append((dt, sale))

    out: list[dict] = []
    for idx, group in enumerate(reversed(groups), start=1):
        opened = group[0][0]
        closed = group[-1][0]
        # Dastur yopiq bo‘lsa sayt ham ochiq deb ko‘rsatmasin
        is_open = False
        gross = sum((_dec(s.get("total")) for _, s in group), Decimal("0"))
        checks = len(group)
        profit = (gross * margin_ratio).quantize(Decimal("0.01")) if margin_ratio > 0 else Decimal("0")
        margin = float((profit / gross * 100) if gross > 0 else 0)
        sale_ids = [str(s.get("id")) for _, s in group if s.get("id")]
        out.append(
            {
                "id": f"session-{opened.strftime('%Y%m%d%H%M')}-{sale_ids[0][:8] if sale_ids else idx}",
                "source": "sales",
                "status": "open" if is_open else "closed",
                "status_label": "Ochiq" if is_open else "Yopilgan",
                "opened_at": opened.isoformat(),
                "closed_at": "" if is_open else closed.isoformat(),
                "opened_display": opened.strftime("%d.%m.%Y %H:%M"),
                "closed_display": "Hozir" if is_open else closed.strftime("%d.%m.%Y %H:%M"),
                "cashier": "Smena",
                "opening_cash": 0.0,
                "closing_cash": 0.0,
                "gross": float(gross),
                "checks": checks,
                "profit": float(profit),
                "margin": margin,
                "sale_ids": sale_ids,
                "price_lists": [],
            }
        )
    return out


def _enrich_shift_with_sales(
    shift: dict,
    sales: list[dict],
    *,
    margin_ratio: Decimal,
) -> dict:
    """API smenasiga mos sotuvlarni vaqt oralig'idan biriktiradi."""
    if shift.get("sale_ids"):
        ids = set(shift["sale_ids"])
        matched = [s for s in sales if str(s.get("id")) in ids]
    else:
        opened = _shift_dt(shift.get("opened_at"))
        closed = _shift_dt(shift.get("closed_at")) or timezone.now()
        matched = []
        if opened:
            for s in sales:
                dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
                if not dt:
                    continue
                if opened <= dt <= closed:
                    matched.append(s)
        shift["sale_ids"] = [str(s.get("id")) for s in matched if s.get("id")]
    if matched:
        prev_checks = int(shift.get("checks") or 0)
        if len(matched) >= prev_checks:
            gross = sum((_dec(s.get("total")) for s in matched), Decimal("0"))
            shift["gross"] = float(gross)
            shift["checks"] = len(matched)
            if not shift.get("profit"):
                shift["profit"] = 0.0
                shift["margin"] = 0.0
    return shift


def _match_shift_sales(
    sales: list[dict],
    *,
    opened_dt: datetime | None,
    closed_dt: datetime | None,
    sale_ids: list[str] | None = None,
    is_open: bool = False,
) -> list[dict]:
    """Smena ochilish–yopilish oralig‘idagi cheklar (00:00 chegarasidan mustaqil)."""
    ids = [str(x) for x in (sale_ids or []) if x]
    if is_open and opened_dt and closed_dt:
        matched = []
        for s in sales:
            dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
            if dt and opened_dt <= dt <= closed_dt:
                matched.append(s)
        return matched
    if ids:
        idset = set(ids)
        matched = [s for s in sales if str(s.get("id")) in idset]
        if opened_dt and closed_dt and len(matched) < len(ids):
            seen = {str(s.get("id")) for s in matched}
            for s in sales:
                sid = str(s.get("id") or "")
                if sid in seen:
                    continue
                dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
                if dt and opened_dt <= dt <= closed_dt:
                    matched.append(s)
                    seen.add(sid)
        return matched
    if opened_dt and closed_dt:
        matched = []
        for s in sales:
            dt = _parse_dt(s.get("completed_at") or s.get("created_at"))
            if dt and opened_dt <= dt <= closed_dt:
                matched.append(s)
        return matched
    return list(sales)


def _sale_details_merged(matched: list[dict], fetched: dict[str, dict]) -> list[dict]:
    """Inline items + alohida fetch — bitta ro‘yxat."""
    out: list[dict] = []
    for s in matched:
        sid = str(s.get("id") or "")
        detail = fetched.get(sid) if sid else None
        if detail and _sale_items(detail):
            out.append(detail)
        elif _sale_items(s):
            out.append(s)
        elif detail:
            out.append(detail)
        elif sid:
            out.append(s)
    return out


def _shift_price_list_stats(
    matched: list[dict],
    fetched: dict[str, dict],
    products_by_id: dict,
    products_by_name: dict,
    price_lists: list[dict],
    gross: float,
    checks: int,
) -> tuple[list[dict], float, float]:
    """Sotuv/Optom statistikasi — to‘liq cheklar bo‘yicha, kerak bo‘lsa scale."""
    details_for_stats = _sale_details_merged(matched, fetched)
    lists = _aggregate_price_list_stats(
        details_for_stats, products_by_id, products_by_name, price_lists
    )
    jami = next((r for r in lists if r.get("is_total")), None)
    sample_checks = int(jami.get("checks") or 0) if jami else len(details_for_stats)
    sample_rev = float(jami.get("revenue") or 0) if jami else 0.0
    need_scale = (
        gross > 0
        and sample_rev > 0
        and checks > 0
        and (
            sample_checks < checks * 0.85
            or abs(sample_rev - gross) / gross > 0.05
        )
    )
    if need_scale:
        lists = _scale_price_list_stats(lists, gross, checks)
        jami = next((r for r in lists if r.get("is_total")), None)

    if jami and float(jami.get("revenue") or 0) > 0:
        rev = float(jami["revenue"])
        ratio = float(jami.get("profit") or 0) / rev if rev else 0.0
        profit = gross * ratio if ratio else float(jami.get("profit") or 0)
        margin = ratio * 100.0
    else:
        profit_dec = Decimal("0")
        for s in matched:
            sid = str(s.get("id") or "")
            detail = fetched.get(sid) or s
            _, pft = _estimate_sale_profit(
                detail, products_by_id, _dec(s.get("total")), products_by_name
            )
            profit_dec += pft
        profit = float(profit_dec)
        margin = (profit / gross * 100.0) if gross > 0 else 0.0
    return lists, profit, margin


def _build_price_list_stats_by_range(
    sales: list[dict],
    details: dict[str, dict],
    today: date,
    products_by_id: dict[str, SimpleNamespace],
    products_by_name: dict[str, SimpleNamespace],
    price_lists: list[dict],
    range_keys: list[str] | None = None,
) -> dict[str, list[dict]]:
    keys = range_keys or ["d1", "d7", "d15", "d30", "m1", "m3", "m6", "y1"]
    out: dict[str, list[dict]] = {}
    for key in keys:
        start, end = _range_window(key, today)
        ranged = []
        for s in sales:
            if not isinstance(s, dict):
                continue
            d = _sale_day(s)
            if not d or d < start or d > end:
                continue
            sid = str(s.get("id") or "")
            detail = details.get(sid)
            if detail:
                ranged.append(detail)
        out[key] = _aggregate_price_list_stats(
            ranged, products_by_id, products_by_name, price_lists
        )
    return out


def _refine_summaries_from_details(
    sales_stats: dict,
    sales: list[dict],
    details: dict[str, dict],
    today: date,
    products_by_id: dict,
    products_by_name: dict,
) -> None:
    """Agar chek detallari bo'lsa — foydani aniqroq hisoblaydi."""
    for key, pack in sales_stats.items():
        start, end = _range_window(key, today)
        profit = 0.0
        gross = 0.0
        used = 0
        total_in_range = 0
        for s in sales:
            if not isinstance(s, dict):
                continue
            d = _sale_day(s)
            if not d or d < start or d > end:
                continue
            total_in_range += 1
            total = float(_dec(s.get("total")))
            gross += total
            sid = str(s.get("id") or "")
            detail = details.get(sid)
            if not detail:
                continue
            used += 1
            cost, sale_profit = _estimate_sale_profit(
                detail, products_by_id, _dec(total), products_by_name
            )
            profit += float(sale_profit)
        summary = pack.get("summary") or {}
        if used > 0 and used >= max(1, int(total_in_range * 0.5)):
            # Detallar yetarli — aniq foyda; qolganlar uchun o'rtacha marja
            if used < total_in_range and profit > 0 and gross > 0:
                sample_gross = 0.0
                for s in sales:
                    if not isinstance(s, dict):
                        continue
                    d = _sale_day(s)
                    if not d or d < start or d > end:
                        continue
                    sid = str(s.get("id") or "")
                    if sid in details:
                        sample_gross += float(_dec(s.get("total")))
                ratio = (profit / sample_gross) if sample_gross > 0 else 0.0
                full_gross = float(summary.get("gross") or gross)
                profit = full_gross * ratio
                gross = full_gross
            else:
                gross = float(summary.get("gross") or gross)
            summary["profit"] = profit
            summary["margin"] = (profit / gross * 100.0) if gross > 0 else 0.0
            pack["summary"] = summary


def _fetch_sale_details(
    token: str,
    server: str,
    sale_ids: list[str],
    limit: int = 40,
    *,
    per_sale_timeout: float = 5.0,
    overall_timeout: float | None = None,
) -> dict[str, dict]:
    """Fetch sale receipts; never raises — timeouts/errors skip that sale."""
    out: dict[str, dict] = {}
    ids = [str(x) for x in sale_ids[:limit] if x]

    def one(sid: str):
        try:
            # get_sale default 5s — tez namuna uchun qisqaroq
            data = tezpos_api.api_request(
                "GET",
                f"/api/sales/{sid}/",
                token=token,
                server_name=server,
                timeout=per_sale_timeout,
            )
            return sid, data if isinstance(data, dict) else None
        except (tezpos_api.TezPosApiError, TimeoutError, OSError):
            return sid, None
        except Exception:
            return sid, None

    if not ids:
        return out
    workers = min(10, max(4, len(ids)))
    deadline = time.time() + overall_timeout if overall_timeout else None
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(one, sid) for sid in ids]
        for fut in as_completed(futures):
            if deadline and time.time() > deadline:
                break
            try:
                sid, data = fut.result()
            except Exception:
                continue
            if data:
                out[sid] = data
    return out


def _fallback_sotuv_only_lists(
    gross: float,
    checks: int,
    margin_ratio: Decimal,
    price_lists: list[dict] | None = None,
) -> list[dict]:
    """Detallar yo‘q — tushum ko‘rinsin, foyda taxmin qilinmasin."""
    del margin_ratio, price_lists
    if gross <= 0:
        return []
    row = {
        "id": SELLING_LIST_ID,
        "name": "Sotuv",
        "qty": 0.0,
        "checks": int(checks),
        "revenue": float(gross),
        "cost": 0.0,
        "costed_revenue": 0.0,
        "profit": 0.0,
        "margin": 0.0,
        "markup": 0.0,
        "share": 100.0,
        "is_total": False,
        "scaled": True,
        "estimated": True,
    }
    jami = dict(row)
    jami["id"] = "__all__"
    jami["name"] = "Jami"
    jami["is_total"] = True
    return [row, jami]


def _catalog_margin_ratio(products: list) -> Decimal:
    """O'rtacha sof marja ulushi: faqat sotib olish narxi kiritilgan tovarlar."""
    ratios = []
    for p in products:
        selling = getattr(p, "selling_price", None) or Decimal("0")
        cost = getattr(p, "cost_price", None) or Decimal("0")
        if selling > 0 and cost > 0 and cost <= selling:
            ratios.append((selling - cost) / selling)
    if not ratios:
        return Decimal("0")
    return sum(ratios, Decimal("0")) / Decimal(len(ratios))


def _item_product_ref(item: dict) -> tuple[str, str, dict]:
    """Chek qatoridagi mahsulot id/nomi va ichki obyekt."""
    nested = item.get("product") if isinstance(item.get("product"), dict) else {}
    pid = item.get("product_id") or item.get("product") or nested.get("id") or ""
    if isinstance(pid, dict):
        nested = pid if not nested else nested
        pid = pid.get("id") or ""
    pid = str(pid or "").strip()
    if pid.startswith("{") or pid in ("None", "null"):
        pid = ""
    name = (
        item.get("product_name")
        or item.get("name")
        or nested.get("name")
        or ""
    )
    return pid, str(name or "").strip(), nested


def _resolve_item_unit_cost(
    item: dict,
    products_by_id: dict[str, SimpleNamespace],
    products_by_name: dict[str, SimpleNamespace] | None = None,
) -> Decimal:
    """Sotib olish narxi; kiritilmagan bo'lsa 0 (foydaga qo'shilmaydi)."""
    pid, name, nested = _item_product_ref(item)
    unit_cost = _dec(
        item.get("unit_cost")
        or item.get("cost_price")
        or item.get("purchase_price")
        or item.get("buy_price")
        or nested.get("cost_price")
        or nested.get("purchase_price")
        or nested.get("buy_price")
        or nested.get("cost")
    )
    if unit_cost > 0:
        return unit_cost
    p = _find_product(
        products_by_id,
        products_by_name,
        product_id=pid,
        product_name=name,
    )
    if p and (p.cost_price or Decimal("0")) > 0:
        return p.cost_price
    return Decimal("0")


def _estimate_sale_cost(
    sale_detail: dict,
    products_by_id: dict[str, SimpleNamespace],
    products_by_name: dict[str, SimpleNamespace] | None = None,
) -> Decimal:
    """Faqat sotib olish narxi bor qatorlar tannarxi."""
    cost = Decimal("0")
    for item in _sale_items(sale_detail):
        unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
        if unit_cost <= 0:
            continue
        cost += _item_qty(item) * unit_cost
    return cost


def _estimate_sale_profit(
    sale_detail: dict,
    products_by_id: dict[str, SimpleNamespace],
    total: Decimal,
    products_by_name: dict[str, SimpleNamespace] | None = None,
    margin_ratio: Decimal | None = None,
) -> tuple[Decimal, Decimal]:
    """
    Qaytaradi: (tannarx, foyda).
    Sotib olish narxi yo'q tovarlar tashlab ketiladi — ularning tushumi foydaga kirmaydi.
    """
    del margin_ratio
    level = _sale_level_cost_profit(sale_detail, total)
    if level is not None:
        return level
    items = _sale_items(sale_detail)
    if not items:
        return Decimal("0"), Decimal("0")

    cost = Decimal("0")
    costed_rev = Decimal("0")
    for item in items:
        qty = _item_qty(item)
        unit_price = _item_unit_price(item)
        line_total = _dec(item.get("total") or item.get("line_total"), str(qty * unit_price))
        unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
        if unit_cost <= 0:
            continue
        cost += qty * unit_cost
        costed_rev += line_total
    profit = (costed_rev - cost).quantize(Decimal("0.01"))
    return cost.quantize(Decimal("0.01")), profit


def _serialize_sale_payload(
    sale_detail: dict,
    cashier: str,
    products_by_id: dict,
    products_by_name: dict | None = None,
    margin_ratio: Decimal | None = None,
) -> dict:
    del margin_ratio
    items = []
    items_cost = Decimal("0")
    costed_rev = Decimal("0")
    for item in _sale_items(sale_detail):
        qty = _item_qty(item)
        unit_price = _item_unit_price(item)
        p = _find_product(
            products_by_id,
            products_by_name,
            product_id=str(item.get("product_id") or item.get("product") or ""),
            product_name=item.get("product_name") or item.get("name") or "",
        )
        unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
        line_total = _dec(item.get("total") or item.get("line_total"), str(qty * unit_price))
        line_cost = qty * unit_cost if unit_cost > 0 else Decimal("0")
        if unit_cost > 0:
            items_cost += line_cost
            costed_rev += line_total
        items.append(
            {
                "name": item.get("product_name") or item.get("name") or (p.name if p else "Mahsulot"),
                "qty": float(qty),
                "unit_price": float(unit_price),
                "unit_cost": float(unit_cost),
                "line_total": float(line_total),
                "line_cost": float(line_cost),
            }
        )
    total_amount = _dec(sale_detail.get("total"))
    level = _sale_level_cost_profit(sale_detail, total_amount)
    if level is not None and not items:
        items_cost, profit = level
    else:
        profit = (costed_rev - items_cost).quantize(Decimal("0.01"))
        if items_cost <= 0 and level is not None:
            items_cost, profit = level
    dt = _parse_dt(sale_detail.get("completed_at") or sale_detail.get("created_at"))
    created_display = timezone.localtime(dt).strftime("%d.%m.%Y, %H:%M") if dt else ""
    method = (
        sale_detail.get("payment_type")
        or sale_detail.get("payment_method")
        or "cash"
    )
    receipt_no = _display_receipt_number(sale_detail) or _receipt_number(sale_detail)
    return {
        "id": str(sale_detail.get("id") or ""),
        "receipt_number": receipt_no,
        "receipt_no": receipt_no,
        "created_at": dt.isoformat() if dt else "",
        "created_display": created_display,
        "time": timezone.localtime(dt).strftime("%H:%M") if dt else "",
        "customer": sale_detail.get("customer_name") or "",
        "cashier": cashier,
        "status": "Yakunlangan",
        "type": "Sotilgan",
        "payment_method": method,
        "payment_label": _payment_label(method),
        "total_amount": float(total_amount),
        "total": float(total_amount),
        "total_cost": float(items_cost),
        "cost": float(items_cost),
        "profit": float(profit),
        "discount": float(_dec(sale_detail.get("discount_amount"))),
        "items": items,
        "needs_detail": not bool(items),
    }


def _build_abc_xyz(products, item_rows, today: date):
    start = today - timedelta(days=29)
    revenue = defaultdict(Decimal)
    daily_qty = defaultdict(lambda: defaultdict(float))
    for row in item_rows:
        pid = str(row.get("product_id") or "")
        if not pid:
            continue
        qty = _dec(row.get("quantity"))
        unit_price = _dec(row.get("unit_price"))
        revenue[pid] += qty * unit_price
        day = row.get("day")
        if isinstance(day, date):
            daily_qty[pid][day] += float(qty)

    total_rev = sum(revenue.values()) or Decimal("1")
    ranked = sorted(revenue.items(), key=lambda x: x[1], reverse=True)
    abc_map = {}
    cumulative = Decimal("0")
    for pid, rev in ranked:
        cumulative += rev
        share = cumulative / total_rev * 100
        if share <= 80:
            abc_map[pid] = "A"
        elif share <= 95:
            abc_map[pid] = "B"
        else:
            abc_map[pid] = "C"
    for p in products:
        abc_map.setdefault(str(p.id), "C")

    xyz_map = {}
    for p in products:
        pid = str(p.id)
        series = [daily_qty[pid].get(start + timedelta(days=i), 0.0) for i in range(30)]
        if sum(series) <= 0:
            xyz_map[pid] = "Z"
            continue
        avg = mean(series)
        if avg <= 0:
            xyz_map[pid] = "Z"
            continue
        cv = pstdev(series) / avg if len(series) > 1 else 0
        if cv < 0.5:
            xyz_map[pid] = "X"
        elif cv < 1.0:
            xyz_map[pid] = "Y"
        else:
            xyz_map[pid] = "Z"

    matrix = {f"{a}{x}": [] for a in "ABC" for x in "XYZ"}
    rows = []
    for p in products:
        pid = str(p.id)
        group = f"{abc_map[pid]}{xyz_map[pid]}"
        rev = revenue.get(pid, Decimal("0"))
        entry = {
            "product": p,
            "abc": abc_map[pid],
            "xyz": xyz_map[pid],
            "group": group,
            "revenue": rev,
            "share": float(rev / total_rev * 100),
        }
        matrix[group].append(entry)
        rows.append(entry)
    rows.sort(key=lambda r: r["revenue"], reverse=True)
    matrix_counts = {k: len(v) for k, v in matrix.items()}
    return rows, matrix_counts, float(total_rev)


def _build_near_min_stock(products):
    rows = []
    for p in products:
        min_qty = p.min_stock if p.min_stock is not None else Decimal("0")
        if p.stock_qty <= min_qty:
            rows.append(
                {
                    "id": p.id,
                    "name": p.name,
                    "barcode": getattr(p, "barcode", "") or "",
                    "unit": getattr(p, "unit", None) or "dona",
                    "stock": float(p.stock_qty),
                    "min_stock": float(min_qty),
                    "title": "Kam qoldiq",
                    "text": f"{p.name} — qoldiq {p.stock_qty}",
                    "channels": ["SMS", "Bildirishnoma"],
                    "level": "warn" if p.stock_qty > 0 else "danger",
                }
            )
    rows.sort(key=lambda r: (r["stock"] / r["min_stock"] if r["min_stock"] else 0, r["stock"]))
    return rows


def _export_daily_sales_excel(day_sales_payload, sale_date, cashier):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Kunlik sotuvlar"
    headers = [
        "Chek raqami",
        "ID",
        "Sana",
        "Kassir",
        "Mijoz",
        "Turi",
        "Status",
        "To‘lov",
        "Jami",
        "Umumiy tannarxi",
        "Foyda",
        "Mahsulotlar",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for payload in day_sales_payload:
        products_txt = " | ".join(
            f"{it['name']} ({it['qty']:g} x {it['unit_price']:g})"
            for it in (payload.get("items") or [])
        )
        receipt = (
            payload.get("receipt_number")
            or payload.get("receipt_no")
            or payload.get("id")
            or ""
        )
        ws.append(
            [
                str(receipt),
                str(payload.get("id") or ""),
                payload.get("created_display") or "",
                cashier or payload.get("cashier") or "",
                payload.get("customer") or "",
                payload.get("type") or "Sotilgan",
                payload.get("status") or "Yakunlangan",
                payload.get("payment_label") or "",
                float(payload.get("total_amount") or payload.get("total") or 0),
                float(payload.get("total_cost") or payload.get("cost") or 0),
                float(payload.get("profit") or 0),
                products_txt,
            ]
        )
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 16 if idx < 12 else 40
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["L"].width = 48
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=11):
        for cell in row:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"kunlik_sotuv_{sale_date.isoformat()}.xlsx"
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _save_product_via_api(request, token: str, server: str, product_id: str | None):
    name = (request.POST.get("name") or "").strip()
    if not name:
        return False, {"name": ["Mahsulot nomi majburiy."]}

    codes = [c.strip() for c in request.POST.getlist("barcodes") if c.strip()]
    barcode = codes[0] if codes else (request.POST.get("barcode") or "").strip()
    payload = {
        "name": name,
        "unit": (request.POST.get("unit") or "dona").strip() or "dona",
        "price": str(_dec(request.POST.get("selling_price"))),
        "cost_price": str(_dec(request.POST.get("cost_price"))),
        "quantity": str(_dec(request.POST.get("stock_qty"))),
        "min_stock": str(_dec(request.POST.get("min_stock"))),
        "barcode": barcode,
        "barcodes": codes or ([barcode] if barcode else []),
        "is_active": True,
    }
    try:
        if product_id:
            tezpos_api.update_product(token, server, product_id, payload)
        else:
            tezpos_api.create_product(token, server, payload)
        return True, {}
    except tezpos_api.TezPosApiError as exc:
        return False, {"__all__": [str(exc)]}


def _product_payload_from_import_row(row: dict) -> tuple[dict | None, str | None]:
    """Import qatoridan TezPOS create payload."""
    name = str(row.get("name") or "").strip()
    if not name:
        return None, "Mahsulot nomi bo‘sh"

    codes = collect_product_barcodes(
        parse_barcodes_cell(row.get("barcodes")),
        parse_barcodes_cell(row.get("barcode")),
    )
    barcode = codes[0] if codes else ""

    selling = _dec(row.get("selling_price") or row.get("price"))
    cost = _dec(row.get("cost_price"))
    qty = _dec(row.get("stock_qty") or row.get("quantity"))
    min_stock = _dec(row.get("min_stock"), "0")
    unit = str(row.get("unit") or "dona").strip() or "dona"

    payload: dict = {
        "name": name,
        "unit": unit,
        "price": str(selling),
        "cost_price": str(cost),
        "quantity": str(qty),
        "min_stock": str(min_stock),
        "barcode": barcode or (codes[0] if codes else ""),
        "barcodes": codes or ([barcode] if barcode else []),
        "is_active": True,
    }

    category = str(row.get("category") or "").strip()
    brand = str(row.get("brand") or "").strip()
    if category:
        payload["category"] = category
    if brand:
        payload["brand"] = brand

    list_prices_in = row.get("list_prices")
    list_prices: dict[str, str] = {}
    if isinstance(list_prices_in, dict):
        for k, v in list_prices_in.items():
            key = str(k).strip()
            if not key:
                continue
            list_prices[key] = str(_dec(v))
    # pl_<id> kalitlari
    for k, v in row.items():
        if not str(k).startswith("pl_"):
            continue
        lid = str(k)[3:].strip()
        if lid:
            list_prices[lid] = str(_dec(v))
    if list_prices:
        payload["list_prices"] = list_prices

    return payload, None


@login_required
@require_POST
def cabinet_products_import(request):
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]

    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Noto‘g‘ri JSON"}, status=400)

    rows = body.get("products") or []
    if not isinstance(rows, list) or not rows:
        return JsonResponse({"error": "Mahsulotlar ro‘yxati bo‘sh"}, status=400)
    if len(rows) > 5000:
        return JsonResponse({"error": "Bir martada eng ko‘pi 5000 ta mahsulot"}, status=400)

    created = 0
    failed: list[dict] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            failed.append({"row": idx, "error": "Noto‘g‘ri qator"})
            continue
        payload, err = _product_payload_from_import_row(row)
        if err or not payload:
            failed.append({"row": idx, "name": str(row.get("name") or ""), "error": err or "Xato"})
            continue
        try:
            tezpos_api.create_product(token, server, payload)
            created += 1
        except tezpos_api.TezPosApiError as exc:
            failed.append({"row": idx, "name": payload.get("name", ""), "error": str(exc)})
        except Exception as exc:  # noqa: BLE001
            failed.append({"row": idx, "name": payload.get("name", ""), "error": str(exc)})

    return JsonResponse(
        {
            "ok": True,
            "created": created,
            "failed": failed,
            "total": len(rows),
        }
    )


_EXPORT_FIELD_LABELS = {
    "name": "Mahsulot nomi",
    "barcode": "Shtrixkod (barkod)",
    "selling_price": "Sotuv narxi",
    "cost_price": "Sotib olish narxi",
    "stock_qty": "Omborda qoldiq",
    "unit": "O‘lchov birligi",
    "category": "Bo‘lim",
    "brand": "Brend",
    "min_stock": "Minimal qoldiq",
}


def _export_cell_value(product: SimpleNamespace, key: str, price_lists_by_id: dict) -> str | float | int:
    list_prices = getattr(product, "list_prices", None) or {}
    if key.startswith("pl_"):
        pid = key[3:]
        return float(list_prices.get(pid) or list_prices.get(str(pid)) or 0)
    if key == "name":
        return product.name or ""
    if key == "barcode":
        return format_barcodes_excel_cell(
            collect_product_barcodes(
                getattr(product, "barcode_list", None),
                getattr(product, "barcodes", None),
                getattr(product, "barcode", None),
            )
        )
    if key == "selling_price":
        return float(product.selling_price or 0)
    if key == "cost_price":
        return float(product.cost_price or 0)
    if key == "stock_qty":
        return float(product.stock_qty or 0)
    if key == "unit":
        return product.unit or "dona"
    if key == "category":
        return product.category or ""
    if key == "brand":
        return product.brand or ""
    if key == "min_stock":
        return float(product.min_stock or 0)
    return ""


@login_required
@require_GET
def cabinet_products_export(request):
    """Barcha mahsulotlarni Excel (.xlsx) — ustunlar avtomatik kenglikda."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]

    raw_fields = [
        x.strip()
        for x in (request.GET.get("fields") or "name,barcode,cost_price,selling_price").split(",")
        if x.strip()
    ]
    if "name" not in raw_fields:
        raw_fields.insert(0, "name")
    # Takrorlarni olib tashlash, tartibni saqlash
    fields: list[str] = []
    seen = set()
    for f in raw_fields:
        if f in seen:
            continue
        seen.add(f)
        fields.append(f)

    try:
        products_raw = tezpos_api.get_catalog_snapshot(token, server, timeout=25) or []
        if len(products_raw) <= 200:
            products_raw = tezpos_api.get_all_products(token, server, timeout=25) or products_raw
        if len(products_raw) <= 200:
            products_raw = (
                tezpos_api.get_products(
                    token, server, max_pages=120, timeout=12, page_size=100
                )
                or products_raw
            )
        price_lists = tezpos_api.get_price_lists(token, server) or []
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"error": str(exc)}, status=502)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"error": str(exc)}, status=504)

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    price_lists_by_id = {
        str(pl.get("id")): pl
        for pl in price_lists
        if isinstance(pl, dict) and pl.get("id")
    }

    headers = []
    for key in fields:
        if key.startswith("pl_"):
            pl = price_lists_by_id.get(key[3:]) or {}
            headers.append(f"Narxlar: {pl.get('name') or key[3:]}")
        else:
            headers.append(_EXPORT_FIELD_LABELS.get(key, key))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({"error": "openpyxl o‘rnatilmagan"}, status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"
    ws.append(headers)
    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    barcode_col = fields.index("barcode") + 1 if "barcode" in fields else 0
    for col_idx in range(1, len(headers) + 1):
        ws.cell(1, col_idx).font = header_font

    for p in products:
        row = [_export_cell_value(p, key, price_lists_by_id) for key in fields]
        ws.append(row)
        r_idx = ws.max_row
        if barcode_col:
            cell = ws.cell(r_idx, barcode_col)
            raw = str(cell.value or "").replace("\r\n", "\n").replace("\r", "\n")
            cell.value = raw
            cell.alignment = wrap_top
            cell.number_format = "@"
            nlines = raw.count("\n") + 1 if raw else 1
            ws.row_dimensions[r_idx].height = min(220, max(18, 15 * nlines))
    if barcode_col:
        for cell in ws[get_column_letter(barcode_col)]:
            cell.alignment = wrap_top
            cell.number_format = "@"

    # Ustun kengligi — shtrixkodda faqat eng uzun qator
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            val = row[0].value
            if val is None:
                continue
            for line in str(val).splitlines() or [str(val)]:
                max_len = max(max_len, len(line))
        cap = 28 if col_idx == barcode_col else 80
        width = min(cap, max(14, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if col_idx == barcode_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 18)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = "barcha_mahsulotlar.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["Cache-Control"] = "no-store"
    return resp


def _cabinet_shell_context(request, tenant, section, *, form=None, sale_date=None):
    """Engil bo‘limlar (bot, qarzdorlar) — og‘ir API siz tezkor ochilish."""
    sale_date = sale_date or timezone.localdate()
    empty_chart = {"labels": [], "totals": [], "counts": []}
    return {
        "tenant": tenant,
        "form": form or _empty_form(),
        "products": [],
        "products_json": "[]",
        "price_lists_json": "[]",
        "shifts_json": "[]",
        "shifts_source": "none",
        "brand_choices": [],
        "category_choices": [],
        "all_products_count": 0,
        "form_errors": None,
        "section": section,
        "total_sales": 0,
        "gross": Decimal("0"),
        "cost": Decimal("0"),
        "profit": Decimal("0"),
        "margin": Decimal("0"),
        "today_count": 0,
        "today_gross": Decimal("0"),
        "today_profit": Decimal("0"),
        "range_checks": 0,
        "price_list_stats_json": "{}",
        "recent_sales": [],
        "sale_date": sale_date,
        "sale_date_iso": sale_date.isoformat(),
        "day_sales_count": 0,
        "day_gross": Decimal("0"),
        "day_cost_total": Decimal("0"),
        "day_profit_total": Decimal("0"),
        "day_pay_totals": {},
        "day_sales_json": "[]",
        "cashier_name": _cashier_name(request),
        "low_stock": [],
        "signals": [],
        "active_signals_count": 0,
        "near_min_json": "[]",
        "abc_rows": [],
        "abc_matrix": {},
        "abc_total": 0,
        "top_products_json": "[]",
        "top_customers_json": "[]",
        "chart_labels_json": "[]",
        "chart_totals_json": "[]",
        "chart_counts_json": "[]",
        "sales_stats_json": "{}",
        "share_items_json": "[]",
        "bot_settings": {
            "enabled": bool(tenant.telegram_enabled),
            "token": tenant.telegram_bot_token or "",
            "token_set": bool(tenant.telegram_bot_token),
            "recipients": tenant.telegram_recipients or "",
            "notify_open": bool(tenant.telegram_notify_open),
            "notify_close": bool(tenant.telegram_notify_close),
        },
        "bot_settings_json": json.dumps(
            {
                "enabled": bool(tenant.telegram_enabled),
                "token_set": bool(tenant.telegram_bot_token),
                "notify_open": bool(tenant.telegram_notify_open),
                "notify_close": bool(tenant.telegram_notify_close),
            },
            ensure_ascii=False,
        ),
        "label_template_json": _label_template_json(request),
        "report_charts_json": json.dumps(
            {
                "daily": empty_chart,
                "weekly": empty_chart,
                "monthly": empty_chart,
            },
            ensure_ascii=False,
        ),
        "fast_nav": True,
    }


@login_required
def cabinet_view(request):
    if not session_has_tezpos(request):
        clear_tezpos_session(request)
        messages.warning(request, "TezPOS hisobiga qayta kiring.")
        return redirect("login")

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    tenant = get_tenant_for_user(request.user)
    display = request.session.get(SESSION_DISPLAY) or tenant.business_name
    if display and tenant.business_name != display:
        tenant.business_name = display
        tenant.save(update_fields=["business_name"])

    section = request.GET.get("section", "overview")
    sale_date = _parse_sale_date(request.GET.get("sale_date"))
    form = _empty_form()
    section_force = None

    if request.method == "POST":
        product_id = (request.POST.get("product_id") or "").strip() or None
        ok, errors = _save_product_via_api(request, token, server, product_id)
        if ok:
            messages.success(request, "Mahsulot TezPOS ga saqlandi.")
            return redirect(f"{request.path}?section=products")
        form = _empty_form(errors)
        section_force = "products"

    section = section_force or request.GET.get("section", "overview")
    allowed = {
        "overview",
        "sales",
        "inventory",
        "products",
        "stock_value",
        "reports",
        "abc",
        "signals",
        "labels",
        "tops",
        "stock_in",
        "shifts",
        "bot",
        "suppliers",
        "client_debts",
    }
    if section not in allowed:
        section = "overview"

    # Engil bo‘limlar — TezPOS API kutmasdan darhol ochiladi (AJAX o‘zi yuklaydi)
    FAST_SHELL = {
        "overview",
        "tops",
        "stock_in",
        "bot",
        "suppliers",
        "client_debts",
        "products",
        "inventory",
        "stock_value",
        "labels",
        "signals",
        "sales",
        "shifts",
        "reports",
        "abc",
    }
    if section in FAST_SHELL and request.method != "POST":
        ctx = _cabinet_shell_context(
            request, tenant, section, form=form, sale_date=sale_date
        )
        ctx["fast_nav"] = True
        resp = render(request, "accounts/cabinet.html", ctx)
        resp["Cache-Control"] = "private, no-cache"
        return resp

    api_warnings = []
    today = timezone.localdate()
    raw_products: list = []
    raw_sales: list = []
    day_sales_raw: list = []

    # Og‘ir SSR faqat kerakli joylarda
    need_products = section in (
        "products",
        "inventory",
        "stock_value",
        "labels",
        "signals",
        "abc",
    )
    need_chart_sales = section in ("abc", "reports")
    need_day_sales = section == "sales"
    # Chek detali SSR da emas — jadval tez ochilsin (foyda taxminiy)
    need_sale_details = False
    need_price_list_stats = False
    need_price_lists = section in (
        "products",
        "inventory",
        "stock_value",
    )
    need_top_stats = section == "abc"
    need_shifts = section == "shifts"
    # all=true ko‘pincha 1 so‘rov — ortiqcha sahifa kutmaslik
    if section in ("products", "inventory", "stock_value", "labels"):
        product_pages = 4
    elif section in ("signals", "sales", "abc"):
        product_pages = 2
    else:
        product_pages = 1

    def _load_products():
        key = f"products:{server}:{product_pages}"
        return _memo_get(
            key,
            90.0,
            lambda: tezpos_api.get_products(token, server, max_pages=product_pages),
        )

    def _load_chart_sales():
        if section == "reports":
            days, pages, timeout = 30, 40, 20
        elif section == "abc":
            days, pages, timeout = 14, 40, 18
        else:
            days, pages, timeout = 7, 40, 16
        key = f"sales:{server}:{days}:{pages}"

        def _loader():
            try:
                return tezpos_api.get_sales(
                    token,
                    server,
                    date_from=(today - timedelta(days=days)).isoformat(),
                    date_to=today.isoformat(),
                    timeout=timeout,
                    max_pages=pages,
                )
            except (tezpos_api.TezPosApiError, TimeoutError, OSError) as exc:
                if getattr(exc, "status", None) in (401, 403):
                    raise
                return tezpos_api.get_sales(
                    token,
                    server,
                    date_from=(today - timedelta(days=2)).isoformat(),
                    date_to=today.isoformat(),
                    timeout=12,
                    max_pages=2,
                )

        return _memo_get(key, 60.0, _loader)

    def _load_price_lists():
        key = f"price_lists:{server}"

        def _loader():
            try:
                return tezpos_api.get_price_lists(token, server)
            except (tezpos_api.TezPosApiError, TimeoutError, OSError):
                return []

        return _memo_get(key, 120.0, _loader)

    def _load_shifts_api():
        key = f"shifts:{server}:{today.isoformat()}"

        def _loader():
            try:
                return tezpos_api.get_shifts(
                    token,
                    server,
                    date_from=(today - timedelta(days=14)).isoformat(),
                    date_to=today.isoformat(),
                    timeout=12,
                    max_pages=20,
                )
            except (tezpos_api.TezPosApiError, TimeoutError, OSError):
                return []

        return _memo_get(key, 45.0, _loader)
    def _auth_fail(exc):
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            messages.error(request, "Sessiya tugadi. Qayta kiring.")
            return True
        return False

    raw_price_lists: list = []
    raw_shifts_api: list = []
    # Mahsulotlar + grafik sotuvlari + kunlik sotuvlar + narxlar ro'yxati parallel
    try:
        with ThreadPoolExecutor(max_workers=5) as pool:
            fut_p = pool.submit(_load_products) if need_products else None
            fut_s = pool.submit(_load_chart_sales) if need_chart_sales else None
            fut_d = (
                pool.submit(tezpos_api.get_sales_for_day, token, server, sale_date.isoformat())
                if need_day_sales
                else None
            )
            fut_pl = pool.submit(_load_price_lists) if need_price_lists else None
            fut_sh = pool.submit(_load_shifts_api) if need_shifts else None
            if fut_p is not None:
                try:
                    raw_products = fut_p.result()
                except (tezpos_api.TezPosApiError, TimeoutError, OSError) as exc:
                    if _auth_fail(exc):
                        return redirect("login")
                    api_warnings.append(f"Mahsulotlar: {exc}")
                    raw_products = []
            if fut_s is not None:
                try:
                    raw_sales = fut_s.result()
                except (tezpos_api.TezPosApiError, TimeoutError, OSError) as exc:
                    if _auth_fail(exc):
                        return redirect("login")
                    api_warnings.append(f"Sotuvlar: {exc}")
                    raw_sales = []
            if fut_d is not None:
                try:
                    day_sales_raw = fut_d.result()
                except (tezpos_api.TezPosApiError, TimeoutError, OSError) as exc:
                    if _auth_fail(exc):
                        return redirect("login")
                    day_sales_raw = [
                        s
                        for s in raw_sales
                        if isinstance(s, dict) and _sale_day(s) == sale_date
                    ]
            if fut_pl is not None:
                raw_price_lists = fut_pl.result() or []
            if fut_sh is not None:
                raw_shifts_api = fut_sh.result() or []
    except Exception as exc:
        api_warnings.append(str(exc))

    if need_day_sales and not day_sales_raw:
        # Bugun sotuv yo'q — oxirgi sotuv kuniga (faqat default sana)
        if not request.GET.get("sale_date") and section in ("overview", "sales"):
            latest = None
            for s in raw_sales:
                if not isinstance(s, dict):
                    continue
                d = _sale_day(s)
                if d and (latest is None or d > latest):
                    latest = d
            if latest and latest != sale_date:
                sale_date = latest
                try:
                    day_sales_raw = tezpos_api.get_sales_for_day(
                        token, server, sale_date.isoformat()
                    )
                except (tezpos_api.TezPosApiError, TimeoutError, OSError):
                    day_sales_raw = [
                        s
                        for s in raw_sales
                        if isinstance(s, dict) and _sale_day(s) == sale_date
                    ]
    if api_warnings:
        clean = []
        for w in api_warnings[:2]:
            s = " ".join(str(w).split())
            if "<!doctype" in s.lower() or "<html" in s.lower():
                s = "TezPOS API dan ma’lumot olinmadi (noto‘g‘ri API manzili yoki endpoint)."
            clean.append(s[:180])
        messages.warning(request, " · ".join(clean))

    products = [_map_product(p) for p in raw_products if isinstance(p, dict)]
    # Faol mahsulotlarni birinchi ko'rsatish
    products.sort(key=lambda p: (not p.is_active, p.name.lower()))
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    price_lists = [pl for pl in raw_price_lists if isinstance(pl, dict) and pl.get("is_active", True)]

    brand_choices = sorted({p.brand for p in products if p.brand})
    category_choices = sorted({p.category for p in products if p.category})

    day_sales_raw = [s for s in day_sales_raw if isinstance(s, dict)]
    day_sales_raw.sort(
        key=lambda s: _parse_dt(s.get("completed_at") or s.get("created_at"))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )

    # Detail — faqat kunlik sotuvlar (limit bilan, tez ochilishi uchun)
    need_detail_ids = []
    detail_limit = 30
    if need_sale_details:
        need_detail_ids = [str(s.get("id")) for s in day_sales_raw if s.get("id")]
        detail_limit = min(max(len(need_detail_ids), 1), 35)
    details = (
        _fetch_sale_details(token, server, need_detail_ids, limit=detail_limit)
        if need_detail_ids
        else {}
    )
    cashier = _cashier_name(request)
    margin_ratio = _catalog_margin_ratio(products)

    day_sales_list = []
    day_sales_payload = []
    day_gross = Decimal("0")
    day_cost_total = Decimal("0")
    day_pay_totals = defaultdict(Decimal)
    for s in day_sales_raw:
        sid = str(s.get("id"))
        detail = details.get(sid) or s
        total = _dec(detail.get("total") or s.get("total"))
        cost, profit = _estimate_sale_profit(
            detail, products_by_id, total, products_by_name, margin_ratio
        )
        method = detail.get("payment_type") or s.get("payment_type") or "cash"
        dt = _parse_dt(
            detail.get("completed_at") or detail.get("created_at") or s.get("created_at")
        )
        row = SimpleNamespace(
            id=sid,
            created_at=dt or timezone.now(),
            customer_name=detail.get("customer_name") or s.get("customer_name") or "",
            total_amount=total,
            total_cost=cost,
            profit=profit,
            payment_method=method,
        )
        day_sales_list.append(row)
        day_gross += total
        day_cost_total += cost
        day_pay_totals[_payment_label(method)] += total
        day_sales_payload.append(
            _serialize_sale_payload(
                detail if detail.get("items") is not None else {**s, "items": []},
                cashier,
                products_by_id,
                products_by_name,
                margin_ratio,
            )
        )

    day_profit_total = day_gross - day_cost_total

    # Tannarx topilmagan cheklar — katalog marjasi (foyda = o'rtadagi summa)
    if day_gross > 0 and margin_ratio > 0:
        for i, row in enumerate(day_sales_list):
            if row.total_amount > 0 and row.total_cost <= 0:
                row.profit = (row.total_amount * margin_ratio).quantize(Decimal("0.01"))
                row.total_cost = row.total_amount - row.profit
                if i < len(day_sales_payload):
                    day_sales_payload[i]["profit"] = float(row.profit)
                    day_sales_payload[i]["total_cost"] = float(row.total_cost)
        day_cost_total = sum((r.total_cost for r in day_sales_list), Decimal("0"))
        day_profit_total = day_gross - day_cost_total

    if section == "sales" and request.GET.get("export") == "excel":
        return _export_daily_sales_excel(day_sales_payload, sale_date, cashier)

    # Top mahsulotlar — API stats; bo'sh bo'lsa oxirgi cheklar detali
    product_qty = defaultdict(Decimal)
    product_rev = defaultdict(Decimal)
    product_meta: dict[str, dict] = {}
    item_rows = []
    top_payload = {"items": []}
    if need_top_stats:
        try:
            top_payload = tezpos_api.get_top_products(token, server, days=30, limit=100)
        except (tezpos_api.TezPosApiError, TimeoutError, OSError):
            top_payload = {"items": []}

    for row in top_payload.get("items") or []:
        if not isinstance(row, dict):
            continue
        pid = str(row.get("product_id") or row.get("id") or "")
        if not pid:
            continue
        qty = _dec(row.get("quantity") or row.get("qty") or row.get("sold_qty"))
        rev = _dec(
            row.get("revenue")
            or row.get("total")
            or row.get("amount")
            or row.get("sum")
        )
        p = products_by_id.get(pid)
        if rev <= 0 and qty > 0:
            unit = _dec(row.get("unit_price") or row.get("price"))
            if unit <= 0 and p:
                unit = p.selling_price
            rev = qty * unit if unit > 0 else Decimal("0")
        if rev <= 0 and p and qty > 0:
            rev = qty * p.selling_price
        if rev <= 0 and qty <= 0:
            continue
        product_qty[pid] = qty
        product_rev[pid] = rev
        name = (
            (row.get("product_name") or row.get("name") or "")
            or (p.name if p else "")
            or f"Mahsulot {pid[:8]}"
        )
        product_meta[pid] = {
            "name": name,
            "image": (p.display_image if p else "") or str(row.get("image") or ""),
            "stock": float(p.stock_qty) if p else float(_dec(row.get("stock") or row.get("quantity_left"))),
            "wholesale": float(p.wholesale_price or p.cost_price) if p else float(_dec(row.get("wholesale_price"))),
            "selling": float(p.selling_price) if p else float(_dec(row.get("selling_price") or row.get("unit_price"))),
        }
        item_rows.append(
            {
                "product_id": pid,
                "quantity": qty,
                "unit_price": (rev / qty) if qty else Decimal("0"),
                "day": today,
                "name": name,
            }
        )

    if need_top_stats and not product_rev:
        recent_ids = []
        for s in raw_sales:
            if not isinstance(s, dict) or not s.get("id"):
                continue
            sid = str(s["id"])
            if sid not in details:
                recent_ids.append(sid)
            if len(recent_ids) >= 40:
                break
        if recent_ids:
            details.update(_fetch_sale_details(token, server, recent_ids, limit=40))
        for detail in details.values():
            d = _sale_day(detail)
            for item in detail.get("items") or []:
                pid = str(item.get("product_id") or "")
                qty = _dec(item.get("quantity"))
                unit_price = _dec(item.get("unit_price"))
                name = (item.get("product_name") or item.get("name") or "").strip()
                if not pid:
                    if not name:
                        continue
                    pid = f"name:{name.casefold()}"
                line_rev = qty * unit_price
                product_qty[pid] += qty
                product_rev[pid] += line_rev
                p = products_by_id.get(pid) if not pid.startswith("name:") else None
                if not p and name:
                    p = _find_product(products_by_id, products_by_name, product_name=name)
                meta = product_meta.get(pid) or {}
                product_meta[pid] = {
                    "name": name or meta.get("name") or (p.name if p else "Mahsulot"),
                    "image": (p.display_image if p else "") or meta.get("image") or "",
                    "stock": float(p.stock_qty) if p else float(meta.get("stock") or 0),
                    "wholesale": float(p.wholesale_price or p.cost_price)
                    if p
                    else float(meta.get("wholesale") or 0),
                    "selling": float(p.selling_price)
                    if p
                    else float(meta.get("selling") or unit_price),
                }
                item_rows.append(
                    {
                        "product_id": pid,
                        "quantity": qty,
                        "unit_price": unit_price,
                        "day": d,
                        "name": product_meta[pid]["name"],
                    }
                )

    sales_stats = _build_charts_from_sales(
        [s for s in raw_sales if isinstance(s, dict)], today
    )
    margin_ratio = _catalog_margin_ratio(products)
    _attach_range_summaries(
        sales_stats,
        [s for s in raw_sales if isinstance(s, dict)],
        today,
        margin_ratio,
    )
    if details:
        _refine_summaries_from_details(
            sales_stats,
            [s for s in raw_sales if isinstance(s, dict)],
            details,
            today,
            products_by_id,
            products_by_name,
        )
    price_list_stats = (
        _build_price_list_stats_by_range(
            [s for s in raw_sales if isinstance(s, dict)],
            details,
            today,
            products_by_id,
            products_by_name,
            price_lists,
            range_keys=["d1", "d7"],
        )
        if need_price_list_stats
        else {}
    )
    chart_labels = sales_stats["d7"]["labels"]
    chart_totals = sales_stats["d7"]["totals"]
    chart_counts = sales_stats["d7"]["counts"]
    d_pack, w_pack, m_pack = sales_stats["d7"], sales_stats["m3"], sales_stats["m6"]

    # Default KPI — tanlangan davr (7 kun)
    d7_summary = (sales_stats.get("d7") or {}).get("summary") or {}
    gross = sum((_dec(s.get("total")) for s in raw_sales if isinstance(s, dict)), Decimal("0"))
    range_checks = int(d7_summary.get("checks") or 0)
    range_gross = Decimal(str(d7_summary.get("gross") or 0))
    range_profit = Decimal(str(d7_summary.get("profit") or 0))
    range_margin = Decimal(str(d7_summary.get("margin") or 0))

    # Bugungi ko'rsatkichlar + sof foyda (sotuv - tannarx)
    if sale_date == today:
        today_count = len(day_sales_list)
        today_gross = day_gross
        today_profit = day_profit_total
        cost = day_cost_total
        profit = day_profit_total
    else:
        today_sales = [s for s in raw_sales if isinstance(s, dict) and _sale_day(s) == today]
        today_count = len(today_sales)
        today_gross = sum((_dec(s.get("total")) for s in today_sales), Decimal("0"))
        today_profit = today_gross  # detail yo'q bo'lsa taxminiy
        cost = day_cost_total
        profit = day_profit_total

    # Overview KPI: tanlangan davr (7 kun default)
    if section == "overview":
        if range_checks > 0 or range_gross > 0:
            profit = range_profit
            cost = range_gross - range_profit
            margin = range_margin
            gross = range_gross
        else:
            margin = Decimal("0")
            profit = Decimal("0")
            cost = Decimal("0")
    else:
        margin = (profit / day_gross * 100) if day_gross > 0 else Decimal("0")
        if not day_sales_list:
            cost = Decimal("0")
            profit = Decimal("0")
            margin = Decimal("0")
            day_profit_total = Decimal("0")
    share_items = []
    for pid, rev in sorted(product_rev.items(), key=lambda x: x[1], reverse=True)[:200]:
        p = products_by_id.get(pid)
        if p:
            share_items.append(
                {
                    "name": p.name,
                    "image": p.display_image,
                    "stock": float(p.stock_qty),
                    "wholesale": float(p.wholesale_price or p.cost_price),
                    "selling": float(p.selling_price),
                    "revenue": float(rev),
                }
            )
    if not share_items:
        # Mahsulotlar bo'yicha ombor qiymati (sotuv bo'lmasa ham to'ldirish)
        for p in sorted(products, key=lambda x: x.selling_price * max(x.stock_qty, 0), reverse=True)[:50]:
            share_items.append(
                {
                    "name": p.name,
                    "image": p.display_image,
                    "stock": float(p.stock_qty),
                    "wholesale": float(p.wholesale_price or p.cost_price),
                    "selling": float(p.selling_price),
                    "revenue": float(p.selling_price * max(p.stock_qty, Decimal("0"))),
                }
            )
    if not share_items:
        share_items = [
            {
                "name": "Hozircha ma'lumot yo'q",
                "image": "",
                "stock": 0,
                "wholesale": 0,
                "selling": 0,
                "revenue": 1,
            }
        ]

    top_products = []
    for pid, rev in sorted(product_rev.items(), key=lambda x: x[1], reverse=True)[:100]:
        p = products_by_id.get(pid)
        meta = product_meta.get(pid) or {}
        if not p and not meta:
            continue
        top_products.append(
            {
                "name": (p.name if p else "") or meta.get("name") or "Mahsulot",
                "image": (p.display_image if p else "") or meta.get("image") or "",
                "qty": float(product_qty[pid]),
                "revenue": float(rev),
                "stock": float(p.stock_qty) if p else float(meta.get("stock") or 0),
                "wholesale": float(p.wholesale_price or p.cost_price)
                if p
                else float(meta.get("wholesale") or 0),
                "selling": float(p.selling_price) if p else float(meta.get("selling") or 0),
            }
        )

    products_payload = [
        {
            "id": p.id,
            "name": p.name,
            "barcode": p.barcode or "",
            "barcodes": p.barcode_list,
            "unit": p.unit or "dona",
            "category": p.category or "",
            "brand": p.brand or "",
            "selling_price": float(p.selling_price),
            "wholesale_price": float(p.wholesale_price or 0),
            "cost_price": float(p.cost_price or 0),
            "list_prices": {
                str(k): float(v) for k, v in (p.list_prices or {}).items()
            },
            "stock_qty": float(p.stock_qty or 0),
            "min_stock": float(p.min_stock or 0),
            "is_favorite": bool(p.is_favorite),
            "image": p.display_image,
            "image_url": p.image_url or "",
            "images": [
                {"id": img.id, "url": img.image.url, "is_primary": img.is_primary}
                for img in (p.images.all() if hasattr(p.images, "all") else [])
            ],
        }
        for p in products
    ]

    price_lists_payload = [
        {
            "id": str(pl.get("id") or ""),
            "name": (pl.get("name") or "").strip() or "Narxlar",
            "is_selling": bool(pl.get("is_selling")),
        }
        for pl in price_lists
        if str(pl.get("id") or "")
    ]

    # Smenalar — TezPOS API yoki sotuvlardan
    shifts_payload: list[dict] = []
    shifts_source = "none"
    if need_shifts:
        sales_for_shifts = [s for s in raw_sales if isinstance(s, dict)]
        if raw_shifts_api:
            shifts_source = "api"
            for raw in raw_shifts_api:
                if not isinstance(raw, dict):
                    continue
                sh = _normalize_api_shift(raw, today)
                sh = _enrich_shift_with_sales(
                    sh, sales_for_shifts, margin_ratio=margin_ratio
                )
                sh.pop("raw", None)
                shifts_payload.append(sh)
        if not shifts_payload:
            shifts_source = "sales"
            shifts_payload = _build_shifts_from_sales(
                sales_for_shifts,
                today=today,
                gap_hours=4.0,
                margin_ratio=margin_ratio,
            )
        # Narxlar ro‘yxati AJAX (shift-detail) orqali — SSR da 8×detail sekin

    cust_agg = defaultdict(lambda: {"orders": 0, "total": Decimal("0")})
    for s in raw_sales:
        if not isinstance(s, dict):
            continue
        name = (s.get("customer_name") or "").strip()
        if not name:
            continue
        cust_agg[name]["orders"] += 1
        cust_agg[name]["total"] += _dec(s.get("total"))
    customer_rows = [
        {"customer_name": n, "orders": v["orders"], "total": float(v["total"])}
        for n, v in sorted(cust_agg.items(), key=lambda x: x[1]["total"], reverse=True)[:100]
    ]

    abc_rows, abc_matrix, abc_total = _build_abc_xyz(products, item_rows, today)
    near_min = _build_near_min_stock(products)
    signals = near_min
    low_stock = near_min[:8]

    section = section_force or request.GET.get("section", "overview")
    allowed = {
        "overview",
        "sales",
        "inventory",
        "products",
        "stock_value",
        "reports",
        "abc",
        "signals",
        "labels",
        "tops",
        "shifts",
        "bot",
        "suppliers",
        "client_debts",
        "stock_in",
    }
    if section not in allowed:
        section = "overview"

    ctx = {
            "tenant": tenant,
            "form": form,
            "products": products,
            "products_json": json.dumps(products_payload, ensure_ascii=False),
            "price_lists_json": json.dumps(price_lists_payload, ensure_ascii=False),
            "shifts_json": json.dumps(shifts_payload, ensure_ascii=False),
            "shifts_source": shifts_source,
            "brand_choices": brand_choices,
            "category_choices": category_choices,
            "all_products_count": len(products),
            "form_errors": form.errors if request.method == "POST" else None,
            "section": section,
            "total_sales": range_checks if section == "overview" else len(raw_sales),
            "gross": gross,
            "cost": cost,
            "profit": profit,
            "margin": margin,
            "today_count": today_count,
            "today_gross": today_gross,
            "today_profit": today_profit,
            "range_checks": range_checks,
            "price_list_stats_json": json.dumps(price_list_stats, ensure_ascii=False),
            "recent_sales": day_sales_list,
            "sale_date": sale_date,
            "sale_date_iso": sale_date.isoformat(),
            "day_sales_count": len(day_sales_list),
            "day_gross": day_gross,
            "day_cost_total": day_cost_total,
            "day_profit_total": day_profit_total,
            "day_pay_totals": dict(day_pay_totals),
            "day_sales_json": json.dumps(day_sales_payload, ensure_ascii=False),
            "cashier_name": cashier,
            "low_stock": low_stock,
            "signals": signals,
            "active_signals_count": len(signals),
            "near_min_json": json.dumps(signals, ensure_ascii=False),
            "abc_rows": abc_rows,
            "abc_matrix": abc_matrix,
            "abc_total": abc_total,
            "top_products_json": json.dumps(top_products, ensure_ascii=False),
            "top_customers_json": json.dumps(customer_rows, ensure_ascii=False),
            "chart_labels_json": json.dumps(chart_labels, ensure_ascii=False),
            "chart_totals_json": json.dumps(chart_totals),
            "chart_counts_json": json.dumps(chart_counts),
            "sales_stats_json": json.dumps(sales_stats, ensure_ascii=False),
            "share_items_json": json.dumps(share_items, ensure_ascii=False),
            "bot_settings": {
                "enabled": bool(tenant.telegram_enabled),
                "token": tenant.telegram_bot_token or "",
                "token_set": bool(tenant.telegram_bot_token),
                "recipients": tenant.telegram_recipients or "",
                "notify_open": bool(tenant.telegram_notify_open),
                "notify_close": bool(tenant.telegram_notify_close),
            },
            "bot_settings_json": json.dumps(
                {
                    "enabled": bool(tenant.telegram_enabled),
                    "token_set": bool(tenant.telegram_bot_token),
                    "notify_open": bool(tenant.telegram_notify_open),
                    "notify_close": bool(tenant.telegram_notify_close),
                },
                ensure_ascii=False,
            ),
            "label_template_json": _label_template_json(request),
            "report_charts_json": json.dumps(
                {
                    "daily": {
                        "labels": d_pack["labels"],
                        "totals": d_pack["totals"],
                        "counts": d_pack["counts"],
                    },
                    "weekly": {
                        "labels": w_pack["labels"],
                        "totals": w_pack["totals"],
                        "counts": w_pack["counts"],
                    },
                    "monthly": {
                        "labels": m_pack["labels"],
                        "totals": m_pack["totals"],
                        "counts": m_pack["counts"],
                    },
                },
                ensure_ascii=False,
            ),
            "fast_nav": True,
        }
    resp = render(request, "accounts/cabinet.html", ctx)
    # Brauzer navigatsiyasini biroz tezlatish (shaxsiy kabinet — kesh emas)
    resp["Cache-Control"] = "private, no-cache"
    return resp


@login_required
@require_GET
def cabinet_range_stats(request):
    """Tanlangan kun/oraliq: grafik + KPI + narxlar ro'yxati (jonli AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    today = timezone.localdate()
    t_start = time.time()

    custom_from = _parse_iso_date(request.GET.get("from"))
    custom_to = _parse_iso_date(request.GET.get("to"))
    range_key = (request.GET.get("range") or "").strip()

    if custom_from and custom_to:
        start, end = custom_from, custom_to
        if end < start:
            start, end = end, start
        # Max 2 yil
        if (end - start).days > 730:
            start = end - timedelta(days=730)
        range_key = f"custom:{start.isoformat()}:{end.isoformat()}"
    else:
        allowed = {"d1", "d7", "d15", "d30", "m1", "m3", "m6", "y1"}
        if range_key not in allowed:
            range_key = "d7"
        start, end = _range_window(range_key, today)

    span = (end - start).days + 1
    fast = (request.GET.get("fast") or "").strip() in ("1", "true", "yes")
    # fast: faqat jami (tez). To‘liq: Optom uchun chek + katalog namuna (deadline ichida).
    single_day = span <= 1
    if fast:
        # Jami to‘liq (barcha cheklar). Foyda — chek qatorlari + katalog tannarxi.
        max_pages = 80 if single_day else (40 if span <= 7 else 25)
        sales_timeout = 16 if single_day else 18
        detail_cap, detail_each, detail_budget = 0, 0.0, 0.0
        hard_deadline = 22.0
    elif single_day:
        max_pages, sales_timeout = 80, 16
        detail_cap, detail_each, detail_budget = 80, 2.2, 10.0
        hard_deadline = 24.0
    elif span <= 7:
        max_pages, sales_timeout = 50, 16
        detail_cap, detail_each, detail_budget = 40, 2.2, 9.0
        hard_deadline = 22.0
    elif span <= 31:
        max_pages, sales_timeout = 40, 18
        detail_cap, detail_each, detail_budget = 30, 2.2, 8.0
        hard_deadline = 22.0
    else:
        max_pages, sales_timeout = 30, 18
        detail_cap, detail_each, detail_budget = 24, 2.0, 8.0
        hard_deadline = 24.0

    memo_prefix = f"{server}|{(token or '')[-12:]}"
    sales: list = []
    products_raw: list = []
    price_lists: list = []
    api_err = ""

    def _load_sales():
        if single_day:
            day = start.isoformat()
            return _memo_get(
                f"{memo_prefix}|dayv3|{day}",
                90.0,
                lambda: tezpos_api.get_sales_for_day(token, server, day) or [],
            )
        return _memo_get(
            f"{memo_prefix}|salesv3|{start}|{end}|{max_pages}",
            60.0,
            lambda: tezpos_api.get_sales(
                token,
                server,
                date_from=start.isoformat(),
                date_to=end.isoformat(),
                timeout=sales_timeout,
                max_pages=max_pages,
            )
            or [],
        )

    def _load_price_lists():
        return _memo_get(
            f"{memo_prefix}|price_lists",
            600.0,
            lambda: tezpos_api.get_price_lists(token, server) or [],
        )

    def _load_products():
        return _memo_get(
            f"{memo_prefix}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=16) or [],
        )

    try:
        workers = 3
        with ThreadPoolExecutor(max_workers=workers) as pool:
            fut_s = pool.submit(_load_sales)
            fut_pl = pool.submit(_load_price_lists)
            fut_p = pool.submit(_load_products)
            remain = max(1.0, hard_deadline - (time.time() - t_start))
            try:
                sales = fut_s.result(timeout=remain) or []
            except FuturesTimeoutError:
                api_err = api_err or "Savdo so‘rovi vaqt limiti"
                sales = []
            remain = max(0.5, hard_deadline - (time.time() - t_start))
            try:
                price_lists = fut_pl.result(timeout=remain) or []
            except FuturesTimeoutError:
                price_lists = []
            except Exception:
                price_lists = []
            if fut_p is not None:
                remain = max(0.5, hard_deadline - (time.time() - t_start))
                try:
                    products_raw = fut_p.result(timeout=remain) or []
                except (FuturesTimeoutError, Exception):
                    products_raw = []
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        api_err = str(exc)
    except (TimeoutError, OSError) as exc:
        api_err = str(exc)

    if not products_raw:
        for key in (
            f"{memo_prefix}|catalog_snap",
            f"{memo_prefix}|products|4",
            f"{memo_prefix}|products|3",
            f"{memo_prefix}|products|2",
        ):
            hit = _TEZPOS_MEMO.get(key)
            if hit and isinstance(hit[1], list) and hit[1]:
                products_raw = hit[1]
                break

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    margin_ratio = _catalog_margin_ratio(products) if products else Decimal("0")
    price_lists = [
        pl for pl in price_lists if isinstance(pl, dict) and pl.get("is_active", True)
    ]

    sales = [s for s in sales if isinstance(s, dict)]
    sales.sort(
        key=lambda s: _parse_dt(s.get("completed_at") or s.get("created_at"))
        or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    sales = [
        s
        for s in sales
        if (d := _sale_day(s)) is not None and start <= d <= end
    ]
    chart = _chart_pack_for_dates(sales, start, end)
    checks = len(sales)
    gross = float(sum((_dec(s.get("total")) for s in sales), Decimal("0")))
    # Desktop /api/sales/stats/daily/ — bugungi jami list kesilgan bo‘lsa ham to‘g‘ri
    if single_day and start == today:
        try:
            daily = tezpos_api.get_daily_stats(token, server) or {}
        except (tezpos_api.TezPosApiError, TimeoutError, OSError, TypeError, ValueError):
            daily = {}
        if isinstance(daily, dict):
            try:
                api_count = int(daily.get("sales_count") or 0)
            except (TypeError, ValueError):
                api_count = 0
            api_rev = daily.get("total_revenue")
            if api_count > checks:
                checks = api_count
            if api_rev is not None:
                api_gross = float(_dec(api_rev))
                if api_gross > gross:
                    gross = api_gross
    profit = 0.0
    margin = 0.0

    lists: list = []
    details_used = 0
    estimated = False
    remain_budget = hard_deadline - (time.time() - t_start)
    if gross > 0:
        details: dict[str, dict] = {}
        for s in sales:
            if not isinstance(s, dict) or not s.get("id"):
                continue
            if _sale_items(s):
                details[str(s.get("id"))] = s
        if detail_cap > 0 and remain_budget >= 3.0:
            need_ids = [
                str(s.get("id"))
                for s in sales
                if s.get("id") and str(s.get("id")) not in details
            ]
            if len(need_ids) > detail_cap:
                step = max(1, len(need_ids) // detail_cap)
                sampled = need_ids[::step][:detail_cap]
                if len(sampled) < detail_cap:
                    for sid in need_ids:
                        if sid not in sampled:
                            sampled.append(sid)
                        if len(sampled) >= detail_cap:
                            break
                need_ids = sampled
            sample_n = min(detail_cap, len(need_ids))
            if sample_n > 0:
                fetched = _fetch_sale_details(
                    token,
                    server,
                    need_ids[:sample_n],
                    limit=sample_n,
                    per_sale_timeout=detail_each,
                    overall_timeout=min(detail_budget, max(2.0, remain_budget - 1.0)),
                )
                details.update(fetched)
        details_used = len(details)
        if details:
            refined = _aggregate_price_list_stats(
                list(details.values()),
                products_by_id,
                products_by_name,
                price_lists,
            )
            if refined:
                lists = _scale_price_list_stats(refined, gross, checks)
                estimated = any(
                    isinstance(r, dict) and r.get("scaled") and not r.get("is_total")
                    for r in lists
                )
                jami = next((r for r in lists if r.get("is_total")), None)
                if jami:
                    profit = float(jami.get("profit") or 0)
                    margin = float(
                        jami.get("markup")
                        if jami.get("markup") is not None
                        else (jami.get("margin") or 0)
                    )
        if not lists:
            lists = _fallback_sotuv_only_lists(gross, checks, margin_ratio, price_lists)
            estimated = True

    has_optom = any(
        isinstance(r, dict)
        and not r.get("is_total")
        and str(r.get("id")) != SELLING_LIST_ID
        and float(r.get("revenue") or 0) > 0
        for r in lists
    )
    payload = {
        "range": range_key,
        "from": start.isoformat(),
        "to": end.isoformat(),
        "summary": {
            "checks": checks,
            "gross": gross,
            "profit": profit,
            "margin": margin,
            "details_used": details_used,
            "products_used": len(products),
        },
        "chart": chart,
        "priceLists": lists,
        "partial": bool(
            estimated
            or (detail_cap > 0 and details_used < max(1, min(checks, detail_cap)))
            or bool(api_err)
        ),
        "estimated": estimated,
        "fast": bool(fast),
        "ms": int((time.time() - t_start) * 1000),
        "api": tezpos_api.normalize_api_base(),
    }
    if api_err:
        payload["error"] = api_err
    _log_slow(
        "range-stats",
        t_start,
        f"span={span} checks={checks} optom={has_optom} details={details_used}",
    )
    # Har doim 200 — brauzer "API timeout" catch ga tushmasin
    return JsonResponse(payload)


def _top_products_from_details(
    details: dict,
    products_by_id: dict,
    products_by_name: dict,
    limit: int = 100,
    price_lists: list[dict] | None = None,
) -> list[dict]:
    price_lists = price_lists or []
    selling_list_ids = {
        str(pl.get("id"))
        for pl in price_lists
        if pl.get("id") and _is_api_selling_list(pl)
    }
    product_qty: dict[str, Decimal] = defaultdict(Decimal)
    product_rev: dict[str, Decimal] = defaultdict(Decimal)
    product_cost: dict[str, Decimal] = defaultdict(Decimal)
    product_costed_rev: dict[str, Decimal] = defaultdict(Decimal)
    product_qty_sell: dict[str, Decimal] = defaultdict(Decimal)
    product_qty_optom: dict[str, Decimal] = defaultdict(Decimal)
    product_rev_sell: dict[str, Decimal] = defaultdict(Decimal)
    product_rev_optom: dict[str, Decimal] = defaultdict(Decimal)
    product_meta: dict[str, dict] = {}

    for detail in details.values():
        if not isinstance(detail, dict):
            continue
        sale_pl = (
            detail.get("price_list_id")
            or detail.get("price_list")
            or detail.get("list_id")
        )
        if isinstance(sale_pl, dict):
            sale_pl = sale_pl.get("id")
        for item in _sale_items(detail):
            pid, name, _nested = _item_product_ref(item)
            qty = _item_qty(item)
            unit_price = _item_unit_price(item)
            line_rev = _dec(
                item.get("total") or item.get("line_total"), str(qty * unit_price)
            )
            if unit_price <= 0 and qty > 0 and line_rev > 0:
                unit_price = (line_rev / qty).quantize(Decimal("0.01"))
            if not pid:
                name = (name or "").strip()
                if not name:
                    continue
                pid = f"name:{name.casefold()}"
            if qty <= 0 and line_rev <= 0:
                continue
            p = products_by_id.get(pid) if not str(pid).startswith("name:") else None
            if not p and name:
                p = _find_product(products_by_id, products_by_name, product_name=name)

            unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
            raw_pl = (
                item.get("price_list_id")
                or item.get("price_list")
                or item.get("list_id")
                or sale_pl
            )
            if isinstance(raw_pl, dict):
                raw_pl = raw_pl.get("id")
            list_id = str(raw_pl).strip() if raw_pl not in (None, "") else ""
            if list_id in ("selling", "retail", SELLING_LIST_ID) or list_id in selling_list_ids:
                list_id = SELLING_LIST_ID
            elif not list_id:
                list_id = _match_price_list_id(unit_price, p, price_lists)
            is_selling = list_id == SELLING_LIST_ID or list_id in selling_list_ids

            product_qty[pid] += qty
            product_rev[pid] += line_rev
            if unit_cost > 0 and qty > 0:
                product_cost[pid] += unit_cost * qty
                product_costed_rev[pid] += line_rev
            if is_selling:
                product_qty_sell[pid] += qty
                product_rev_sell[pid] += line_rev
            else:
                product_qty_optom[pid] += qty
                product_rev_optom[pid] += line_rev

            wholesale = Decimal("0")
            selling = Decimal("0")
            cost_show = unit_cost
            if p:
                selling = Decimal(str(p.selling_price or 0))
                wholesale = Decimal(str(p.wholesale_price or 0))
                if wholesale <= 0:
                    list_prices = getattr(p, "list_prices", None) or {}
                    vals = [
                        Decimal(str(v))
                        for lid, v in list_prices.items()
                        if lid not in selling_list_ids and Decimal(str(v or 0)) > 0
                    ]
                    if vals:
                        wholesale = min(vals)
                if cost_show <= 0 and (p.cost_price or Decimal("0")) > 0:
                    cost_show = p.cost_price
            meta = product_meta.get(pid) or {}
            product_meta[pid] = {
                "name": name or meta.get("name") or (p.name if p else "Mahsulot"),
                "image": (p.display_image if p else "") or meta.get("image") or "",
                "stock": float(p.stock_qty) if p else float(meta.get("stock") or 0),
                "wholesale": float(wholesale or meta.get("wholesale") or 0),
                "selling": float(selling or meta.get("selling") or unit_price),
                "cost": float(cost_show or meta.get("cost") or 0),
            }

    out = []
    ranked = sorted(
        product_qty.items(),
        key=lambda x: (float(x[1]), float(product_rev.get(x[0]) or 0)),
        reverse=True,
    )[:limit]
    for pid, qty in ranked:
        rev = product_rev.get(pid) or Decimal("0")
        cost_total = product_cost.get(pid) or Decimal("0")
        costed_rev = product_costed_rev.get(pid) or Decimal("0")
        profit = (costed_rev - cost_total) if cost_total > 0 else Decimal("0")
        meta = product_meta.get(pid) or {}
        p = products_by_id.get(pid) if not str(pid).startswith("name:") else None
        cost_unit = float(meta.get("cost") or 0)
        if cost_unit <= 0 and p and (p.cost_price or Decimal("0")) > 0:
            cost_unit = float(p.cost_price)
        out.append(
            {
                "id": str(pid),
                "name": (p.name if p else "") or meta.get("name") or "Mahsulot",
                "image": (p.display_image if p else "") or meta.get("image") or "",
                "qty": float(qty),
                "qty_selling": float(product_qty_sell.get(pid) or 0),
                "qty_wholesale": float(product_qty_optom.get(pid) or 0),
                "revenue": float(rev),
                "revenue_selling": float(product_rev_sell.get(pid) or 0),
                "revenue_wholesale": float(product_rev_optom.get(pid) or 0),
                "cost": cost_unit,
                "cost_total": float(cost_total),
                "profit": float(profit),
                "stock": float(p.stock_qty) if p else float(meta.get("stock") or 0),
                "wholesale": float(p.wholesale_price or 0)
                if p and (p.wholesale_price or 0)
                else float(meta.get("wholesale") or 0),
                "selling": float(p.selling_price)
                if p
                else float(meta.get("selling") or 0),
            }
        )
    return out


def _top_products_qty_sum(rows: list[dict]) -> float:
    return sum(float(r.get("qty") or 0) for r in (rows or []) if isinstance(r, dict))


def _receipt_items(detail: dict | None) -> list[dict]:
    if not isinstance(detail, dict):
        return []
    for key in ("items", "lines", "products", "details", "receipt_items"):
        rows = detail.get(key)
        if isinstance(rows, list) and rows:
            return [x for x in rows if isinstance(x, dict)]
    return []


def _receipt_day(raw: dict) -> date | None:
    dt = _parse_dt(
        raw.get("completed_at")
        or raw.get("received_at")
        or raw.get("created_at")
        or raw.get("date")
        or raw.get("posted_at")
    )
    if not dt:
        return None
    return timezone.localtime(dt).date()


def _serialize_stock_receipt(
    raw: dict,
    products_by_id: dict,
    products_by_name: dict | None = None,
) -> dict:
    items_out = []
    total_qty = Decimal("0")
    total_cost = Decimal("0")
    for item in _receipt_items(raw):
        qty = _dec(item.get("quantity") or item.get("qty") or item.get("count"))
        unit_cost = _dec(
            item.get("cost_price")
            or item.get("unit_cost")
            or item.get("purchase_price")
            or item.get("price")
        )
        line_cost = _dec(
            item.get("total") or item.get("line_total") or item.get("total_cost"),
            str(qty * unit_cost),
        )
        pid = str(item.get("product_id") or item.get("product") or "").strip()
        if isinstance(item.get("product"), dict):
            nested = item["product"]
            pid = str(nested.get("id") or pid).strip()
            name = (nested.get("name") or item.get("product_name") or "").strip()
        else:
            name = (item.get("product_name") or item.get("name") or "").strip()
        p = _find_product(
            products_by_id,
            products_by_name,
            product_id=pid,
            product_name=name,
        )
        if not name and p:
            name = p.name
        if unit_cost <= 0 and p:
            unit_cost = _dec(getattr(p, "cost_price", 0))
            line_cost = qty * unit_cost
        total_qty += qty
        total_cost += line_cost
        items_out.append(
            {
                "id": pid,
                "name": name or "Mahsulot",
                "qty": float(qty),
                "unit_cost": float(unit_cost),
                "line_cost": float(line_cost),
                "image": (p.display_image if p else "") or "",
            }
        )
    if not items_out:
        # Ba'zi API lar faqat jami beradi
        total_cost = _dec(
            raw.get("total_cost")
            or raw.get("total")
            or raw.get("amount")
            or raw.get("cost_total")
        )
        total_qty = _dec(raw.get("total_qty") or raw.get("items_count") or raw.get("quantity"))

    dt = _parse_dt(
        raw.get("completed_at")
        or raw.get("received_at")
        or raw.get("created_at")
        or raw.get("date")
    )
    created_display = timezone.localtime(dt).strftime("%d.%m.%Y, %H:%M") if dt else ""
    supplier = (
        raw.get("supplier_name")
        or raw.get("supplier")
        or ""
    )
    if isinstance(supplier, dict):
        supplier = supplier.get("name") or ""
    return {
        "id": str(raw.get("id") or raw.get("uuid") or ""),
        "created_at": dt.isoformat() if dt else "",
        "created_display": created_display,
        "time": timezone.localtime(dt).strftime("%H:%M") if dt else "",
        "supplier": str(supplier or "").strip() or "—",
        "warehouse": str(raw.get("warehouse") or raw.get("warehouse_name") or "").strip() or "—",
        "items_count": len(items_out) if items_out else int(total_qty or 0),
        "total_qty": float(total_qty),
        "total_cost": float(total_cost),
        "items": items_out,
    }


def _aggregate_stock_in_products(receipts: list[dict]) -> list[dict]:
    """Kunlik kirim — mahsulot bo‘yicha yig‘indi."""
    buckets: dict[str, dict] = {}
    for rec in receipts:
        for it in rec.get("items") or []:
            if not isinstance(it, dict):
                continue
            key = str(it.get("id") or "").strip() or (it.get("name") or "").casefold()
            if not key:
                continue
            b = buckets.setdefault(
                key,
                {
                    "id": str(it.get("id") or ""),
                    "name": it.get("name") or "Mahsulot",
                    "image": it.get("image") or "",
                    "qty": 0.0,
                    "cost": 0.0,
                },
            )
            b["qty"] += float(it.get("qty") or 0)
            b["cost"] += float(it.get("line_cost") or 0)
            if not b.get("image") and it.get("image"):
                b["image"] = it["image"]
            if b.get("name") in ("", "Mahsulot") and it.get("name"):
                b["name"] = it["name"]
    return sorted(
        buckets.values(),
        key=lambda r: (float(r.get("cost") or 0), float(r.get("qty") or 0)),
        reverse=True,
    )


def _top_products_from_api_items(
    items: list,
    products_by_id: dict,
    limit: int = 100,
) -> list[dict]:
    out = []
    for row in items or []:
        if not isinstance(row, dict):
            continue
        pid = str(row.get("product_id") or row.get("id") or "")
        qty = _dec(row.get("quantity") or row.get("qty") or row.get("sold_qty"))
        rev = _dec(
            row.get("revenue")
            or row.get("total")
            or row.get("amount")
            or row.get("sum")
        )
        p = products_by_id.get(pid) if pid else None
        if rev <= 0 and qty > 0:
            unit = _dec(row.get("unit_price") or row.get("price"))
            if unit <= 0 and p:
                unit = p.selling_price
            rev = qty * unit if unit > 0 else Decimal("0")
        if rev <= 0 and p and qty > 0:
            rev = qty * p.selling_price
        if rev <= 0 and qty <= 0:
            continue
        name = (
            (row.get("product_name") or row.get("name") or "")
            or (p.name if p else "")
            or (f"Mahsulot {pid[:8]}" if pid else "Mahsulot")
        )
        cost_unit = float(p.cost_price) if p and (p.cost_price or 0) else 0.0
        cost_total = float(qty) * cost_unit if cost_unit > 0 else 0.0
        profit = float(rev) - cost_total if cost_total > 0 else 0.0
        wholesale = float(p.wholesale_price or 0) if p else float(_dec(row.get("wholesale_price")))
        selling = (
            float(p.selling_price)
            if p
            else float(_dec(row.get("selling_price") or row.get("unit_price")))
        )
        out.append(
            {
                "id": pid,
                "name": name,
                "image": (p.display_image if p else "") or str(row.get("image") or ""),
                "qty": float(qty),
                "qty_selling": float(qty),
                "qty_wholesale": 0.0,
                "revenue": float(rev),
                "revenue_selling": float(rev),
                "revenue_wholesale": 0.0,
                "cost": cost_unit,
                "cost_total": cost_total,
                "profit": profit,
                "stock": float(p.stock_qty) if p else float(_dec(row.get("stock"))),
                "wholesale": wholesale,
                "selling": selling,
            }
        )
        if len(out) >= limit:
            break
    out.sort(
        key=lambda r: (float(r.get("qty") or 0), float(r.get("revenue") or 0)),
        reverse=True,
    )
    return out[:limit]


@login_required
@require_GET
def cabinet_top_stats(request):
    """Top tovarlar — tanlangan oralig‘dagi BARCHA cheklar yig‘indisi (AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    today = timezone.localdate()

    custom_from = _parse_iso_date(request.GET.get("from"))
    custom_to = _parse_iso_date(request.GET.get("to"))
    if custom_from and custom_to:
        start, end = custom_from, custom_to
        if end < start:
            start, end = end, start
        if (end - start).days > 730:
            start = end - timedelta(days=730)
    else:
        start, end = today, today

    span = (end - start).days + 1
    try:
        limit = max(1, min(500, int(request.GET.get("limit") or 100)))
    except (TypeError, ValueError):
        limit = 100

    pack = _build_top_products_pack(
        token, server, start=start, end=end, limit=limit
    )
    if pack.get("error") == "auth":
        clear_tezpos_session(request)
        return JsonResponse({"error": "auth"}, status=401)
    if pack.get("error"):
        return JsonResponse({"error": pack["error"]}, status=502)

    return JsonResponse(
        {
            "from": start.isoformat(),
            "to": end.isoformat(),
            "topProducts": pack.get("topProducts") or [],
            "count": len(pack.get("topProducts") or []),
            "checks": pack.get("checks") or 0,
            "details_used": pack.get("details_used") or 0,
            "source": pack.get("source") or "sales",
            "span_days": span,
        }
    )


def _build_top_products_pack(
    token: str,
    server: str,
    *,
    start: date,
    end: date,
    limit: int = 100,
) -> dict:
    """
    Belgilangan kun(lar)dagi barcha cheklar bo‘yicha mahsulot yig‘indisi.
    Bitta mijozga katta savdo emas — kunlik/oralig‘ umumiy sotilgan miqdor.
    """
    span = (end - start).days + 1
    memo_prefix = f"{server}|{(token or '')[-12:]}"
    if span <= 1:
        max_pages, detail_cap, overall = 80, 400, 55.0
    elif span <= 7:
        max_pages, detail_cap, overall = 60, 300, 50.0
    elif span <= 31:
        max_pages, detail_cap, overall = 50, 200, 45.0
    else:
        max_pages, detail_cap, overall = 40, 120, 40.0

    try:
        sales = _memo_get(
            f"{memo_prefix}|topsales4|{start}|{end}|{max_pages}",
            60.0,
            lambda: tezpos_api.get_sales(
                token,
                server,
                date_from=start.isoformat(),
                date_to=end.isoformat(),
                timeout=22,
                max_pages=max_pages,
            ),
        ) or []
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            return {"error": "auth"}
        return {"error": str(exc)}
    except (TimeoutError, OSError) as exc:
        return {"error": str(exc)}

    products_raw = []
    try:
        products_raw = _memo_get(
            f"{memo_prefix}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=12) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError, Exception):
        products_raw = []

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)

    price_lists: list[dict] = []
    try:
        price_lists = _memo_get(
            f"{memo_prefix}|price_lists",
            120.0,
            lambda: tezpos_api.get_price_lists(token, server) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError, Exception):
        price_lists = []
    price_lists = [
        pl for pl in price_lists if isinstance(pl, dict) and pl.get("is_active", True)
    ]

    sales = [s for s in sales if isinstance(s, dict)]
    sales = [
        s
        for s in sales
        if (d := _sale_day(s)) is not None and start <= d <= end
    ]

    details_map: dict[str, dict] = {}
    need_fetch: list[str] = []
    for s in sales:
        sid = str(s.get("id") or "")
        if not sid:
            continue
        if _sale_items(s):
            details_map[sid] = s
        else:
            need_fetch.append(sid)

    if need_fetch:
        fetched = _fetch_sale_details(
            token,
            server,
            need_fetch,
            limit=min(detail_cap, max(len(need_fetch), 1)),
            per_sale_timeout=2.2,
            overall_timeout=overall,
        )
        details_map.update(fetched)

    top_products = _top_products_from_details(
        details_map,
        products_by_id,
        products_by_name,
        limit=limit,
        price_lists=price_lists,
    )

    # Agar cheklar itemsiz qolsa — API ni faqat zaxira sifatida
    if not top_products:
        try:
            top_payload = tezpos_api.get_top_products(
                token,
                server,
                days=span,
                limit=limit,
                date_from=start.isoformat(),
                date_to=end.isoformat(),
            ) or {"items": []}
        except (tezpos_api.TezPosApiError, TimeoutError, OSError):
            top_payload = {"items": []}
        top_products = _top_products_from_api_items(
            top_payload.get("items") or [], products_by_id, limit=limit
        )
        # API ham revenue bo‘yicha — qty ga o‘tkazamiz
        top_products.sort(
            key=lambda r: (float(r.get("qty") or 0), float(r.get("revenue") or 0)),
            reverse=True,
        )
        source = "api"
    else:
        source = "sales"

    return {
        "topProducts": top_products,
        "checks": len(sales),
        "details_used": len(details_map),
        "source": source,
        "products_by_id": products_by_id,
    }


@login_required
@require_GET
def cabinet_top_export(request):
    """Top reyting — tanlangan sana oralig‘i Excel."""
    if not session_has_tezpos(request):
        return redirect("login")

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    today = timezone.localdate()
    custom_from = _parse_iso_date(request.GET.get("from"))
    custom_to = _parse_iso_date(request.GET.get("to"))
    if custom_from and custom_to:
        start, end = custom_from, custom_to
        if end < start:
            start, end = end, start
        if (end - start).days > 730:
            start = end - timedelta(days=730)
    else:
        start, end = today, today

    try:
        limit = max(1, min(500, int(request.GET.get("limit") or 100)))
    except (TypeError, ValueError):
        limit = 100

    pack = _build_top_products_pack(
        token, server, start=start, end=end, limit=limit
    )
    if pack.get("error") == "auth":
        clear_tezpos_session(request)
        return redirect("login")
    if pack.get("error"):
        return HttpResponse(
            pack["error"], status=502, content_type="text/plain; charset=utf-8"
        )

    rows = pack.get("topProducts") or []
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Top reyting"
    ws.append(
        [
            "#",
            "Mahsulot",
            "Sotildi",
            "Sotuvda (dona)",
            "Optomda (dona)",
            "Sotib olish",
            "Sotuv narxi",
            "Optom narxi",
            "Tushum",
            "Tushum (sotuv)",
            "Tushum (optom)",
            "Jami tannarx",
            "Jami foyda",
            "Sana dan",
            "Sana gacha",
            "Cheklar",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, row in enumerate(rows, start=1):
        ws.append(
            [
                i,
                row.get("name") or "",
                float(row.get("qty") or 0),
                float(row.get("qty_selling") or 0),
                float(row.get("qty_wholesale") or 0),
                float(row.get("cost") or 0),
                float(row.get("selling") or 0),
                float(row.get("wholesale") or 0),
                float(row.get("revenue") or 0),
                float(row.get("revenue_selling") or 0),
                float(row.get("revenue_wholesale") or 0),
                float(row.get("cost_total") or 0),
                float(row.get("profit") or 0),
                start.isoformat(),
                end.isoformat(),
                int(pack.get("checks") or 0),
            ]
        )
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    for col in ("C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
        ws.column_dimensions[col].width = 14
        for cell in ws[col][1:]:
            cell.number_format = "#,##0.##"
            cell.alignment = Alignment(horizontal="right")
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 12
    ws.column_dimensions["P"].width = 10

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"top_reyting_{start.isoformat()}_{end.isoformat()}.xlsx"
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


@login_required
@require_GET
def cabinet_stock_in(request):
    """Kirim qilingan mahsulotlar — kalendar kuni bo‘yicha (AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    day = _parse_sale_date(request.GET.get("date") or request.GET.get("sale_date"))
    memo_prefix = f"{server}|{(token or '')[-12:]}"

    try:
        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_r = pool.submit(
                lambda: _memo_get(
                    f"{memo_prefix}|stockin|{day.isoformat()}",
                    60.0,
                    lambda: tezpos_api.get_stock_receipts(
                        token,
                        server,
                        date_from=day.isoformat(),
                        date_to=day.isoformat(),
                        timeout=18,
                        max_pages=30,
                    ),
                )
            )
            fut_p = pool.submit(
                lambda: _memo_get(
                    f"{memo_prefix}|catalog_snap",
                    120.0,
                    lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=14)
                    or [],
                )
            )
            raw_receipts = fut_r.result() or []
            products_raw = fut_p.result() or []
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc)}, status=200)
    except (TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=200)

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)

    raw_receipts = [r for r in raw_receipts if isinstance(r, dict)]
    # Sana filtri API da ishlamasa — lokal filtrlash
    filtered = []
    for r in raw_receipts:
        d = _receipt_day(r)
        if d is None or d == day:
            filtered.append(r)
    if not filtered and raw_receipts:
        # date_from/to qo‘llab-quvvatlanmasa — kengroq so‘rov + filtrlash
        try:
            wider = tezpos_api.get_stock_receipts(
                token,
                server,
                date_from=(day - timedelta(days=7)).isoformat(),
                date_to=(day + timedelta(days=1)).isoformat(),
                timeout=16,
                max_pages=40,
            )
        except (tezpos_api.TezPosApiError, TimeoutError, OSError):
            wider = []
        for r in wider or []:
            if isinstance(r, dict) and _receipt_day(r) == day:
                filtered.append(r)

    # Items yo‘q bo‘lsa — detail fetch
    need_ids = [
        str(r.get("id") or "")
        for r in filtered
        if r.get("id") and not _receipt_items(r)
    ]
    details: dict[str, dict] = {}
    if need_ids:
        def _one(rid: str):
            try:
                return rid, tezpos_api.get_stock_receipt(token, server, rid)
            except (tezpos_api.TezPosApiError, TimeoutError, OSError):
                return rid, None

        with ThreadPoolExecutor(max_workers=min(8, max(2, len(need_ids)))) as pool:
            for fut in as_completed([pool.submit(_one, rid) for rid in need_ids[:80]]):
                try:
                    rid, data = fut.result()
                except Exception:
                    continue
                if isinstance(data, dict):
                    details[rid] = data

    receipts_payload = []
    for r in filtered:
        rid = str(r.get("id") or "")
        detail = details.get(rid) or r
        if not _receipt_items(detail) and _receipt_items(r):
            detail = r
        receipts_payload.append(
            _serialize_stock_receipt(detail, products_by_id, products_by_name)
        )

    receipts_payload.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    products_agg = _aggregate_stock_in_products(receipts_payload)
    total_cost = sum(float(x.get("total_cost") or 0) for x in receipts_payload)
    total_qty = sum(float(x.get("total_qty") or 0) for x in receipts_payload)
    sku_count = len(products_agg)

    return JsonResponse(
        {
            "ok": True,
            "date": day.isoformat(),
            "receipts_count": len(receipts_payload),
            "sku_count": sku_count,
            "total_qty": total_qty,
            "total_cost": total_cost,
            "receipts": receipts_payload,
            "products": products_agg,
        }
    )


@login_required
@require_GET
def cabinet_shift_detail(request):
    """Bitta smena: sotuv/foyda + narxlar ro'yxati (AJAX)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    shift_id = (request.GET.get("id") or "").strip()
    opened = (request.GET.get("opened_at") or "").strip()
    closed = (request.GET.get("closed_at") or "").strip()
    sale_ids = [x for x in (request.GET.get("sale_ids") or "").split(",") if x.strip()]
    is_open = not closed or (request.GET.get("status") or "").strip() == "open"

    today = timezone.localdate()
    try:
        with ThreadPoolExecutor(max_workers=3) as pool:
            fut_pl = pool.submit(tezpos_api.get_price_lists, token, server)
            fut_p = pool.submit(
                lambda: tezpos_api.get_products(token, server, max_pages=3, timeout=12)
            )
            api_shift = {}
            if shift_id and not shift_id.startswith("session-"):
                fut_sh = pool.submit(tezpos_api.get_shift, token, server, shift_id)
            else:
                fut_sh = None
            price_lists = fut_pl.result(timeout=12) or []
            products_raw = fut_p.result(timeout=14) or []
            if fut_sh is not None:
                try:
                    api_shift = fut_sh.result(timeout=8) or {}
                except (tezpos_api.TezPosApiError, FuturesTimeoutError):
                    api_shift = {}
    except tezpos_api.TezPosApiError as exc:
        if getattr(exc, "status", None) in (401, 403):
            clear_tezpos_session(request)
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"ok": False, "error": str(exc)}, status=200)
    except (TimeoutError, OSError, FuturesTimeoutError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=200)

    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    price_lists = [pl for pl in price_lists if isinstance(pl, dict) and pl.get("is_active", True)]
    margin_ratio = _catalog_margin_ratio(products)

    opened_dt = _parse_dt(opened) or _parse_dt(api_shift.get("opened_at"))
    closed_dt = _parse_dt(closed) or _parse_dt(api_shift.get("closed_at"))
    if is_open or not closed_dt:
        closed_dt = timezone.now()
        is_open = True
    if opened_dt:
        date_from = timezone.localtime(opened_dt).date().isoformat()
        date_to = timezone.localtime(closed_dt).date().isoformat()
    else:
        date_from = (today - timedelta(days=2)).isoformat()
        date_to = today.isoformat()

    # Smena ochilgandan yopilguncha — barcha cheklar
    sale_pages = 60 if is_open else 40
    try:
        sales = tezpos_api.get_sales(
            token,
            server,
            date_from=date_from,
            date_to=date_to,
            timeout=16,
            max_pages=sale_pages,
        )
    except (tezpos_api.TezPosApiError, TimeoutError, OSError) as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=200)

    sales = [s for s in sales if isinstance(s, dict)]
    matched = _match_shift_sales(
        sales,
        opened_dt=opened_dt,
        closed_dt=closed_dt,
        sale_ids=sale_ids,
        is_open=is_open and bool(opened_dt),
    )

    gross = float(sum((_dec(s.get("total")) for s in matched), Decimal("0")))
    checks = len(matched)
    need_fetch = [
        str(s.get("id"))
        for s in matched
        if s.get("id") and not _sale_items(s)
    ]
    detail_cap = min(len(need_fetch) or 1, 35 if is_open or (opened_dt and date_from == date_to) else 28)
    details = _fetch_sale_details(
        token,
        server,
        need_fetch,
        limit=detail_cap,
        per_sale_timeout=2.2,
        overall_timeout=10.0,
    )
    lists, profit, margin = _shift_price_list_stats(
        matched,
        details,
        products_by_id,
        products_by_name,
        price_lists,
        gross,
        checks,
    )
    jami = next((r for r in lists if r.get("is_total")), None)

    return JsonResponse(
        {
            "id": shift_id,
            "summary": {
                "checks": checks,
                "gross": gross,
                "profit": profit,
                "margin": margin,
                "cost": gross - profit,
                "details_used": len(details),
            },
            "priceLists": [r for r in lists if not r.get("is_total")],
            "total": jami,
        }
    )

# --- telegram bot helpers & endpoints (appended) ---


def _shift_report_bundle(
    token: str,
    server: str,
    shift: dict,
) -> dict:
    """Smena uchun xabar + Excel ma'lumotlarini yig'adi."""
    from . import telegram_bot as tg

    today = timezone.localdate()
    try:
        products_raw = _memo_get(
            f"{server}|{(token or '')[-12:]}|catalog_snap",
            120.0,
            lambda: tezpos_api.get_catalog_snapshot(token, server, timeout=12) or [],
        ) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError):
        tg_logger.exception("shift bundle products failed")
        products_raw = []
    try:
        price_lists = tezpos_api.get_price_lists(token, server) or []
    except (tezpos_api.TezPosApiError, TimeoutError, OSError):
        price_lists = []
    products = [_map_product(p) for p in products_raw if isinstance(p, dict)]
    products_by_id = {str(p.id): p for p in products}
    products_by_name = _products_by_name(products)
    price_lists = [pl for pl in price_lists if isinstance(pl, dict) and pl.get("is_active", True)]
    margin_ratio = _catalog_margin_ratio(products)

    opened_dt = _parse_dt(shift.get("opened_at"))
    closed_dt = _parse_dt(shift.get("closed_at"))
    is_open = (shift.get("status") or "") == "open" or not closed_dt
    if is_open or not closed_dt:
        closed_dt = timezone.now()
        is_open = True
    if opened_dt:
        date_from = timezone.localtime(opened_dt).date().isoformat()
        date_to = timezone.localtime(closed_dt).date().isoformat()
    else:
        date_from = (today - timedelta(days=1)).isoformat()
        date_to = today.isoformat()

    sale_pages = 60 if is_open else 50
    sales = tezpos_api.get_sales(
        token,
        server,
        date_from=date_from,
        date_to=date_to,
        timeout=18,
        max_pages=sale_pages,
    )
    sales = [s for s in sales if isinstance(s, dict)]
    shift_sale_ids = [str(x) for x in (shift.get("sale_ids") or []) if x]
    matched = _match_shift_sales(
        sales,
        opened_dt=opened_dt,
        closed_dt=closed_dt,
        sale_ids=shift_sale_ids,
        is_open=is_open and bool(opened_dt),
    )

    gross = float(sum((_dec(s.get("total")) for s in matched), Decimal("0")))
    checks = len(matched)
    need_fetch = [
        str(s.get("id"))
        for s in matched
        if s.get("id") and not _sale_items(s)
    ]
    fetch_limit = min(len(need_fetch) or 1, 200)
    details = _fetch_sale_details(
        token,
        server,
        need_fetch,
        limit=fetch_limit,
        per_sale_timeout=2.5,
        overall_timeout=55.0,
    )
    lists, profit, margin = _shift_price_list_stats(
        matched,
        details,
        products_by_id,
        products_by_name,
        price_lists,
        gross,
        checks,
    )
    jami = next((r for r in lists if r.get("is_total")), None)

    selling_rev = 0.0
    wholesale_rev = 0.0
    pl_rows = []
    for r in lists:
        if r.get("is_total"):
            continue
        rid = str(r.get("id") or "")
        rev = float(r.get("revenue") or 0)
        if rid == SELLING_LIST_ID:
            selling_rev += rev
        else:
            wholesale_rev += rev
        pl_rows.append(
            {
                "name": r.get("name") or "",
                "checks": r.get("checks") or 0,
                "revenue": rev,
                "profit": float(r.get("profit") or 0),
                "share": float(r.get("share") or 0),
            }
        )

    credit_agg: dict[str, dict] = defaultdict(lambda: {"orders": 0, "total": 0.0})
    sales_rows = []
    sold_agg: dict[str, dict] = {}
    for s in matched:
        method = (
            s.get("payment_method")
            or s.get("payment_type")
            or s.get("payment")
            or ""
        )
        method = str(method).strip().lower()
        total = float(_dec(s.get("total")))
        cust = (s.get("customer_name") or s.get("customer") or "").strip() or "—"
        dt = _parse_dt(s.get("completed_at") or s.get("created_at") or s.get("sold_at"))
        time_s = timezone.localtime(dt).strftime("%d.%m.%Y %H:%M") if dt else ""
        sid = str(s.get("id") or "")
        detail = details.get(sid) if sid else None
        if not isinstance(detail, dict):
            detail = {}
        pay_src = (
            detail.get("payment_method")
            or detail.get("payment_type")
            or s.get("payment_method")
            or s.get("payment_type")
            or s.get("payment")
            or method
        )
        pay_label = _payment_label(pay_src)
        items = _sale_items(detail) or _sale_items(s)
        products_text = _sale_products_text(items, products_by_id, products_by_name)
        # To‘liq chek foydasi: sotish − sotib olish
        line_cost = Decimal("0")
        line_rev = Decimal("0")
        for item in items:
            qty = _item_qty(item)
            if qty <= 0:
                continue
            unit_price = _item_unit_price(item)
            line_total = _dec(item.get("total") or item.get("line_total"), str(qty * unit_price))
            unit_cost = _resolve_item_unit_cost(item, products_by_id, products_by_name)
            line_rev += line_total
            line_cost += qty * unit_cost
            pid = str(item.get("product_id") or item.get("product") or "").strip()
            name = (item.get("product_name") or item.get("name") or "").strip()
            p = _find_product(
                products_by_id,
                products_by_name,
                product_id=pid,
                product_name=name,
            )
            if not name and p:
                name = p.name
            if not name:
                name = "Mahsulot"
            key = pid or name.lower()
            bucket = sold_agg.setdefault(
                key,
                {
                    "name": name,
                    "barcode": (p.barcode if p else "") or (item.get("barcode") or ""),
                    "unit": _item_unit(item, p),
                    "qty": 0.0,
                    "revenue": 0.0,
                    "profit": 0.0,
                },
            )
            bucket["qty"] += float(qty)
            bucket["revenue"] += float(line_total)
            bucket["profit"] += float(line_total - (qty * unit_cost))
            if not bucket.get("barcode") and p:
                bucket["barcode"] = p.barcode or ""
        if items:
            sale_profit = float((line_rev - line_cost).quantize(Decimal("0.01")))
            if line_rev > 0 and abs(total - float(line_rev)) > 1:
                # Chegirma/jami farqi — chek jami asosida
                sale_profit = float(Decimal(str(total)) - line_cost)
        elif detail:
            _, pft = _estimate_sale_profit(
                detail, products_by_id, _dec(total), products_by_name, margin_ratio
            )
            sale_profit = float(pft)
        else:
            sale_profit = total * float(margin_ratio)
        sales_rows.append(
            {
                "receipt_no": _display_receipt_number(detail, s) or _receipt_number(s, detail),
                "time": time_s,
                "customer": cust,
                "products_text": products_text,
                "payment": pay_label,
                "total": total,
                "profit": sale_profit,
            }
        )
        if method in ("credit", "qarz", "debt", "nasiya") or pay_label == "Qarz":
            credit_agg[cust]["orders"] += 1
            credit_agg[cust]["total"] += total

    credit_rows = [
        {"customer": n, "orders": v["orders"], "total": v["total"]}
        for n, v in sorted(credit_agg.items(), key=lambda x: x[1]["total"], reverse=True)
    ]
    credit_total = sum(v["total"] for v in credit_rows)

    sold_product_rows = sorted(
        (r for r in sold_agg.values() if float(r.get("qty") or 0) > 0),
        key=lambda r: (float(r.get("qty") or 0), float(r.get("revenue") or 0)),
        reverse=True,
    )

    # To'liq qarzdorlar ro'yxati (API)
    debtors_rows: list[dict] = []
    debtors_total = Decimal("0")
    try:
        customers = tezpos_api.get_customers(token, server) or []
        for c in customers:
            if not isinstance(c, dict):
                continue
            debt = _debt_amount(c.get("debt"))
            if debt <= 0:
                continue
            debtors_total += debt
            debtors_rows.append(
                {
                    "name": (c.get("name") or "").strip() or "Mijoz",
                    "phone": (c.get("phone") or "").strip(),
                    "debt": float(debt),
                }
            )
        debtors_rows.sort(key=lambda r: r["debt"], reverse=True)
    except Exception:
        debtors_rows = []
        debtors_total = Decimal("0")

    # Kam qoldiq
    low_stock_rows = _build_near_min_stock(products)

    # Kirim qilingan mahsulotlar (smena kunlari)
    stock_in_rows: list[dict] = []
    stock_in_sku = 0
    stock_in_qty = 0.0
    stock_in_cost = 0.0
    try:
        raw_receipts = tezpos_api.get_stock_receipts(
            token,
            server,
            date_from=date_from,
            date_to=date_to,
            timeout=14,
            max_pages=30,
        ) or []
        filtered_receipts = []
        for r in raw_receipts:
            if not isinstance(r, dict):
                continue
            rdt = _parse_dt(
                r.get("completed_at")
                or r.get("received_at")
                or r.get("created_at")
                or r.get("date")
            )
            d = _receipt_day(r)
            if opened_dt and closed_dt and rdt:
                if opened_dt <= rdt <= closed_dt:
                    filtered_receipts.append(r)
            elif d and date_from <= d.isoformat() <= date_to:
                filtered_receipts.append(r)
        need_ids = [
            str(r.get("id") or "")
            for r in filtered_receipts
            if r.get("id") and not _receipt_items(r)
        ]
        fetched_rc: dict[str, dict] = {}
        if need_ids:
            def _one_rc(rid: str):
                try:
                    return rid, tezpos_api.get_stock_receipt(token, server, rid)
                except (tezpos_api.TezPosApiError, TimeoutError, OSError):
                    return rid, None

            with ThreadPoolExecutor(max_workers=min(8, max(2, len(need_ids)))) as pool:
                for fut in as_completed(
                    [pool.submit(_one_rc, rid) for rid in need_ids[:80]]
                ):
                    try:
                        rid, data = fut.result()
                    except Exception:
                        continue
                    if isinstance(data, dict):
                        fetched_rc[rid] = data
        serialized = []
        for r in filtered_receipts:
            rid = str(r.get("id") or "")
            detail = fetched_rc.get(rid) or r
            serialized.append(
                _serialize_stock_receipt(detail, products_by_id, products_by_name)
            )
        stock_in_rows = _aggregate_stock_in_products(serialized)
        stock_in_sku = len(stock_in_rows)
        stock_in_qty = float(sum(float(x.get("qty") or 0) for x in stock_in_rows))
        stock_in_cost = float(sum(float(x.get("cost") or 0) for x in stock_in_rows))
        if stock_in_cost <= 0:
            stock_in_cost = float(
                sum(float(x.get("total_cost") or 0) for x in serialized)
            )
    except Exception:
        tg_logger.exception("shift stock-in load failed")
        stock_in_rows = []

    opened_disp = shift.get("opened_display") or shift.get("opened_at") or ""
    closed_disp = shift.get("closed_display") or shift.get("closed_at") or ""
    # Namuna: 09.08.2026 · 10:17
    if opened_disp and " · " not in str(opened_disp):
        opened_disp = str(opened_disp).replace(" ", " · ", 1)
    if closed_disp and closed_disp != "—" and " · " not in str(closed_disp):
        closed_disp = str(closed_disp).replace(" ", " · ", 1)

    duration_label = tg.format_duration(
        str(shift.get("opened_at") or ""),
        str(shift.get("closed_at") or ""),
    )

    enriched = {
        **shift,
        "checks": checks,
        "gross": gross,
        "profit": profit,
        "margin": margin,
        "selling_revenue": selling_rev,
        "wholesale_revenue": wholesale_rev,
        "credit_total": credit_total,
        "credit_customers": credit_rows[:20],
        "debtors": debtors_rows,
        "debtors_count": len(debtors_rows),
        "debtors_total": float(debtors_total),
        "low_stock": low_stock_rows,
        "low_stock_count": len(low_stock_rows),
        "stock_in_count": stock_in_sku,
        "stock_in_qty": stock_in_qty,
        "stock_in_total": stock_in_cost,
        "stock_in_products": stock_in_rows[:80],
        "opened_at_display": opened_disp,
        "closed_at_display": closed_disp,
        "duration_label": duration_label,
    }
    biz = shift.get("business_name") or "TezPOS"
    return {
        "shift": enriched,
        "sales_rows": sales_rows,
        "price_lists": pl_rows,
        "credit_rows": credit_rows,
        "debtors_rows": debtors_rows,
        "low_stock_rows": low_stock_rows,
        "sold_product_rows": sold_product_rows,
        "stock_in_rows": stock_in_rows,
        "excel": tg.build_shift_excel(
            business_name=biz,
            shift=enriched,
            sales_rows=sales_rows,
            price_lists=pl_rows,
            credit_rows=credit_rows,
            debtors_rows=debtors_rows,
            low_stock_rows=low_stock_rows,
            sold_product_rows=sold_product_rows,
            stock_in_rows=stock_in_rows,
        ),
    }


@login_required
def cabinet_label_template(request):
    """Do‘kon bo‘yicha umumiy narx belgisi shabloni (GET/POST)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    shop_key = _label_shop_key(request)
    if request.method == "GET":
        row = LabelTemplate.objects.filter(shop_key=shop_key).first()
        data = row.data if row and isinstance(row.data, dict) else {}
        return JsonResponse(
            {
                "ok": True,
                "has_template": bool(row and data),
                "template": data or {},
            }
        )
    if request.method != "POST":
        return JsonResponse({"error": "method"}, status=405)

    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Noto‘g‘ri JSON"}, status=400)
    if not isinstance(body, dict):
        return JsonResponse({"error": "Noto‘g‘ri JSON"}, status=400)

    display = (request.session.get(SESSION_DISPLAY) or request.user.get_username() or "")[:180]
    if body.get("reset"):
        LabelTemplate.objects.filter(shop_key=shop_key).delete()
        return JsonResponse({"ok": True, "has_template": False, "template": {}})

    template = _sanitize_label_template(body)
    LabelTemplate.objects.update_or_create(
        shop_key=shop_key,
        defaults={"data": template, "updated_by": display},
    )
    return JsonResponse({"ok": True, "has_template": True, "template": template})


@login_required
@require_POST
def cabinet_bot_settings_save(request):
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    tenant = get_tenant_for_user(request.user)
    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Noto‘g‘ri JSON"}, status=400)

    token = (body.get("token") or "").strip()
    recipients = (body.get("recipients") or "").strip()
    enabled = bool(body.get("enabled"))
    notify_open = bool(body.get("notify_open", True))
    notify_close = bool(body.get("notify_close", True))
    test_send = bool(body.get("test"))

    if enabled and not token:
        return JsonResponse({"error": "Bot token majburiy"}, status=400)
    if enabled and not recipients:
        return JsonResponse({"error": "Kamida bitta qabul qiluvchi kiriting"}, status=400)

    from . import telegram_bot as tg

    parsed = tg.parse_recipients(recipients)
    if enabled and not parsed:
        return JsonResponse(
            {
                "error": "Qabul qiluvchilar noto‘g‘ri. Chat ID, @username yoki t.me/... kiriting.",
            },
            status=400,
        )

    bot_info = None
    if token:
        bot_info = tg.get_me(token)
        if not bot_info.get("ok"):
            return JsonResponse(
                {
                    "error": f"Token ishlamadi: {bot_info.get('description') or 'xato'}",
                },
                status=400,
            )

    tenant.telegram_bot_token = token
    tenant.telegram_recipients = recipients
    tenant.telegram_enabled = enabled
    tenant.telegram_notify_open = notify_open
    tenant.telegram_notify_close = notify_close
    # Birinchi yoqishda eski smenalarni yubormaslik
    if enabled and not (tenant.telegram_notified_events or {}):
        tenant.telegram_notified_events = {"__seeded__": timezone.now().isoformat()}
    tenant.save(
        update_fields=[
            "telegram_bot_token",
            "telegram_recipients",
            "telegram_enabled",
            "telegram_notify_open",
            "telegram_notify_close",
            "telegram_notified_events",
        ]
    )

    test_results = []
    if test_send and token and parsed:
        uname = ((bot_info or {}).get("result") or {}).get("username") or "bot"
        msg = (
            f"✅ <b>TezPOS bot ulandi</b>\n"
            f"Bot: @{uname}\n"
            f"Biznes: {tenant.business_name}\n"
            f"Smena ochilish/yopilish xabarlari shu yerga keladi."
        )
        test_results = tg.broadcast_message(token, parsed, msg)

    return JsonResponse(
        {
            "ok": True,
            "enabled": tenant.telegram_enabled,
            "bot": ((bot_info or {}).get("result") or {}),
            "recipients": parsed,
            "test": test_results,
        }
    )


@login_required
@require_GET
def cabinet_bot_chats(request):
    """Botga /start yozgan chatlarni ko‘rsatadi (Chat ID olish)."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    tenant = get_tenant_for_user(request.user)
    token = (request.GET.get("token") or tenant.telegram_bot_token or "").strip()
    if not token:
        return JsonResponse({"error": "Avval bot tokenini kiriting"}, status=400)
    from . import telegram_bot as tg

    bot_info = tg.get_me(token)
    if not bot_info.get("ok"):
        return JsonResponse(
            {"error": f"Token ishlamadi: {bot_info.get('description') or 'xato'}"},
            status=400,
        )
    chats = tg.list_recent_chats(token)
    return JsonResponse(
        {
            "ok": True,
            "bot": bot_info.get("result") or {},
            "chats": chats,
            "hint": (
                "Ro‘yxat bo‘sh bo‘lsa: Telegramda botga kirib /start yozing, "
                "keyin shu tugmani qayta bosing."
            ),
        }
    )


@login_required
@require_GET
def cabinet_telegram_sync(request):
    """Web worker bloklanmasin: faqat status + cron uchun token yangilash."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)

    tenant = get_tenant_for_user(request.user)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    if token and server:
        TenantProfile.objects.filter(pk=tenant.pk).update(
            tezpos_api_token=token,
            tezpos_server_name=server,
        )
    last = tenant.telegram_last_sync or {}
    return JsonResponse(
        {
            "ok": True,
            "skipped": True,
            "reason": "background_only",
            "enabled": bool(tenant.telegram_enabled and tenant.telegram_bot_token),
            "last_sync": last,
        }
    )


@require_GET
def telegram_cron_sync(request):
    """Faqat oxirgi sync holati. Haqiqiy sync — manage.py / systemd."""
    from django.conf import settings as dj_settings

    expected = (getattr(dj_settings, "TELEGRAM_CRON_SECRET", "") or "").strip()
    got = (
        request.headers.get("X-Cron-Secret")
        or request.GET.get("secret")
        or ""
    ).strip()
    if not expected or got != expected:
        return JsonResponse({"error": "forbidden"}, status=403)

    tenants = TenantProfile.objects.filter(telegram_enabled=True).exclude(
        telegram_bot_token=""
    )
    results = [
        {
            "tenant": t.business_name,
            "last_sync": t.telegram_last_sync or {},
        }
        for t in tenants.only("business_name", "telegram_last_sync")
    ]
    return JsonResponse({"ok": True, "sync": "systemd", "results": results})


@csrf_exempt
@require_http_methods(["POST", "OPTIONS"])
def telegram_shift_ping(request):
    """
    Desktop TezPOS: smena ochilishi/yopilishi bilan darhol Telegram sync.
    Body: { "server": "kuloloptom", "token": "<tezpos api token>", "event": "open"|"close" }
    """
    if request.method == "OPTIONS":
        resp = HttpResponse(status=204)
        resp["Access-Control-Allow-Origin"] = "*"
        resp["Access-Control-Allow-Methods"] = "POST, OPTIONS"
        resp["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    def _cors(payload, status=200):
        r = JsonResponse(payload, status=status)
        r["Access-Control-Allow-Origin"] = "*"
        return r

    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return _cors({"ok": False, "error": "json"}, status=400)

    server = re.sub(r"[^a-z0-9_-]", "", str(body.get("server") or "").strip().lower())
    token = str(body.get("token") or "").strip()
    if not server or not token:
        return _cors({"ok": False, "error": "server_token_required"}, status=400)

    tenant = (
        TenantProfile.objects.filter(telegram_enabled=True)
        .exclude(telegram_bot_token="")
        .filter(tezpos_server_name__iexact=server)
        .first()
    )
    if not tenant:
        tenant = (
            TenantProfile.objects.filter(telegram_enabled=True, tezpos_api_token=token)
            .exclude(telegram_bot_token="")
            .first()
        )
    if not tenant:
        return _cors({"ok": False, "error": "bot_not_configured", "server": server}, status=404)

    TenantProfile.objects.filter(pk=tenant.pk).update(
        tezpos_api_token=token,
        tezpos_server_name=server,
    )
    tenant.tezpos_api_token = token
    tenant.tezpos_server_name = server

    def _run():
        try:
            # Yopilishda API smena statusi biroz kechikishi mumkin
            time.sleep(1.2)
            sync_telegram_shifts_for_tenant(tenant, token, server)
        except Exception:
            tg_logger.exception(
                "telegram_shift_ping failed tenant=%s server=%s",
                getattr(tenant, "business_name", ""),
                server,
            )

    threading.Thread(
        target=_run,
        daemon=True,
        name=f"tg-ping-{server[:20]}",
    ).start()
    return _cors({"ok": True, "queued": True, "tenant": tenant.business_name})


def _tg_any_ok(results: list | None) -> bool:
    return any(bool(r.get("ok")) for r in (results or []) if isinstance(r, dict))


def sync_telegram_shifts_for_tenant(tenant, token: str, server: str) -> dict:
    """
    Smena ochilish/yopilishni tekshiradi.
    Yopilishda: to‘liq hisobot (xabar + Excel) guruh/kanal/shaxsiy chatlarga.
    Muvaffaqiyatsiz yuborish qayta uriniladi (notified faqat ok bo‘lsa).
    """
    from . import telegram_bot as tg

    t0 = time.time()
    api_s = 0.0
    db_s = 0.0
    deadline = t0 + 140.0
    max_heavy = 2  # yopilish xabari + Excel (fon, web emas)

    def _persist_meta(extra: dict | None = None) -> None:
        nonlocal db_s
        meta = {
            "at": timezone.now().isoformat(),
            "duration_s": round(time.time() - t0, 2),
            "api_s": round(api_s, 2),
            "db_s": round(db_s, 2),
        }
        if extra:
            meta.update(extra)
        tenant.telegram_last_sync = meta
        tdb = time.time()
        tenant.save(update_fields=["telegram_notified_events", "telegram_last_sync"])
        db_s += time.time() - tdb

    recipients = tg.parse_recipients(tenant.telegram_recipients)
    if not recipients:
        result = {"ok": True, "skipped": True, "reason": "no_recipients"}
        tenant.telegram_last_sync = {**result, "at": timezone.now().isoformat()}
        tenant.save(update_fields=["telegram_last_sync"])
        return result

    today = timezone.localdate()

    t_api = time.time()
    try:
        shifts, shifts_source = _collect_shifts_payload(token, server, days=30)
    except (tezpos_api.TezPosApiError, TimeoutError, OSError):
        tg_logger.exception("get_shifts failed tenant=%s", tenant.business_name)
        raise
    api_s += time.time() - t_api

    # Faqat API smenalari — session-* (sotuvdan taxmin) spam qilmasin
    if shifts_source == "api":
        shifts = [sh for sh in shifts if str(sh.get("id") or "").strip() and not str(sh.get("id")).startswith("session-")]
    elif not shifts:
        t_api = time.time()
        try:
            sales = tezpos_api.get_sales(
                token,
                server,
                date_from=(today - timedelta(days=2)).isoformat(),
                date_to=today.isoformat(),
                timeout=14,
                max_pages=20,
            )
        except (tezpos_api.TezPosApiError, TimeoutError, OSError):
            tg_logger.exception("sales fallback failed tenant=%s", tenant.business_name)
            sales = []
        api_s += time.time() - t_api
        shifts = _build_shifts_from_sales(
            [s for s in sales if isinstance(s, dict)],
            today=today,
            gap_hours=4.0,
            margin_ratio=Decimal("0"),
        )

    notified = dict(tenant.telegram_notified_events or {})
    sent: list[dict] = []
    heavy_used = 0
    # Birinchi sync: mavjud smenalarni faqat belgilash (spam bo‘lmasin)
    seed_only = "__seeded__" in notified and len(notified) <= 2
    if seed_only:
        for sh in shifts[:20]:
            sid = str(sh.get("id") or "").strip()
            if not sid:
                continue
            st = sh.get("status") or "closed"
            if st == "open":
                notified[f"{sid}:open"] = timezone.now().isoformat()
            else:
                notified[f"{sid}:close"] = timezone.now().isoformat()
                notified.setdefault(f"{sid}:open", timezone.now().isoformat())
                notified.setdefault(f"{sid}:excel", timezone.now().isoformat())
        notified.pop("__seeded__", None)
        tenant.telegram_notified_events = notified
        _persist_meta({"checked": len(shifts), "seeded": True, "sent": 0})
        return {"ok": True, "sent": [], "checked": len(shifts), "seeded": True}

    for sh in shifts[:12]:
        if time.time() >= deadline:
            break
        sid = str(sh.get("id") or "").strip()
        if not sid:
            continue
        status = sh.get("status") or "closed"
        for event, enabled in (
            ("open", tenant.telegram_notify_open and status == "open"),
            ("close", tenant.telegram_notify_close and status == "closed"),
        ):
            if not enabled:
                continue
            key = f"{sid}:{event}"
            if key in notified:
                continue
            if event == "open":
                if str(sid).startswith("session-"):
                    continue
                opened_dt = _parse_dt(sh.get("opened_at"))
                if opened_dt and timezone.now() - opened_dt > timedelta(hours=36):
                    notified[key] = timezone.now().isoformat()
                    continue
            if event == "close":
                closed_dt = _parse_dt(sh.get("closed_at")) or _parse_dt(sh.get("opened_at"))
                if closed_dt and timezone.now() - closed_dt > timedelta(days=3):
                    notified[key] = timezone.now().isoformat()
                    notified.setdefault(f"{sid}:excel", timezone.now().isoformat())
                    continue

            try:
                doc_res: list = []
                if event == "close":
                    if heavy_used >= max_heavy or time.time() >= deadline:
                        continue
                    heavy_used += 1
                    t_api = time.time()
                    bundle = _shift_report_bundle(
                        token, server, {**sh, "business_name": tenant.business_name}
                    )
                    api_s += time.time() - t_api
                    shift_payload = bundle.get("shift") or dict(sh)
                    shift_payload["business_name"] = tenant.business_name
                    text = tg.build_shift_message(
                        business_name=tenant.business_name,
                        event=event,
                        shift=shift_payload,
                    )
                    msg_res = tg.broadcast_message(
                        tenant.telegram_bot_token, recipients, text
                    )
                    if not _tg_any_ok(msg_res):
                        sent.append(
                            {
                                "shift_id": sid,
                                "event": event,
                                "messages": msg_res,
                                "error": "xabar yetmadi — qayta uriniladi",
                            }
                        )
                        continue

                    notified[key] = timezone.now().isoformat()
                    tenant.telegram_notified_events = notified
                    tenant.save(update_fields=["telegram_notified_events"])

                    excel_key = f"{sid}:excel"
                    if excel_key not in notified and bundle.get("excel"):
                        day = timezone.localdate().isoformat()
                        fname = f"kunlik_hisobot_{day}_{sid[:8]}.xlsx"
                        doc_res = tg.broadcast_document(
                            tenant.telegram_bot_token,
                            recipients,
                            fname,
                            bundle["excel"],
                            caption=(
                                f"📎 Kunlik hisobot — {tenant.business_name}\n"
                                f"Sotuv · Kirim · Mahsulotlar · Qarzdorlar · Kam qoldiq"
                            ),
                        )
                        if _tg_any_ok(doc_res):
                            notified[excel_key] = timezone.now().isoformat()
                            tenant.telegram_notified_events = notified
                            tdb = time.time()
                            tenant.save(update_fields=["telegram_notified_events"])
                            db_s += time.time() - tdb
                else:
                    quick = dict(sh)
                    quick["business_name"] = tenant.business_name
                    opened_disp = sh.get("opened_display") or sh.get("opened_at") or ""
                    if opened_disp and " · " not in str(opened_disp):
                        opened_disp = str(opened_disp).replace(" ", " · ", 1)
                    quick["opened_at_display"] = opened_disp
                    quick["closed_at_display"] = sh.get("closed_display") or ""
                    text = tg.build_shift_message(
                        business_name=tenant.business_name,
                        event=event,
                        shift=quick,
                    )
                    msg_res = tg.broadcast_message(
                        tenant.telegram_bot_token, recipients, text
                    )
                    if not _tg_any_ok(msg_res):
                        sent.append(
                            {
                                "shift_id": sid,
                                "event": event,
                                "messages": msg_res,
                                "error": "xabar yetmadi — qayta uriniladi",
                            }
                        )
                        continue
                    notified[key] = timezone.now().isoformat()
                    tenant.telegram_notified_events = notified
                    tenant.save(update_fields=["telegram_notified_events"])

                sent.append(
                    {
                        "shift_id": sid,
                        "event": event,
                        "messages": msg_res,
                        "documents": doc_res,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                tg_logger.exception(
                    "telegram event failed tenant=%s shift=%s event=%s",
                    tenant.business_name,
                    sid,
                    event,
                )
                sent.append({"shift_id": sid, "event": event, "error": str(exc)})

    # Yopilgan smenalar uchun Excel hali ketmagan bo‘lsa — alohida urinish
    if tenant.telegram_notify_close:
        for sh in shifts[:12]:
            if time.time() >= deadline or heavy_used >= max_heavy:
                break
            sid = str(sh.get("id") or "").strip()
            if not sid or (sh.get("status") or "") != "closed":
                continue
            close_key = f"{sid}:close"
            excel_key = f"{sid}:excel"
            if close_key not in notified or excel_key in notified:
                continue
            try:
                heavy_used += 1
                t_api = time.time()
                bundle = _shift_report_bundle(
                    token, server, {**sh, "business_name": tenant.business_name}
                )
                api_s += time.time() - t_api
                if bundle.get("excel"):
                    day = timezone.localdate().isoformat()
                    fname = f"kunlik_hisobot_{day}_{sid[:8]}.xlsx"
                    doc_res = tg.broadcast_document(
                        tenant.telegram_bot_token,
                        recipients,
                        fname,
                        bundle["excel"],
                        caption=f"📎 Kunlik hisobot — {tenant.business_name}",
                    )
                    if _tg_any_ok(doc_res):
                        notified[excel_key] = timezone.now().isoformat()
                    sent.append(
                        {"shift_id": sid, "event": "excel", "documents": doc_res}
                    )
            except Exception as exc:  # noqa: BLE001
                tg_logger.exception(
                    "telegram excel failed tenant=%s shift=%s",
                    tenant.business_name,
                    sid,
                )
                sent.append({"shift_id": sid, "event": "excel", "error": str(exc)})

    if len(notified) > 120:
        items = sorted(notified.items(), key=lambda x: x[1])
        notified = dict(items[-80:])

    tenant.telegram_notified_events = notified
    duration = round(time.time() - t0, 2)
    _persist_meta(
        {
            "checked": len(shifts),
            "sent": len(sent),
            "errors": sum(1 for x in sent if x.get("error")),
        }
    )
    result = {
        "ok": True,
        "sent": sent,
        "checked": len(shifts),
        "recipients": recipients,
        "duration_s": duration,
        "api_s": round(api_s, 2),
        "db_s": round(db_s, 2),
    }
    tg_logger.info(
        "tenant=%s checked=%s sent=%s duration=%.2fs api=%.2fs db=%.2fs",
        tenant.business_name,
        len(shifts),
        len(sent),
        duration,
        api_s,
        db_s,
    )
    return result


def _debt_amount(value) -> Decimal:
    try:
        return Decimal(str(value or 0).replace(" ", "").replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _site_check_url(request, check_path_or_url: str) -> str:
    """API yoki nisbiy chek yo‘lini shu saytdagi /check/... URL ga aylantiradi."""
    raw = (check_path_or_url or "").strip()
    if not raw:
        return ""
    m = re.search(r"/check/([^/]+)/([^/]+)/?", raw)
    if m:
        path = reverse(
            "public-receipt-check",
            kwargs={"server_name": m.group(1), "ref": m.group(2).rstrip("/")},
        )
        return request.build_absolute_uri(path)
    if raw.startswith("http://") or raw.startswith("https://"):
        return raw
    if not raw.startswith("/"):
        raw = "/" + raw
    return request.build_absolute_uri(raw)


def _sms_check_url(request, check_path_or_url: str) -> str:
    """SMS uchun ochiladigan chek — localhost o‘rniga tez-pos.uz yoki API."""
    url = _site_check_url(request, check_path_or_url)
    if url and "127.0.0.1" not in url and "localhost" not in url:
        return url
    m = re.search(r"/check/([^/]+)/([^/]+)/?", check_path_or_url or url or "")
    if not m:
        return url
    slug, ref = m.group(1), m.group(2).rstrip("/")
    # Production sayt
    return f"https://tez-pos.uz/check/{slug}/{ref}/"


def _money_sms(value) -> str:
    try:
        n = Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        n = Decimal("0")
    return f"{n:,.0f}".replace(",", " ")


def _build_debt_payment_sms(
    *,
    store: str,
    cashier: str,
    paid,
    balance,
    check_url: str,
) -> str:
    """Qarz to‘lovi — TezPOS MpBuildDebtMessage."""
    shop = (cashier or store or "TezPOS").strip() or "TezPOS"
    branch = (store or "").strip() if store and store != shop else ""
    return devsms.build_debt_message(
        shop=shop,
        branch=branch,
        debt_amount=-abs(Decimal(str(paid or 0))),
        balance=balance,
        check_link=check_url,
    )


@login_required
@require_GET
def cabinet_debtors(request):
    """TezPOS API dan qarzdorlar ro‘yxati."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    try:
        customers = tezpos_api.get_customers(token, server)
    except tezpos_api.TezPosApiError as exc:
        if exc.status in (401, 403):
            return JsonResponse({"error": "auth"}, status=401)
        return JsonResponse({"error": str(exc)}, status=502)
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=502)

    rows = []
    total = Decimal("0")
    for c in customers or []:
        if not isinstance(c, dict):
            continue
        debt = _debt_amount(c.get("debt"))
        if debt <= 0:
            continue
        total += debt
        name = (c.get("name") or "").strip() or "Mijoz"
        phone = (c.get("phone") or "").strip()
        rows.append(
            {
                "id": str(c.get("id") or ""),
                "name": name,
                "phone": phone,
                "debt": float(debt),
                "debt_display": f"{debt:,.0f}".replace(",", " "),
            }
        )
    rows.sort(key=lambda r: r["debt"], reverse=True)
    return JsonResponse(
        {
            "ok": True,
            "count": len(rows),
            "total_debt": float(total),
            "total_debt_display": f"{total:,.0f}".replace(",", " "),
            "debtors": rows,
        }
    )


@login_required
@require_POST
def cabinet_debtor_pay(request):
    """Qarzning bir qismi yoki to‘liq to‘lovi — TezPOS pay-debt API."""
    if not session_has_tezpos(request):
        return JsonResponse({"error": "auth"}, status=401)
    token = request.session[SESSION_TOKEN]
    server = request.session[SESSION_SERVER]
    try:
        body = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Noto‘g‘ri JSON"}, status=400)

    customer_id = str(body.get("customer_id") or "").strip()
    if not customer_id:
        return JsonResponse({"error": "Mijoz tanlanmagan."}, status=400)

    amount = _debt_amount(body.get("amount"))
    if amount <= 0:
        return JsonResponse({"error": "To‘lov summasi 0 dan katta bo‘lishi kerak."}, status=400)

    payment_type = str(body.get("payment_type") or "cash").strip().lower()
    if payment_type not in ("cash", "card"):
        payment_type = "cash"
    note = str(body.get("note") or "").strip()[:255]

    try:
        site_base = request.build_absolute_uri("/").rstrip("/")
        # Localhost DevSMS/chek uchun yaroqsiz
        if "127.0.0.1" in site_base or "localhost" in site_base:
            check_base = "https://tez-pos.uz"
        else:
            check_base = site_base
        result = tezpos_api.pay_customer_debt(
            token,
            server,
            customer_id,
            amount,
            payment_type=payment_type,
            note=note,
            check_base_url=check_base,
        )
    except tezpos_api.TezPosApiError as exc:
        if exc.status in (401, 403):
            return JsonResponse({"error": "auth"}, status=401)
        msg = str(exc)
        if isinstance(exc.payload, (bytes, bytearray)):
            try:
                payload = json.loads(exc.payload.decode("utf-8", errors="replace"))
                if isinstance(payload, dict) and payload.get("detail"):
                    msg = str(payload["detail"])
            except Exception:
                pass
        return JsonResponse({"error": msg}, status=exc.status or 502)
    except Exception as exc:  # noqa: BLE001
        return JsonResponse({"error": str(exc)}, status=502)

    if not isinstance(result, dict):
        return JsonResponse({"error": "Server javobi noto‘g‘ri"}, status=502)

    check_raw = (
        result.get("check_path")
        or result.get("check_url")
        or result.get("receipt_url")
        or ""
    )
    check_url = _site_check_url(request, str(check_raw))
    sms_check_url = _sms_check_url(request, str(check_raw) or check_url)

    customer = result.get("customer") if isinstance(result.get("customer"), dict) else {}
    balance = _debt_amount(result.get("balance", customer.get("debt")))
    paid = _debt_amount(result.get("paid", amount))

    phone = (
        str(customer.get("phone") or "")
        or str(result.get("phone") or "")
        or str(body.get("phone") or "")
    ).strip()
    if not phone:
        try:
            for c in tezpos_api.get_customers(token, server) or []:
                if isinstance(c, dict) and str(c.get("id") or "") == customer_id:
                    phone = str(c.get("phone") or "").strip()
                    if not customer.get("name") and c.get("name"):
                        customer["name"] = c.get("name")
                    break
        except Exception:
            pass

    sms_sent = bool(result.get("sms_sent"))
    if not sms_sent and isinstance(result.get("sms"), dict):
        sms_sent = bool(result["sms"].get("ok") or result["sms"].get("sent"))
    elif not sms_sent and result.get("sms") is True:
        sms_sent = True

    sms_error = ""
    # TezPOS DevSMS — to'g'ridan-to'g'ri (backend SMS ishonchsiz bo'lishi mumkin)
    from . import devsms

    store = str(
        result.get("store_name")
        or request.session.get(SESSION_DISPLAY)
        or server
        or "TezPOS"
    )
    cashier = str(result.get("cashier") or result.get("cashier_name") or "admin")
    sms_text = devsms.build_debt_message(
        shop=cashier,
        branch=store if store and store != cashier else "",
        debt_amount=-abs(paid),
        balance=balance,
        check_link=sms_check_url,
    )
    sms_res = devsms.send_dev_sms(phone=phone, message=sms_text)
    sms_sent = bool(sms_res.get("ok"))
    if not sms_sent:
        sms_error = str(sms_res.get("error") or "SMS yuborilmadi (DevSMS).")

    return JsonResponse(
        {
            "ok": True,
            "paid": float(paid),
            "paid_display": f"{paid:,.0f}".replace(",", " "),
            "balance": float(balance),
            "balance_display": f"{balance:,.0f}".replace(",", " "),
            "receipt_number": result.get("receipt_number"),
            "payment_type": result.get("payment_type") or payment_type,
            "check_url": check_url,
            "sms_sent": sms_sent,
            "sms_error": sms_error,
            "sms_phone": phone,
            "sms_preview": sms_text,
            "customer": {
                "id": str(customer.get("id") or customer_id),
                "name": customer.get("name") or "",
                "phone": phone,
                "debt": float(balance),
            },
        }
    )
