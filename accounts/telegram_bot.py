"""Telegram Bot API — smena xabarlari va Excel yuborish."""
from __future__ import annotations

import json
import logging
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from io import BytesIO
from typing import Any

logger = logging.getLogger("tezpos.telegram")


TG_API = "https://api.telegram.org"


def parse_recipient_line(raw: str) -> str | None:
    """Qatorni Telegram chat_id, @username yoki invite havolasiga aylantiradi."""
    s = (raw or "").strip()
    if not s or s.startswith("#"):
        return None
    s = s.replace("https://", "").replace("http://", "")
    if s.startswith("t.me/"):
        s = s[5:]
    if s.startswith("telegram.me/"):
        s = s[12:]
    s = s.strip().strip("/")

    # Private channel/supergroup: t.me/c/1234567890/11 → -1001234567890
    m = re.match(r"^c/(\d+)(?:/\d+)?$", s)
    if m:
        return f"-100{m.group(1)}"

    # Invite: t.me/+HASH yoki t.me/joinchat/HASH
    if s.startswith("joinchat/"):
        h = s.split("/", 1)[1].split("/")[0].strip()
        return f"invite:{h}" if h else None
    if s.startswith("+") and not re.fullmatch(r"\+\d{5,}", s):
        h = s[1:].split("/")[0].strip()
        return f"invite:{h}" if len(h) >= 6 else None

    if s.startswith("@"):
        return s
    # -100... guruh/kanal ID yoki shaxsiy raqamli ID
    if re.fullmatch(r"-100\d{5,}", s) or re.fullmatch(r"-?\d{5,}", s):
        return s
    # Oddiy username (kanal/guruh public)
    if re.fullmatch(r"[A-Za-z0-9_]{4,}", s):
        return f"@{s}"
    # t.me/name/msg
    m2 = re.fullmatch(r"([A-Za-z0-9_]{4,})(?:/\d+)?", s)
    if m2:
        return f"@{m2.group(1)}"
    return None


_RECIPIENT_TOKEN_RE = re.compile(
    r"https?://(?:t\.me|telegram\.me)/[^\s,;]+"
    r"|t\.me/[^\s,;]+"
    r"|@[A-Za-z0-9_]{4,}"
    r"|-100\d{5,}"
    r"|-\d{6,}"
    r"|\+[A-Za-z0-9_-]{8,}"
    r"|(?<![A-Za-z0-9_-])\d{6,}"
)


def _split_recipient_tokens(line: str) -> list[str]:
    s = (line or "").strip()
    if not s or s.startswith("#"):
        return []
    found = [m.group(0).rstrip(".,;") for m in _RECIPIENT_TOKEN_RE.finditer(s)]
    return found if found else [s]


def parse_recipients(text: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for line in (text or "").replace(",", "\n").splitlines():
        for token in _split_recipient_tokens(line):
            chat = parse_recipient_line(token)
            if not chat or chat in seen:
                continue
            seen.add(chat)
            out.append(chat)
    return out


def _api_call(token: str, method: str, payload: dict | None = None, timeout: float = 12) -> dict:
    url = f"{TG_API}/bot{token}/{method}"
    data = None
    headers = {"Content-Type": "application/json"}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST" if data else "GET")
    last_err = "telegram request failed"
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                body = json.loads(resp.read().decode("utf-8") or "{}")
            return body if isinstance(body, dict) else {"ok": False, "description": "bad response"}
        except urllib.error.HTTPError as exc:
            err_body = exc.read().decode("utf-8", errors="replace")
            try:
                body = json.loads(err_body)
            except json.JSONDecodeError:
                body = {"ok": False, "description": err_body or str(exc)}
            logger.warning("telegram %s HTTP %s: %s", method, exc.code, body.get("description"))
            return body if isinstance(body, dict) else {"ok": False, "description": str(exc)}
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_err = str(exc)
            logger.warning("telegram %s attempt %s failed: %s", method, attempt + 1, exc)
            if attempt < 2:
                time.sleep(0.4 * (2 ** attempt))
    logger.error("telegram %s failed after retries: %s", method, last_err)
    return {"ok": False, "description": last_err}


def get_me(token: str) -> dict:
    return _api_call(token, "getMe")


def get_updates(token: str, limit: int = 50) -> dict:
    return _api_call(token, "getUpdates", {"offset": -limit, "limit": limit, "timeout": 0})


def list_recent_chats(token: str) -> list[dict[str, Any]]:
    """Botga yozgan foydalanuvchi/guruh/kanallar (getUpdates)."""
    data = get_updates(token, limit=80)
    if not data.get("ok"):
        return []
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    for upd in reversed(data.get("result") or []):
        if not isinstance(upd, dict):
            continue
        msg = (
            upd.get("message")
            or upd.get("edited_message")
            or upd.get("channel_post")
            or upd.get("my_chat_member")
            or {}
        )
        chat = msg.get("chat") if isinstance(msg, dict) else None
        if not isinstance(chat, dict):
            continue
        cid = str(chat.get("id") or "")
        if not cid or cid in seen:
            continue
        seen.add(cid)
        uname = (chat.get("username") or "").strip()
        title = (chat.get("title") or "").strip()
        first = (chat.get("first_name") or "").strip()
        last = (chat.get("last_name") or "").strip()
        name = title or " ".join(x for x in (first, last) if x) or uname or cid
        out.append(
            {
                "id": cid,
                "username": f"@{uname}" if uname else "",
                "name": name,
                "type": chat.get("type") or "",
            }
        )
        if len(out) >= 25:
            break
    return out


def resolve_chat_id(token: str, recipient: str) -> tuple[str, str | None]:
    """
    @username / invite / raqamli ID ni yuborish uchun chat_id ga aylantiradi.
    Guruh invite: bot guruhda bo‘lishi kerak (getUpdates orqali topiladi).
    """
    r = (recipient or "").strip()
    if not r:
        return r, "bo‘sh manzil"
    if r.startswith("invite:"):
        chats = list_recent_chats(token)
        groups = [
            c
            for c in chats
            if (c.get("type") or "") in ("group", "supergroup", "channel")
        ]
        if len(groups) == 1:
            return str(groups[0]["id"]), None
        if groups:
            # Bir nechta guruh — oxirgi (bot qo‘shilgan) ni sinab ko‘ramiz
            return str(groups[0]["id"]), None
        return r, (
            "invite havola: botni guruhga qo‘shing (admin), "
            "keyin «Chatlarni topish» yoki -100… ID ni yozing"
        )
    if re.fullmatch(r"-?\d+", r):
        return r, None

    uname = r[1:] if r.startswith("@") else r
    uname_l = uname.lower()
    for chat in list_recent_chats(token):
        cu = (chat.get("username") or "").lstrip("@").lower()
        if cu and cu == uname_l:
            return str(chat["id"]), None
        if str(chat.get("id")) == r:
            return str(chat["id"]), None
    probe = _api_call(token, "getChat", {"chat_id": r if r.startswith("@") else f"@{uname}"})
    if probe.get("ok") and isinstance(probe.get("result"), dict):
        cid = probe["result"].get("id")
        if cid is not None:
            return str(cid), None
    return r, None


def explain_tg_error(description: str, recipient: str) -> str:
    d = (description or "").lower()
    if "chat not found" in d:
        return (
            f"{recipient}: chat topilmadi. "
            "Shaxsiy chat uchun avval botga kirib /start bosing, "
            "keyin «Chatlarni topish» orqali raqamli ID ni qo‘ying. "
            "Guruh/kanal: botni qo‘shing, admin qiling, "
            "so‘ng -100… ID yoki t.me/+invite havolasini yozing."
        )
    if "blocked" in d or "deactivated" in d:
        return f"{recipient}: foydalanuvchi botni bloklagan yoki o‘chirilgan."
    if "not enough rights" in d or "need administrator" in d:
        return f"{recipient}: botga yozish huquqi yo‘q — admin qiling."
    if "chat_id is empty" in d:
        return f"{recipient}: chat ID bo‘sh."
    return f"{recipient}: {description or 'xato'}"


def send_message(token: str, chat_id: str, text: str, parse_mode: str = "HTML") -> dict:
    resolved, _hint = resolve_chat_id(token, chat_id)
    return _api_call(
        token,
        "sendMessage",
        {
            "chat_id": resolved,
            "text": text[:4000],
            "parse_mode": parse_mode,
            "disable_web_page_preview": True,
        },
    )


def send_document(
    token: str,
    chat_id: str,
    filename: str,
    file_bytes: bytes,
    caption: str = "",
) -> dict:
    """multipart/form-data orqali Excel yuborish."""
    resolved, _ = resolve_chat_id(token, chat_id)
    chat_id = resolved
    boundary = "----TezPosBoundary7MA4YWxkTrZu0gW"
    parts: list[bytes] = []

    def add_field(name: str, value: str) -> None:
        parts.append(f"--{boundary}\r\n".encode())
        parts.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        parts.append(f"{value}\r\n".encode())

    add_field("chat_id", str(chat_id))
    if caption:
        add_field("caption", caption[:1000])
        add_field("parse_mode", "HTML")

    parts.append(f"--{boundary}\r\n".encode())
    parts.append(
        (
            f'Content-Disposition: form-data; name="document"; filename="{filename}"\r\n'
            f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode()
    )
    parts.append(file_bytes)
    parts.append(b"\r\n")
    parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(parts)

    url = f"{TG_API}/bot{token}/sendDocument"
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    last_err = "sendDocument failed"
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                return json.loads(resp.read().decode("utf-8") or "{}")
        except urllib.error.HTTPError as exc:
            err_body = exc.read().decode("utf-8", errors="replace")
            try:
                return json.loads(err_body)
            except json.JSONDecodeError:
                return {"ok": False, "description": err_body or str(exc)}
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_err = str(exc)
            logger.warning("sendDocument attempt %s failed: %s", attempt + 1, exc)
            if attempt < 2:
                time.sleep(0.4 * (2 ** attempt))
    logger.error("sendDocument failed after retries: %s", last_err)
    return {"ok": False, "description": last_err}


def broadcast_message(token: str, recipients: list[str], text: str) -> list[dict[str, Any]]:
    results = []
    for chat in recipients:
        resolved, hint = resolve_chat_id(token, chat)
        if hint and str(resolved).startswith("invite:"):
            results.append(
                {
                    "chat": chat,
                    "resolved": resolved,
                    "ok": False,
                    "error": hint,
                    "raw": {},
                }
            )
            continue
        res = send_message(token, resolved, text)
        ok = bool(res.get("ok"))
        desc = res.get("description") or ""
        results.append(
            {
                "chat": chat,
                "resolved": resolved,
                "ok": ok,
                "error": "" if ok else explain_tg_error(desc, chat),
                "raw": res,
            }
        )
    return results


def broadcast_document(
    token: str,
    recipients: list[str],
    filename: str,
    file_bytes: bytes,
    caption: str = "",
) -> list[dict[str, Any]]:
    results = []
    for chat in recipients:
        resolved, hint = resolve_chat_id(token, chat)
        if hint and str(resolved).startswith("invite:"):
            results.append(
                {
                    "chat": chat,
                    "resolved": resolved,
                    "ok": False,
                    "error": hint,
                    "raw": {},
                }
            )
            continue
        res = send_document(token, resolved, filename, file_bytes, caption=caption)
        ok = bool(res.get("ok"))
        desc = res.get("description") or ""
        results.append(
            {
                "chat": chat,
                "resolved": resolved,
                "ok": ok,
                "error": "" if ok else explain_tg_error(desc, chat),
                "raw": res,
            }
        )
    return results


def build_shift_excel(
    *,
    business_name: str,
    shift: dict,
    sales_rows: list[dict],
    price_lists: list[dict],
    credit_rows: list[dict],
    debtors_rows: list[dict] | None = None,
    low_stock_rows: list[dict] | None = None,
    sold_product_rows: list[dict] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Xulosa"
    summary = [
        ("Biznes", business_name),
        ("Smena", shift.get("status_label") or shift.get("status") or ""),
        ("Kassir", shift.get("cashier") or ""),
        ("Ochilgan", shift.get("opened_at_display") or shift.get("opened_at") or ""),
        ("Yopilgan", shift.get("closed_at_display") or shift.get("closed_at") or "—"),
        ("Davomiylik", shift.get("duration_label") or ""),
        ("Cheklar", shift.get("checks") or 0),
        ("Jami savdo", shift.get("gross") or 0),
        ("Sof foyda", shift.get("profit") or 0),
        ("Marja %", shift.get("margin") or 0),
        ("Sotuv narxida", shift.get("selling_revenue") or 0),
        ("Narxlar ro‘yxatida (optom)", shift.get("wholesale_revenue") or 0),
        ("Qarzga savdo (smena)", shift.get("credit_total") or 0),
        ("Qarzdorlar (jami)", shift.get("debtors_count") or 0),
        ("Qarzdorlar summasi", shift.get("debtors_total") or 0),
        ("Kam qoldiq (tovar)", shift.get("low_stock_count") or 0),
    ]
    money_keys = {
        "Jami savdo",
        "Sof foyda",
        "Sotuv narxida",
        "Narxlar ro‘yxatida (optom)",
        "Qarzga savdo (smena)",
        "Qarzdorlar summasi",
    }
    ws.append(["Ko‘rsatkich", "Qiymat"])
    for k, v in summary:
        if k in money_keys:
            ws.append([k, format_money_som(v)])
        elif k == "Marja %":
            ws.append([k, format_margin(v)])
        else:
            ws.append([k, v])

    ws2 = wb.create_sheet("Kunlik sotuv")
    ws2.append(
        [
            "Chek raqami",
            "Vaqt",
            "Mijoz",
            "Mahsulotlar",
            "To‘lov",
            "Summa",
            "Foyda",
        ]
    )
    wrap_top = Alignment(wrap_text=True, vertical="top")
    for row in sales_rows:
        products_text = (row.get("products_text") or "").strip()
        line_count = products_text.count("\n") + 1 if products_text else 1
        ws2.append(
            [
                row.get("receipt_no") or row.get("receipt_number") or "",
                row.get("time") or "",
                row.get("customer") or "",
                products_text,
                row.get("payment") or "",
                format_money_som(row.get("total") or 0),
                format_money_som(row.get("profit") or 0),
            ]
        )
        r_idx = ws2.max_row
        ws2.cell(r_idx, 4).alignment = wrap_top
        if line_count > 1:
            ws2.row_dimensions[r_idx].height = min(180, 15 * line_count)

    ws_sold = wb.create_sheet("Sotilgan mahsulotlar")
    ws_sold.append(["#", "Mahsulot nomi", "Shtrixkod", "Miqdor", "Birlik", "Tushum", "Foyda"])
    sold_i = 0
    for row in sold_product_rows or []:
        qty = float(row.get("qty") or 0)
        if qty <= 0:
            continue
        sold_i += 1
        ws_sold.append(
            [
                sold_i,
                row.get("name") or "",
                row.get("barcode") or "",
                format_qty(qty),
                row.get("unit") or "dona",
                format_money_som(row.get("revenue") or 0),
                format_money_som(row.get("profit") or 0),
            ]
        )

    ws3 = wb.create_sheet("Narxlar")
    ws3.append(["Ro‘yxat", "Cheklar", "Tushum", "Foyda", "Ulush %"])
    for row in price_lists:
        ws3.append(
            [
                row.get("name") or "",
                row.get("checks") or 0,
                format_money_som(row.get("revenue") or 0),
                format_money_som(row.get("profit") or 0),
                format_margin(row.get("share") or 0),
            ]
        )

    ws4 = wb.create_sheet("Smena qarzlari")
    ws4.append(["Mijoz", "Cheklar", "Qarz summa"])
    for row in credit_rows:
        ws4.append(
            [
                row.get("customer") or "",
                row.get("orders") or 0,
                format_money_som(row.get("total") or 0),
            ]
        )

    ws5 = wb.create_sheet("Qarzdorlar")
    ws5.append(["#", "Mijoz", "Telefon", "Qarz"])
    for i, row in enumerate(debtors_rows or [], start=1):
        ws5.append(
            [
                i,
                row.get("name") or row.get("customer") or "",
                row.get("phone") or "",
                format_money_som(row.get("debt") or row.get("total") or 0),
            ]
        )

    ws6 = wb.create_sheet("Kam qoldiq")
    ws6.append(["#", "Mahsulot nomi", "Shtrixkod", "Qoldiq", "Minimal", "Birlik"])
    for i, row in enumerate(low_stock_rows or [], start=1):
        ws6.append(
            [
                i,
                row.get("name") or "",
                row.get("barcode") or "",
                row.get("stock") if row.get("stock") is not None else row.get("qty") or 0,
                row.get("min_stock") if row.get("min_stock") is not None else "",
                row.get("unit") or "dona",
            ]
        )

    wrap = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True)
    # Ustun indekslari (1-based): nom kesilmasin
    name_cols_by_sheet = {
        "Xulosa": {2},
        "Kunlik sotuv": {3, 4},
        "Sotilgan mahsulotlar": {2},
        "Narxlar": {1},
        "Smena qarzlari": {1},
        "Qarzdorlar": {2},
        "Kam qoldiq": {2},
    }

    for sheet in wb.worksheets:
        name_cols = name_cols_by_sheet.get(sheet.title, set())
        for cell in sheet[1]:
            cell.font = header_font
        for col in sheet.columns:
            col_idx = col[0].column
            letter = get_column_letter(col_idx)
            width = 14
            for cell in col:
                val = str(cell.value if cell.value is not None else "")
                width = max(width, len(val) + 2)
                if col_idx in name_cols:
                    cell.alignment = wrap
            # Mahsulot/mijoz nomlari to‘liq ko‘rinsin
            if col_idx in name_cols:
                sheet.column_dimensions[letter].width = min(90, max(28, width))
            else:
                sheet.column_dimensions[letter].width = min(36, max(10, width))
        if sheet.title == "Kunlik sotuv":
            sheet.column_dimensions["A"].width = 14
            sheet.column_dimensions["D"].width = 55

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def format_qty(n: float | int) -> str:
    try:
        v = float(n or 0)
    except (TypeError, ValueError):
        v = 0.0
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    text = f"{v:.3f}".rstrip("0").rstrip(".")
    return text or "0"


def format_sale_product_line(
    name: str,
    qty: float | int,
    unit: str,
    unit_price: float | int,
    line_total: float | int | None = None,
) -> str:
    """bon saryog 10 шт x 33.000 = 330.000"""
    q = float(qty or 0)
    price = float(unit_price or 0)
    total = float(line_total) if line_total is not None else q * price
    unit_s = (unit or "dona").strip() or "dona"
    label = (name or "Mahsulot").strip() or "Mahsulot"
    return f"{label} {format_qty(q)} {unit_s} x {format_money(price)} = {format_money(total)}"


def format_money(n: float | int) -> str:
    """1.000 / 10.000 / 100.000 / 1.000.000"""
    try:
        v = float(n or 0)
    except (TypeError, ValueError):
        v = 0.0
    sign = "-" if v < 0 else ""
    whole = int(abs(round(v)))
    grouped = f"{whole:,}".replace(",", ".")
    return f"{sign}{grouped}"


def format_money_som(n: float | int) -> str:
    return f"{format_money(n)} so'm"


def format_margin(n: float | int) -> str:
    try:
        v = float(n or 0)
    except (TypeError, ValueError):
        v = 0.0
    # 41,0% — o‘zbekcha vergul
    return f"{v:.1f}".replace(".", ",")


def format_duration(opened_iso: str, closed_iso: str) -> str:
    """Smena davomiyligi: '2 soat 51 daqiqa'."""
    from datetime import datetime

    def _parse(raw: str):
        s = (raw or "").strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s.replace("Z", "+00:00"))
        except ValueError:
            return None

    a = _parse(opened_iso)
    b = _parse(closed_iso)
    if not a or not b:
        return ""
    secs = int(abs((b - a).total_seconds()))
    hours, rem = divmod(secs, 3600)
    mins = rem // 60
    if hours and mins:
        return f"{hours} soat {mins} daqiqa"
    if hours:
        return f"{hours} soat"
    return f"{mins} daqiqa"


def build_shift_message(
    *,
    business_name: str,
    event: str,
    shift: dict,
) -> str:
    """Telegram smena xabari — foydalanuvchi namunasi bilan bir xil ko‘rinish."""
    status = "ochildi" if event == "open" else "yopildi"
    emoji = "🟢" if event == "open" else "🔴"
    opened = (
        shift.get("opened_at_display")
        or shift.get("opened_display")
        or shift.get("opened_at")
        or "—"
    )
    closed = (
        shift.get("closed_at_display")
        or shift.get("closed_display")
        or shift.get("closed_at")
        or "—"
    )
    duration = shift.get("duration_label") or ""
    if not duration and event == "close":
        duration = format_duration(
            str(shift.get("opened_at") or ""),
            str(shift.get("closed_at") or ""),
        )

    lines = [
        f"{emoji} <b>{business_name}</b>",
        f"<b>Smena {status}</b>",
        "",
        f"👤 Kassir: <b>{shift.get('cashier') or '—'}</b>",
    ]
    if event == "close":
        lines.append(f"🕐 {opened} → {closed}")
        if duration:
            lines.append(f"⏱️ Smena davomiyligi: <b>{duration}</b>")
    else:
        lines.append(f"🕐 {opened}")

    lines += [
        "",
        "━━━━━━━━━━━━━━━━━━",
        "",
        f"🧾 <b>Cheklar:</b> {int(shift.get('checks') or 0)} ta",
        f"💰 <b>Jami savdo:</b> {format_money(shift.get('gross') or 0)} so‘m",
        f"📈 <b>Sof foyda:</b> {format_money(shift.get('profit') or 0)} so‘m",
        f"📊 <b>Marja:</b> {format_margin(shift.get('margin') or 0)}%",
        "",
        "━━━━━━━━━━━━━━━━━━",
        "",
        f"💵 <b>Savdo:</b> {format_money(shift.get('selling_revenue') or shift.get('gross') or 0)} so‘m",
        f"📦 <b>Optom savdo:</b> {format_money(shift.get('wholesale_revenue') or 0)} so‘m",
        f"💳 <b>Smenadagi qarz:</b> {format_money(shift.get('credit_total') or 0)} so‘m",
    ]

    if event == "close":
        lines += [
            "",
            "━━━━━━━━━━━━━━━━━━",
            "",
            f"👥 <b>Qarzdorlar:</b> {int(shift.get('debtors_count') or 0)} ta",
            f"💰 Jami qarz: <b>{format_money(shift.get('debtors_total') or 0)} so‘m</b>",
            "",
            f"⚠️ <b>Kam qoldiq:</b> {int(shift.get('low_stock_count') or 0)} ta mahsulot",
            "",
            "━━━━━━━━━━━━━━━━━━",
            "",
            "📎 <b>Excel hisobot tayyor</b>",
            "• Kunlik sotuvlar",
            "• Sotilgan mahsulotlar",
            "• Qarzdorlar ro‘yxati",
            "• Kam qoldiq mahsulotlar",
        ]

        debtors = shift.get("debtors") or []
        if debtors:
            lines += ["", "📋 <b>Qarzdorlar</b>"]
            for c in debtors[:15]:
                name = c.get("name") or c.get("customer") or "—"
                lines.append(
                    f"• {name} — <b>{format_money(c.get('debt') or c.get('total') or 0)} so‘m</b>"
                )
            if len(debtors) > 15:
                lines.append(f"… yana {len(debtors) - 15} ta (Excelda)")

        low_count = int(shift.get("low_stock_count") or 0)
        if low_count:
            lines += [
                "",
                "⚠️ <b>Kam qoldiq mahsulotlar</b>",
                f"{low_count} ta mahsulot Excel faylida batafsil ko‘rsatilgan.",
            ]
    else:
        credits = shift.get("credit_customers") or []
        if credits:
            lines += ["", "<b>Qarzga sotuv:</b>"]
            for c in credits[:12]:
                lines.append(
                    f"• {c.get('customer')}: {format_money(c.get('total') or 0)} so‘m"
                )

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3980] + "\n…"
    return text
